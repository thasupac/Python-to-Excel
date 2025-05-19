from openpyxl import load_workbook
# from tkinter import *
from tkinter import ttk, Tk
# from excel_function_tool import *
from tkinter import filedialog, messagebox, Toplevel, END, LabelFrame, Label, StringVar, Text
from PIL import ImageTk
from datetime import datetime
from excel_function_tool import ToolReg
# from send_mail import *
import PIL.Image
import configparser
import tkinter as tk
exT = ToolReg()




###---Font
FONT1 = ('Angsana New',25,'bold')
FONT2 = ('Angsana New',18)
FONT3 = ('Angsana New',12)

### technician
techname = ['Thanongsak Su','Khunakorn R','Pichet T','Wasan R',
            'Don P','Somchai L','Adirek C','Sangworn D',
            'Pratchaya S','Supot P','Kriangsak H','Anan C',
            'Thanatorn K','Anong J','Thanongsak D','Apicha K','Sompong L']

class ToolCheck:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.toolpath = self.config['DATABASE']['dbtoolpath']
        self.toolregsheet = self.config['DATABASE']['regtoolsheet']
        self.toolphotopath = self.config['DATABASE']['dbphototoolpath']

        from excel_function_tool import ExcelF, ExcelB, stenCil
        self.exF = ExcelF()
        self.exB = ExcelB()
        self.exSC = stenCil()

        from send_mail import SendMail
        self.songmail = SendMail()
    
    #reg check
    # tsid,model,modelNum,customer,clsaaTypes,qty,unit,line,regBy,desc,check_photo_save
    def reg_check(self,tsid,model,modelNum,modelNum_for_check,customer,clsaaTypes,clsaaTypes_check,qty,unit,line,regBy,desc,check_photo_save):

        self.wb = load_workbook(self.toolpath)
        self.wb.active = self.wb[self.toolregsheet]
        self.sheet = self.wb[self.toolregsheet]
        rows = self.sheet.max_row

        data = ['CHECK','']
        for i in range(1,rows+1):
            model_form = (self.sheet.cell(row=i,column=3).value).lower().replace(' ','').strip()
            types_form = (self.sheet.cell(row=i,column=5).value).lower()
            if model_form == modelNum_for_check and  types_form == clsaaTypes_check:

                tsidf = self.sheet.cell(row=i,column=1).value
                modelf = self.sheet.cell(row=i,column=2).value
                modelNumf = self.sheet.cell(row=i,column=3).value
                customerf = self.sheet.cell(row=i,column=4).value
                clsaaTypesf = self.sheet.cell(row=i,column=5).value
                qtyf = self.sheet.cell(row=i,column=6).value
                unitf = self.sheet.cell(row=i,column=7).value
                linef = self.sheet.cell(row=i,column=8).value
                regByf = self.sheet.cell(row=i,column=9).value
                regdatef = self.sheet.cell(row=i,column=10).value
                descf = self.sheet.cell(row=i,column=11).value

                # print(regByf)
                # print(regdatef)

                data.remove('')
                data.append('NOTFOUND')

                break
        
        if data[1] != '':
            messagebox.showerror('Tooling', 'มี Tool นี้อยู่แล้วให้ทำการอัปเดต')
            # print(tsidf)

            toolViews = Toplevel()
            toolViews.geometry('800x500-50+50')
            toolViews.title('Tooling management')
            L = Label(toolViews, text='มี Tool นี้อยู่แล้วให้ทำการอัปเดต', font=FONT1)
            L.pack()

            #function update
            def upDatetool():
                U_model = v_model.get()
                U_modelNum = v_modelNum.get()
                U_customer = v_customer.get()
                U_types = v_types.get()
                U_qty = v_qty.get()
                U_unit = v_unit.get()
                U_line = v_line.get()
                U_desc = v_desc.get() 
                U_regby = v_regBy.get()
                U_regdate = v_regDate.get()
                file_path_add = v_photo_path.get()
                
                check = messagebox.askyesno('ยืนยันการอัปเดต','คุณต้องการอัปเดตข้อมูลใช่หรือไม่?')
                        
                if check == True:
                    exT.StartToolUpdatE(tsidf,U_model,U_modelNum,U_customer,U_types,U_qty,U_unit,U_line,U_desc,file_path_add)
                    closegui()

            def closegui():
                toolViews.destroy()
                toolViews.update()

            #main frame
            MF = LabelFrame(toolViews)
            MF.pack(padx=10, pady=5)

            #general frame
            GD = LabelFrame(MF)
            GD.grid(row=0, column=0, padx=10, pady=10)

            #tsid
            L = Label(GD, text=f'TSID : {tsidf}', font=FONT3)
            L.grid(row=0, column=0)

            #model
            L = Label(GD, text='Model :', font=FONT2)
            L.grid(row=1, column=0, padx=20, pady=5, sticky='w')
            v_model = StringVar()
            v_model.set(modelf)
            E = ttk.Entry(GD, textvariable=v_model, font=FONT2)
            E.grid(row=1, column=1, padx=20, pady=5)

            #model num
            L = Label(GD, text='Model number :', font=FONT2)
            L.grid(row=2, column=0, padx=20, pady=5, sticky='w')
            v_modelNum = StringVar()
            v_modelNum.set(modelNumf)
            E = ttk.Entry(GD, textvariable=v_modelNum, font=FONT2)
            E.grid(row=2, column=1, padx=20, pady=5)

            #customer
            L = Label(GD, text='Customer :', font=FONT2)
            L.grid(row=3, column=0, padx=20, pady=5, sticky='w')
            v_customer = StringVar()
            v_customer.set(customerf)
            E = ttk.Entry(GD, textvariable=v_customer, font=FONT2)
            E.grid(row=3, column=1, padx=20, pady=5)

            #types
            L = Label(GD, text='Types :', font=FONT2)
            L.grid(row=4, column=0, padx=20, pady=5, sticky='w')
            v_types = StringVar()
            v_types.set(clsaaTypesf)
            E = ttk.Entry(GD, textvariable=v_types, font=FONT2)
            E.grid(row=4, column=1, padx=20, pady=5)

            #qty
            L = Label(GD, text="Q'ty :", font=FONT2)
            L.grid(row=5, column=0, padx=20, pady=5)
            v_qty = StringVar()
            v_qty.set(qtyf)
            E = ttk.Entry(GD, textvariable=v_qty, font=FONT2)
            E.grid(row=5, column=1, padx=20, pady=5)

            #unit
            L = Label(GD, text='Unit :', font=FONT2)
            L.grid(row=6, column=0, padx=20, pady=5)
            v_unit = StringVar()
            v_unit.set(unitf)
            E = ttk.Entry(GD, textvariable=v_unit, font=FONT2)
            E.grid(row=6, column=1, padx=20, pady=5)

            #line
            L = Label(GD, text='Line :', font=FONT2)
            L.grid(row=7, column=0, padx=20, pady=5)
            v_line = StringVar()
            v_line.set(linef)
            E = ttk.Entry(GD, textvariable=v_line, font=FONT2)
            E.grid(row=7, column=1, padx=20, pady=5)

            #regis by
            L = Label(GD, text='Reg by:', font=FONT2)
            L.grid(row=8, column=0, padx=20, pady=5)
            v_regBy = StringVar()
            v_regBy.set(regByf)
            E = ttk.Entry(GD, textvariable=v_regBy, font=FONT2)
            E.grid(row=8, column=1, padx=20, pady=5)
            
            #reg date
            L = Label(GD, text='Reg date :', font=FONT2)
            L.grid(row=9, column=0, padx=20, pady=10)
            v_regDate = StringVar()
            v_regDate.set(regdatef)
            E = ttk.Entry(GD, textvariable=v_regDate, font=FONT2)
            E.grid(row=9, column=1, padx=20, pady=5)

            #ldescriptions
            L = Label(GD, text='Descriptions :', font=FONT2)
            L.grid(row=10, column=0, padx=20, pady=5)
            v_desc = StringVar()
            v_desc.set(descf)
            E = ttk.Entry(GD, textvariable=v_desc, font=FONT2)
            E.grid(row=10, column=1, padx=20, pady=5)

            #frame photo
            F = LabelFrame(MF, text='Photo', font=FONT3, width=350, height=350)
            F.grid(row=0, column=1, padx=10, pady=10)

            #buttom select photo
            def add_photo():
                file_path_add = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                v_photo_path.set(file_path_add)
                try:
                    photo_mc = tk.Label(F)
                    photo_mc.pack()
                    image = PIL.Image.open(file_path_add)
                    image = image.resize((350, 350))
                    photo = ImageTk.PhotoImage(image)
                    photo_mc.config(image=photo)
                    photo_mc.image = photo
                except Exception as e:
                    messagebox.showerror('Register photo', f'Failed to load photo {e}')

            b_select = ttk.Button(MF, text='Add photo', command=add_photo)
            b_select.grid(row=0, column=2, padx=10)

            #photo path
            v_photo_path = StringVar()
            L_path = ttk.Entry(MF, textvariable=v_photo_path)
            L_path.grid(row=0, column=4, padx=10)

            #try to previews
            try:
                photo_mc = tk.Label(F)
                photo_mc.pack()
                image = PIL.Image.open(f"{self.toolphotopath}\\{tsidf}.png")
                image = image.resize((350, 350))
                photo = ImageTk.PhotoImage(image)
                photo_mc.config(image=photo)
                photo_mc.image = photo

            except Exception as e:
                L = Label(F, text='ไม่มีรูปภาพสำหรับ TSID นี้')
                L.pack()
                ()

            #button update
            B = ttk.Button(MF, text='Update', command=upDatetool)
            B.grid(row=0, column=3, padx=10)

        if data[1] != 'NOTFOUND':
            exT.StartToolReG(tsid,model,modelNum,customer,clsaaTypes,qty,unit,line,regBy,desc,check_photo_save)
            
    def fixTure(self,tsid,fixid,side,line,customer,qty,by,fixid_for_check,comment,photofix): ##tsid,fixid,side,line,customer,qty,by,fixid_for_chec

        self.fixid = fixid
        self.side = side
        self.line = line
        self.customer = customer
        self.qty = qty
        self.by = by
        self.comment = comment
        self.photofix = photofix

        self.wb = load_workbook(self.toolpath)
        self.wb.active = self.wb[self.toolregsheet]
        self.sheet = self.wb[self.toolregsheet]
        rows = self.sheet.max_row
        
        data = ['CHECK','']
        for i in range(1,rows+1):
            if (self.sheet.cell(row=i,column=3).value).lower().replace(' ','').strip() == fixid_for_check:
                data.remove('')
                data.append('NOTFOUND')
                self.exF.StartToolWithdrawFixture(tsid,fixid,side,line,customer,qty,by,comment,photofix)  #tsid,fixid,side,line,customer,qty,by,fixid_for_check
                break

        if data[1] != 'NOTFOUND':
            
            try:
                check = messagebox.askyesno('Fixture register','หมายเลขนี้ยังไม่มีในฐานข้อมูลคุณต้องบันทึกลงในฐานข้อมูล!\n คลิ๊ก Yes เพื่อบันทึกข้อมูลและบันทึกการยืม')
                # messagebox.showwarning("showwarning", "Warning") 
                if check:
                    def center_windows(w,h):
                        ws = REGTOOLCHECK.winfo_screenwidth() #screen width
                        hs = REGTOOLCHECK.winfo_screenheight() #screen height
                        x = (ws/2) - (w/2)
                        y = (hs/2) - (h/2)
                        return f'{w}x{h}+{x:.0f}+{y:.0f}'

                    REGTOOLCHECK = Toplevel()
                    REGTOOLCHECK.title('Fixture register')
                    win_size = center_windows(1000,600)
                    REGTOOLCHECK.geometry(win_size)
                    
                    #save reg
                    def saveToolReg(event=None):
                        tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
                        model = v_model.get()
                        modelNum = v_model_number.get()
                        customer = v_customer.get()
                        clsaaTypes = v_types.get()
                        qtyr = v_qty.get()
                        unit = v_unit.get()
                        line = v_line.get()
                        desc = E_desc.get('1.0',END).strip()
                        regBy = v_reg_by.get()
                        check_photo_save = v_photo_path.get()
                        
                        if model and modelNum and customer and clsaaTypes and qtyr and unit and line and regBy:

                            try:
                                self.exTR.StartToolReG(tsid,model,modelNum,customer,clsaaTypes,qtyr,unit,line,regBy,desc,check_photo_save)
                                self.exF.StartToolWithdrawFixture(tsid,self.fixid,self.side,self.line,self.customer,self.qty,self.by,self.comment,self.photofix) #tsid,fixid,side,line,customer,qty,by,fixid_for_check
                                reset()
                                CloSE()
                            except Exception as e:
                                messagebox.showerror("Error", f"Failed to data: {e}")
                            
                        else:
                            messagebox.showinfo('Fixture register', 'โปรดกรอกข้อมูลให้ครบถ้วน')

                    def CloSE():
                        REGTOOLCHECK.destroy()
                        REGTOOLCHECK.update()

                    def reset():
                        v_model.set('')
                        v_model_number.set('')
                        v_customer.set('')
                        v_types.set('')
                        v_qty.set('')
                        v_unit.set('')
                        v_line.set('')
                        E_desc.delete('1.0',END)
                        v_reg_by.set('')
                        v_photo_path.set('')
                        photofix.config(image=None)
                        photofix.image = None

                    #select photo
                    def selectPhoto():
                        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                        v_photo_path.set(file_path)

                        ##
                        try:
                            image = PIL.Image.open(v_photo_path.get())
                            image = image.resize((190, 140))
                            photo = ImageTk.PhotoImage(image)
                            photofix.config(image=photo)
                            photofix.image = photo

                        except Exception as e:
                            ()

                    #registertool
                    MF = LabelFrame(REGTOOLCHECK)
                    MF.pack(padx=10, pady=10)

                    ##mainframe
                    GD = LabelFrame(MF, text='general detail', font=FONT3, width=800, height=500)
                    GD.grid(row=0, column=0, padx=10, pady=10)

                    #model
                    L = Label(GD, text='Model :', font=FONT2)
                    L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
                    v_model = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_model, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Ciena','Other']), state='readonly')
                    E.grid(row=0, column=1, padx=30, pady=10)

                    # #model number
                    L = Label(GD, text='Model number :', font=FONT2)
                    L.grid(row=1, column=0, padx=50, pady=10, sticky='e')
                    v_model_number = StringVar()
                    v_model_number.set(fixid)
                    E = ttk.Entry(GD, textvariable=v_model_number, font=FONT2)
                    E.grid(row=1, column=1, padx=30, pady=10)

                    # #customer
                    L = Label(GD, text='Customer :', font=FONT2)
                    L.grid(row=2, column=0, padx=50, pady=10, sticky='e')
                    v_customer = StringVar()
                    v_customer.set(customer)
                    E = ttk.Combobox(GD, textvariable=v_customer, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Ciena','Other']), state='readonly')
                    E.grid(row=2, column=1, padx=30, pady=10)

                    # #types
                    L = Label(GD, text='Types :', font=FONT2)
                    L.grid(row=3, column=0, padx=50, pady=10, sticky='e')
                    v_types = StringVar()
                    v_types.set('Fixture')
                    E = ttk.Combobox(GD, textvariable=v_types, font=FONT2, values=(['Fixture','Bord profile','Stencil','PPtool']), state='readonly')
                    E.grid(row=3, column=1, padx=30, pady=10)

                    # #q'ty
                    L = Label(GD, text="Q'ty (จำนวน Fixture):", font=FONT2)
                    L.grid(row=4, column=0, padx=50, pady=10, sticky='e')
                    v_qty = StringVar()
                    E = ttk.Entry(GD, textvariable=v_qty, font=FONT2)
                    E.grid(row=4, column=1, padx=30, pady=10)

                    # #unit
                    L = Label(GD, text="Unit :", font=FONT2)
                    L.grid(row=5, column=0, padx=50, pady=10, sticky='e')
                    v_unit = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_unit, font=FONT2, values=(['ชิ้น','คู่',
                                                                                    'ชุด']), state='readonly')
                    E.grid(row=5, column=1, padx=30, pady=10)

                    # #line
                    L = Label(GD, text='Line :', font=FONT2)
                    L.grid(row=6, column=0, padx=50, pady=10, sticky='e')
                    v_line = StringVar()
                    v_line.set(line)
                    E = ttk.Combobox(GD, textvariable=v_line, font=FONT2, values=(['Bld4#2','Bld5#10','Bld6#15']), state='readonly')
                    E.grid(row=6, column=1, padx=30, pady=10)

                    # #reg by
                    L = Label(GD, text='Reg by :', font=FONT2)
                    L.grid(row=7, column=0, padx=50, pady=10, sticky='e')
                    v_reg_by = StringVar()
                    v_reg_by.set(by)
                    E = ttk.Combobox(GD, textvariable=v_reg_by, font=FONT2, values=techname)
                    E.grid(row=7, column=1, padx=30, pady=10)

                    #other frame
                    OD = LabelFrame(MF, text='other details', width=200, height=200)
                    OD.grid(row=0, column=1, padx=10, pady=10)

                    # #descriptions
                    L = Label(OD, text="Descriptions :", font=FONT2)
                    L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
                    E_desc = Text(OD, font=FONT2, width=20, height=2)
                    E_desc.grid(row=0, column=1, padx=30, pady=10)

                    # #photo frame
                    F2 = LabelFrame(OD, text='Photo', font=FONT3, width=200, height=80)
                    F2.grid(row=1, column=0, padx=10, pady=10)

                    # photo show
                    PS = LabelFrame(OD, width=200, height=150)
                    PS.grid(row=1, column=1, padx=10, pady=10)

                    #label for show photo
                    photofix = Label(PS)
                    photofix.pack()

                    #entry
                    L = Label(OD, text='')
                    L.grid(row=2, column=0, pady=50)

                    # #photo path
                    v_photo_path = StringVar()
                    E = ttk.Entry(F2, textvariable=v_photo_path, font=FONT2)
                    E.pack(padx=10, pady=10)

                    # #button
                    B = ttk.Button(F2, text='Select', command=selectPhoto)
                    B.pack(padx=10, pady=10)

                    B2 = ttk.Button(OD, text='Save', command=saveToolReg)
                    B2.grid(row=2, column=1, sticky='w')

                    B3 = ttk.Button(OD, text='Clear', command=reset)
                    B3.grid(row=2, column=1, sticky='e', padx=20)

            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

    def bordProfile(self,tsid,model,num_model,line,side,with_by,physical,fixture_status,signal_status,desc,bordNumcheck,photo):

        self.model = model
        self.nummodel = num_model
        self.line = line
        self.side = side
        self.with_by = with_by
        self.physical = physical
        self.fixture_status = fixture_status
        self.signal_status = signal_status
        self.desc = desc
        self.photo = photo

        self.wb = load_workbook(self.toolpath)
        self.wb.active = self.wb[self.toolregsheet]
        self.sheet = self.wb[self.toolregsheet]
        rows = self.sheet.max_row

        data = ['CHECK','']
        for i in range(1,rows+1):
            if (self.sheet.cell(row=i,column=3).value).lower().replace(' ','').strip() == bordNumcheck:
                data.remove('')
                data.append('NOTFOUND')
                self.exB.StartToolWithdrawBorD(tsid,model,num_model,line,side,with_by,physical,fixture_status,signal_status,desc,photo)
                break

        if data[1] != 'NOTFOUND':
            
            try:
                check = messagebox.askyesno('Bordprofile register','หมายเลขนี้ยังไม่มีในฐานข้อมูลคุณต้องบันทึกลงในฐานข้อมูล!\n คลิ๊ก Yes เพื่อบันทึกข้อมูลและบันทึกการยืม')
                if check:
                    def center_windows(w,h):
                        ws = REGTOOLCHECK.winfo_screenwidth() #screen width
                        hs = REGTOOLCHECK.winfo_screenheight() #screen height
                        x = (ws/2) - (w/2)
                        y = (hs/2) - (h/2)
                        return f'{w}x{h}+{x:.0f}+{y:.0f}'

                    REGTOOLCHECK = Toplevel()
                    REGTOOLCHECK.title('Bordprofile register')
                    win_size = center_windows(800,600)
                    REGTOOLCHECK.geometry(win_size)
                    
                    #save reg
                    def saveToolReg(event=None):
                        tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
                        model = v_model.get()
                        modelNum = v_model_number.get()
                        customer = v_customer.get()
                        clsaaTypes = v_types.get()
                        qty = v_qty.get()
                        unit = v_unit.get()
                        line = v_line.get()
                        # desc = E_desc.get('1.0',END).strip()
                        regBy = v_reg_by.get()
                        check_photo_save = v_photoBord.get()
                        if model and modelNum and customer and clsaaTypes and qty and unit and line and regBy:

                            try:
                                self.exTR.StartToolReG(tsid,model,modelNum,customer,clsaaTypes,qty,unit,line,regBy,desc,check_photo_save)
                                self.exB.StartToolWithdrawBorD(tsid,self.model,self.nummodel,self.line,self.side,self.with_by,self.physical,self.fixture_status,self.signal_status,self.desc,self.photo) #เจอหรือไม่เจอใน dataBase ก็ต้องบันทึก
                                reset()
                                CloSE()
                            except Exception as e:
                                messagebox.showerror("Error", f"Failed to data: {e}")
                            
                        else:
                            messagebox.showinfo('Bordprofile register', 'โปรดกรอกข้อมูลให้ครบถ้วน')

                    def CloSE():
                        REGTOOLCHECK.destroy()
                        REGTOOLCHECK.update()

                    def reset():
                        v_model.set('')
                        v_model_number.set('')
                        v_customer.set('')
                        v_types.set('')
                        v_qty.set('')
                        v_unit.set('')
                        v_line.set('')
                        E_desc.delete('1.0',END)
                        v_reg_by.set('')
                        v_photoBord.set('')
                        photoBord.config(image=None)
                        photoBord.image = None

                    #select photo
                    def selectPhoto():
                        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                        v_photoBord.set(file_path)

                        ##
                        try:
                            image = PIL.Image.open(v_photoBord.get())
                            image = image.resize((190, 140))
                            photo = ImageTk.PhotoImage(image)
                            photoBord.config(image=photo)
                            photoBord.image = photo

                        except Exception as e:
                            messagebox.showinfo('Equipment', 'Failed to load photo {e}')

                    #registertool
                    MF = LabelFrame(REGTOOLCHECK)
                    MF.pack(padx=10, pady=10)

                    ##mainframe
                    GD = LabelFrame(MF, text='general detail', font=FONT3, width=800, height=500)
                    GD.grid(row=0, column=0, padx=10, pady=10)

                    #model
                    L = Label(GD, text='Model :', font=FONT2)
                    L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
                    v_model = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_model, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Other']), state='readonly')
                    E.grid(row=0, column=1, padx=30, pady=10)

                    # #model number
                    L = Label(GD, text='Model number :', font=FONT2)
                    L.grid(row=1, column=0, padx=50, pady=10, sticky='e')
                    v_model_number = StringVar()
                    v_model_number.set(num_model)
                    E = ttk.Entry(GD, textvariable=v_model_number, font=FONT2)
                    E.grid(row=1, column=1, padx=30, pady=10)

                    # #customer
                    L = Label(GD, text='Customer :', font=FONT2)
                    L.grid(row=2, column=0, padx=50, pady=10, sticky='e')
                    v_customer = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_customer, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Other']), state='readonly')
                    E.grid(row=2, column=1, padx=30, pady=10)

                    # #types
                    L = Label(GD, text='Types :', font=FONT2)
                    L.grid(row=3, column=0, padx=50, pady=10, sticky='e')
                    v_types = StringVar()
                    v_types.set('Bord profile')
                    E = ttk.Combobox(GD, textvariable=v_types, font=FONT2, values=(['Fixture','Bord profile','Stencil','PPtool']), state='readonly')
                    E.grid(row=3, column=1, padx=30, pady=10)

                    # #q'ty
                    L = Label(GD, text="Q'ty (จำนวน Bord) :", font=FONT2)
                    L.grid(row=4, column=0, padx=50, pady=10, sticky='e')
                    v_qty = StringVar()
                    E = ttk.Entry(GD, textvariable=v_qty, font=FONT2)
                    E.grid(row=4, column=1, padx=30, pady=10)

                    # #unit
                    L = Label(GD, text="Unit :", font=FONT2)
                    L.grid(row=5, column=0, padx=50, pady=10, sticky='e')
                    v_unit = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_unit, font=FONT2, values=(['ชิ้น','คู่',
                                                                                    'ชุด']), state='readonly')
                    E.grid(row=5, column=1, padx=30, pady=10)

                    # #line
                    L = Label(GD, text='Line :', font=FONT2)
                    L.grid(row=6, column=0, padx=50, pady=10, sticky='e')
                    v_line = StringVar()
                    v_line.set(line)
                    E = ttk.Combobox(GD, textvariable=v_line, font=FONT2, values=(['Bld4#2','Bld5#10','Bld6#15']), state='readonly')
                    E.grid(row=6, column=1, padx=30, pady=10)

                    # #reg by
                    L = Label(GD, text='Reg by :', font=FONT2)
                    L.grid(row=7, column=0, padx=50, pady=10, sticky='e')
                    v_reg_by = StringVar()
                    v_reg_by.set(with_by)
                    E = ttk.Combobox(GD, textvariable=v_reg_by, font=FONT2, values=techname)
                    E.grid(row=7, column=1, padx=30, pady=10)

                    #other frame
                    OD = LabelFrame(MF, text='other details', width=200, height=200)
                    OD.grid(row=0, column=1, padx=10, pady=10)

                    # #descriptions
                    L = Label(OD, text="Descriptions :", font=FONT2)
                    L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
                    E_desc = Text(OD, font=FONT2, width=20, height=2)
                    E_desc.grid(row=0, column=1, padx=30, pady=10)

                    # #photo frame
                    F2 = LabelFrame(OD, text='Photo', font=FONT3, width=200, height=80)
                    F2.grid(row=1, column=0, padx=10, pady=10)

                    # photo show
                    PS = LabelFrame(OD, width=200, height=150)
                    PS.grid(row=1, column=1, padx=10, pady=10)

                    #label for show photo
                    photoBord = Label(PS)
                    photoBord.pack()

                    #entry
                    L = Label(OD, text='')
                    L.grid(row=2, column=0, pady=50)

                    # #photo path
                    v_photoBord = StringVar()
                    E = ttk.Entry(F2, textvariable=v_photoBord, font=FONT2)
                    E.pack(padx=10, pady=10)

                    # #button
                    B = ttk.Button(F2, text='Select', command=selectPhoto)
                    B.pack(padx=10, pady=10)

                    B2 = ttk.Button(OD, text='Save', command=saveToolReg)
                    B2.grid(row=2, column=1, sticky='w')

                    B3 = ttk.Button(OD, text='Clear', command=reset)
                    B3.grid(row=2, column=1, sticky='e', padx=20)

            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")


    def stenCil(self,tsid,stenNum,line,slotNum,by,date,stencilCheck,photoStencilSave,comment):

        self.stenNum = stenNum
        self.line = line
        self.slotNum = slotNum
        self.by = by
        self.date = date
        self.photoStencilSave = photoStencilSave
        self.comment = comment

        self.wb = load_workbook(self.toolpath)
        self.wb.active = self.wb[self.toolregsheet]
        self.sheet = self.wb[self.toolregsheet]
        rows = self.sheet.max_row

        data = ['CHECK','']
        for i in range(1,rows+1):
            if (self.sheet.cell(row=i,column=3).value).lower().replace(' ','').strip() == stencilCheck:
                data.remove('')
                data.append('NOTFOUND')
                self.exSC.StartToolWithdrawstencil(tsid,stenNum,line,slotNum,by,date,photoStencilSave,comment)
                break

        if data[1] != 'NOTFOUND':
            
            try:
                check = messagebox.askyesno('Stencil register','หมายเลขนี้ยังไม่มีในฐานข้อมูลคุณต้องบันทึกลงในฐานข้อมูล!\n คลิ๊ก Yes เพื่อบันทึกข้อมูลและบันทึกการยืม')
                if check:
                    def center_windows(w,h):
                        ws = REGTOOLCHECK.winfo_screenwidth() #screen width
                        hs = REGTOOLCHECK.winfo_screenheight() #screen height
                        x = (ws/2) - (w/2)
                        y = (hs/2) - (h/2)
                        return f'{w}x{h}+{x:.0f}+{y:.0f}'

                    REGTOOLCHECK = Toplevel()
                    REGTOOLCHECK.title('Stencil register')
                    win_size = center_windows(800,600)
                    REGTOOLCHECK.geometry(win_size)
                    
                    #save reg
                    def saveToolReg(event=None):
                        tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
                        model = v_model.get()
                        modelNum = v_model_number.get()
                        customer = v_customer.get()
                        clsaaTypes = v_types.get()
                        qty = v_qty.get()
                        unit = v_unit.get()
                        line = v_line.get()
                        desc = E_desc.get('1.0',END).strip()
                        regBy = v_reg_by.get()
                        check_photo_save = v_photoBord.get()
                        if model and modelNum and customer and clsaaTypes and qty and unit and line and regBy:

                            try:
                                self.exTR.StartToolReG(tsid,model,modelNum,customer,clsaaTypes,qty,unit,line,regBy,desc,check_photo_save)
                                self.exSC.StartToolWithdrawstencil(tsid,self.stenNum,self.line,self.slotNum,self.by,self.date,self.photoStencilSave,self.comment) #เจอหรือไม่เจอใน dataBase ก็ต้องบันทึก
                                reset()
                                CloSE()
                            except Exception as e:
                                messagebox.showerror("Error", f"Failed to data: {e}")
                            
                        else:
                            messagebox.showinfo('Stencil register', 'โปรดกรอกข้อมูลให้ครบถ้วน')

                    def CloSE():
                        REGTOOLCHECK.destroy()
                        REGTOOLCHECK.update()

                    def reset():
                        v_model.set('')
                        v_model_number.set('')
                        v_customer.set('')
                        v_types.set('')
                        v_qty.set('')
                        v_unit.set('')
                        v_line.set('')
                        E_desc.delete('1.0',END)
                        v_reg_by.set('')
                        v_photoBord.set('')
                        photoBord.config(image=None)
                        photoBord.image = None

                    #select photo
                    def selectPhoto():
                        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                        v_photoBord.set(file_path)

                        ##
                        try:
                            image = PIL.Image.open(v_photoBord.get())
                            image = image.resize((190, 140))
                            photo = ImageTk.PhotoImage(image)
                            photoBord.config(image=photo)
                            photoBord.image = photo

                        except Exception as e:
                            messagebox.showinfo('Equipment', 'Failed to load photo {e}')

                    #registertool
                    MF = LabelFrame(REGTOOLCHECK)
                    MF.pack(padx=10, pady=10)

                    ##mainframe
                    GD = LabelFrame(MF, text='general detail', font=FONT3, width=800, height=500)
                    GD.grid(row=0, column=0, padx=10, pady=10)

                    #model
                    L = Label(GD, text='Model :', font=FONT2)
                    L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
                    v_model = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_model, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Ciena','Other']), state='readonly')
                    E.grid(row=0, column=1, padx=30, pady=10)

                    # #model number
                    L = Label(GD, text='Model number :', font=FONT2)
                    L.grid(row=1, column=0, padx=50, pady=10, sticky='e')
                    v_model_number = StringVar()
                    v_model_number.set(stenNum)
                    E = ttk.Entry(GD, textvariable=v_model_number, font=FONT2)
                    E.grid(row=1, column=1, padx=30, pady=10)

                    # #customer
                    L = Label(GD, text='Customer :', font=FONT2)
                    L.grid(row=2, column=0, padx=50, pady=10, sticky='e')
                    v_customer = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_customer, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Ciena','Other']), state='readonly')
                    E.grid(row=2, column=1, padx=30, pady=10)

                    # #types
                    L = Label(GD, text='Types :', font=FONT2)
                    L.grid(row=3, column=0, padx=50, pady=10, sticky='e')
                    v_types = StringVar()
                    v_types.set('Stencil')
                    E = ttk.Combobox(GD, textvariable=v_types, font=FONT2, values=(['Fixture','Bord profile','Stencil','PPtool']), state='readonly')
                    E.grid(row=3, column=1, padx=30, pady=10)

                    # #q'ty
                    L = Label(GD, text="Q'ty (จำนวน Stencil) :", font=FONT2)
                    L.grid(row=4, column=0, padx=50, pady=10, sticky='e')
                    v_qty = StringVar()
                    E = ttk.Entry(GD, textvariable=v_qty, font=FONT2)
                    E.grid(row=4, column=1, padx=30, pady=10)

                    # #unit
                    L = Label(GD, text="Unit :", font=FONT2)
                    L.grid(row=5, column=0, padx=50, pady=10, sticky='e')
                    v_unit = StringVar()
                    E = ttk.Combobox(GD, textvariable=v_unit, font=FONT2, values=(['ชิ้น','คู่',
                                                                                    'ชุด']), state='readonly')
                    E.grid(row=5, column=1, padx=30, pady=10)

                    # #line
                    L = Label(GD, text='Line :', font=FONT2)
                    L.grid(row=6, column=0, padx=50, pady=10, sticky='e')
                    v_line = StringVar()
                    v_line.set(line)
                    E = ttk.Combobox(GD, textvariable=v_line, font=FONT2, values=(['Bld4#2','Bld5#10','Bld6#15']), state='readonly')
                    E.grid(row=6, column=1, padx=30, pady=10)

                    # #reg by
                    L = Label(GD, text='Reg by :', font=FONT2)
                    L.grid(row=7, column=0, padx=50, pady=10, sticky='e')
                    v_reg_by = StringVar()
                    v_reg_by.set(by)
                    E = ttk.Combobox(GD, textvariable=v_reg_by, font=FONT2, values=techname)
                    E.grid(row=7, column=1, padx=30, pady=10)

                    #other frame
                    OD = LabelFrame(MF, text='other details', width=200, height=200)
                    OD.grid(row=0, column=1, padx=10, pady=10)

                    # #descriptions
                    L = Label(OD, text="Descriptions :", font=FONT2)
                    L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
                    E_desc = Text(OD, font=FONT2, width=20, height=2)
                    E_desc.grid(row=0, column=1, padx=30, pady=10)

                    # #photo frame
                    F2 = LabelFrame(OD, text='Photo', font=FONT3, width=200, height=80)
                    F2.grid(row=1, column=0, padx=10, pady=10)

                    # photo show
                    PS = LabelFrame(OD, width=200, height=150)
                    PS.grid(row=1, column=1, padx=10, pady=10)

                    #label for show photo
                    photoBord = Label(PS)
                    photoBord.pack()

                    #entry
                    L = Label(OD, text='')
                    L.grid(row=2, column=0, pady=50)

                    # #photo path
                    v_photoBord = StringVar()
                    E = ttk.Entry(F2, textvariable=v_photoBord, font=FONT2)
                    E.pack(padx=10, pady=10)

                    # #button
                    B = ttk.Button(F2, text='Select', command=selectPhoto)
                    B.pack(padx=10, pady=10)

                    B2 = ttk.Button(OD, text='Save', command=saveToolReg)
                    B2.grid(row=2, column=1, sticky='w')

                    B3 = ttk.Button(OD, text='Clear', command=reset)
                    B3.grid(row=2, column=1, sticky='e', padx=20)

            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

# เรียกมาใช้ก่อนน
# gui = Tk()
# obj_y = ToolCheck()
# obj_y.fixTure(1,2,3,4,5,6,7,8,9,10)
# gui.mainloop()
