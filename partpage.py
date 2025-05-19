from tkinter import ttk, Frame, END, LabelFrame, Label, StringVar, IntVar, Text, Toplevel, Menu, Scrollbar, Y, RIGHT
from tkinter import messagebox
from tkinter import filedialog, messagebox
from PIL import ImageTk
from openpyxl import load_workbook
from datetime import datetime
import subprocess
import tkinter as tk
import PIL.Image
import threading
import configparser

#font size
FONT1 = ('Angsana New', 25, 'bold')
FONT2 = ('Angsana New', 18, )
FONT3 = ('Angsana New', 16, )

#techname
techname = ['Thanongsak Su',
            'Khunakorn R',
            'Pichet T',
            'Wasan R',
            'Don P',
            'Somchai L',
            'Adirek C',
            'Sangworn D',
            'Pratchaya S',
            'Supot P',
            'Kriangsak H',
            'Anan C',
            'Thanatorn K',
            'Anong J',
            'Thanongsak D',
            'Apicha K',
            'Sompong L']

##--rec part
class PartReg(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        ##--function save
        def reg_part(event=None):
            from excel_function_spare import SpareparT
            excel = SpareparT()
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            p_name = v_partname.get()
            p_name_check = v_partname.get().lower().replace(' ','').strip()
            p_no = v_partno.get()
            usefor = v_usedfor.get()
            pcs_use_fre = v_freuse.get() 
            freq = v_period.get()
            qty = v_qty.get()
            stocklow = v_minimum.get()
            leadtime = v_leadtime.get()
            subcon = v_sub.get()
            subphone = v_contact.get()
            des = note.get('1.0',END).strip()
            photo_path_check = v_path_photo.get()
            
            if p_name and p_no and usefor and freq and pcs_use_fre and qty and stocklow and leadtime != (''):
                if qty>0:
                    try:
                        excel.StartpartReG(tsid,p_name,p_name_check,p_no,usefor,freq,pcs_use_fre,qty,stocklow,leadtime,subcon,subphone,des,photo_path_check)
                        reset()
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to data: {e}")
                else:
                    messagebox.showinfo('Spare part',"โปรดกรอกจำนวน Q'ty เป็นตัวเลข")
            
            else:
                messagebox.showinfo('Spare part', 'โปรดกรอกข้อมูลให้ครบถ้วน')

        def reset():
            v_partname.set('')
            v_partno.set('')
            v_usedfor.set('')
            v_freuse.set('')
            v_qty.set('')
            v_minimum.set('')
            v_leadtime.set('')
            note.delete('1.0',END)
            v_path_photo.set('')
            photoRemove()
            
        ##--function savephoto
        def select_photo():
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            v_path_photo.set(file_path)

            try:
                image = PIL.Image.open(v_path_photo.get())
                image = image.resize((300,240))
                photo = ImageTk.PhotoImage(image)
                photo_show.config(image=photo)
                photo_show.image = photo
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def photoRemove():
            photo_show.config(image=None)
            photo_show.image = None

        ##--frame
        PR = LabelFrame(self, width=400, height=420, font=FONT3)
        PR.grid(row=0, column=0)

        #general detail
        GD = LabelFrame(PR, text='general detail', width=500, height=500)
        GD.grid(row=0, column=0, padx=20, pady=10)

        #part name
        L = Label(GD, text='Part name :', font=FONT2)
        L.grid(row=0, column=0, padx=40, pady=10)
        v_partname = StringVar()
        E = ttk.Combobox(GD, textvariable=v_partname, font=FONT2, values='N/A')
        E.grid(row=0, column=1, padx=30, pady=10)

        ##--Part no :
        L = Label(GD, text='Part no :', font=FONT2)
        L.grid(row=1, column=0, padx=40, pady=10)
        v_partno = StringVar()
        E = ttk.Combobox(GD, textvariable=v_partno, font=FONT2, values='N/A')
        E.grid(row=1, column=1, padx=30, pady=10)

        ##--used for :
        L = Label(GD, text='Used for :', font=FONT2)
        L.grid(row=2, column=0, padx=40, pady=10)
        v_usedfor = StringVar()
        E = ttk.Combobox(GD, textvariable=v_usedfor, font=FONT2, values=['Machine','Accessories', 'Other'])
        E.grid(row=2, column=1, padx=30, pady=10)

        ##--usage fre :
        L = Label(GD, text='Usage frequency (pcs) :', font=FONT2)
        L.grid(row=4, column=0, padx=40, pady=10)
        v_freuse = IntVar()
        E = ttk.Entry(GD, textvariable=v_freuse, font=FONT2, width=10)
        E.grid(row=4, column=1, pady=10)

         ##--fre/period :
        v_period = StringVar()
        E = ttk.Combobox(GD, textvariable=v_period, font=FONT3, values=(['/Day','/Month','6Month','/Quarter','/Year']), state='readonly', width=5)
        E.grid(row=4, column=1, padx=2, pady=10, sticky='e')

        ##--จำนวนที่สั่งซื้อมา :
        L = Label(GD, text="Q'ty :", font=FONT2)
        L.grid(row=6, column=0, padx=30, pady=10)
        v_qty = IntVar()
        E = ttk.Entry(GD, textvariable=v_qty, font=FONT2)
        E.grid(row=6, column=1, padx=30, pady=10)

        ##--minimum stock
        L = Label(GD, text='Minimum stock :', font=FONT2)
        L.grid(row=7, column=0, padx=30, pady=10)
        v_minimum = IntVar()
        E = ttk.Entry(GD, textvariable=v_minimum, font=FONT2)
        E.grid(row=7, column=1, padx=30, pady=10)

        ##--lead time
        L = Label(GD, text='Lead time ARO (day):', font=FONT2)
        L.grid(row=8, column=0, pady=10)
        v_leadtime = IntVar()
        E = ttk.Entry(GD, textvariable=v_leadtime, font=FONT2)
        E.grid(row=8, column=1, pady=10)

        #other detail
        OD = LabelFrame(PR, text='other detail', width=500, height=500)
        OD.grid(row=0, column=1, padx=20, pady=10)

        sub = LabelFrame(OD, text='supplier detail', width=400, height=120, font=FONT3)
        sub.grid(row=0, column=0, padx=10, pady=10)

        ##sub
        L = Label(sub, text='Supplier :', font=FONT2)
        L.grid(row=0, column=0, padx=20, pady=10)
        v_sub = StringVar()
        E = ttk.Entry(sub, textvariable=v_sub, font=FONT2)
        E.grid(row=0, column=1, padx=20, pady=10)

        ##contrac phone
        L = Label(sub, text='Contact phone :', font=FONT2)
        L.grid(row=1, column=0, padx=20, pady=10)
        v_contact = StringVar()
        E = ttk.Entry(sub, textvariable=v_contact, font=FONT2)
        E.grid(row=1, column=1, padx=20, pady=10)

        ##note
        L = Label(sub, text='Note :', font=FONT2)
        L.grid(row=2, column=0)
        note = Text(sub, width=20, height=4)
        note.grid(row=2, column=1, pady=10)

        photo = LabelFrame(OD, text='photo', font=FONT3)
        photo.grid(row=0, column=1, padx=10, pady=10)

         #label photo part
        photo_show = tk.Label(photo)
        photo_show.pack(padx=10, pady=10)

        ##--path entry
        v_path_photo = StringVar()
        path_photo = ttk.Entry(photo, textvariable=v_path_photo, font=FONT3)
        path_photo.pack(padx=20, pady=10)

        #button reg photo part
        BRPPV = ttk.Button(photo, text='Select photo', command=select_photo)
        BRPPV.pack(pady=10)

        #button reg part
        BRPS = ttk.Button(OD, text='Save', command=reg_part)
        BRPS.grid(row=1, column=0, pady=5)
        BRPC = ttk.Button(OD, text='Clear', command=reset)
        BRPC.grid(row=2, column=0, pady=5)

        
    #back to home
    ##--progress
    def StartmCSwitchPagE(self):
        def center_windows(self,w,h):
            ws = self.winfo_screenwidth() #screen width
            hs = self.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        win_size = center_windows(self,200,100)
        window = Toplevel()
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กรุณารอสักครู่...', font=FONT1)
        label.pack()
        window.update()
        thread = threading.Thread(target = PartReg.Switch(self))
        thread.start()
        window.destroy()
        window.update()
 
    def Switch(self):
        subprocess.Popen(["python", "ASMT_Store_Part.py"])
        self.quit()

##--view part new
class PartViews(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.sparepath = self.config['DATABASE']['dbsparepath']
        self.sparesheet = self.config['DATABASE']['sparepartsheet']
        self.sparephotopath = self.config['DATABASE']['dbphotosparepath']

        #search_type
        search_type = Label(self, text='Search type :', font=FONT2)
        search_type.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('All')
        search_type_combo = ttk.Combobox(self, values=['Part name','Part no','All'], state='readonly', textvariable=v_search_type)
        search_type_combo.place(x=120, y=10, height=30)        

        #function search
        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()

            ##S/N search type:
            if check_search_type == ('Part name'):    #S/N search
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=0, max_col=16, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[col_idx - 1]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(), str(row[1]).strip(),
                                    str(row[2]).strip(), str(row[3]).strip(),
                                    str(row[9]).strip(), str(row[14]).strip()])
                return results
            
            #Part name
            elif check_search_type == ('Part no'):    
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=0, max_col=15, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[col_idx]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(), str(row[1]).strip(),
                                    str(row[2]).strip(), str(row[3]).strip(),
                                    str(row[9]).strip(), str(row[14]).strip()])
                return results
            
            #all
            elif check_search_type == ('All'):
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=0, max_col=16, values_only=True):
                    results.append([str(row[0]).strip(), str(row[1]).strip(),
                                    str(row[2]).strip(), str(row[3]).strip(),
                                    str(row[9]).strip(), str(row[14]).strip()])
                return results

        def on_key_release(event):
            search_prefix = enter.get().strip()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [3], search_prefix)  # Adjust columns as needed
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def display_results(results):
            for iter in P_list.get_children():
                P_list.delete(iter)
            if results:
                for row in results:
                    #row = tuple(row)
                    P_list['columns']=("TSID","Part name", "Part no" , 'Used for',"Stock (pcs)","Note")
                    P_list.column('TSID', anchor="center", width=5)
                    P_list.column('Part name', anchor="center", width=5)
                    P_list.column('Part no', anchor="center", width=5)
                    P_list.column('Used for', anchor="center", width=5)  
                    P_list.column('Stock (pcs)', anchor="center", width=5) 
                    P_list.column('Note', anchor="center", width=5) 

                    ##heading
                    P_list.heading('TSID', text='TSID',anchor="center")
                    P_list.heading('Part name', text='Part name', anchor="center")
                    P_list.heading('Part no', text='Part no',anchor="center")
                    P_list.heading('Used for', text='Used for', anchor="center")  
                    P_list.heading('Stock (pcs)', text='Stock (pcs)', anchor="center")  
                    P_list.heading('Note', text='Note', anchor="center")  
                    P_list.insert('', 'end', values=row)

            else:
                P_list['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                P_list.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                P_list.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")
            
            def delete_right_click(event=None):
                from excel_function_spare import SpareparT
                excel = SpareparT()
                select = P_list.selection()
                output = P_list.item(select)
                tsid = output['values'][0]
                check = messagebox.askyesno('ยืนยันการลบ','คุณต้องการลบข้อมูล?...')
                if check == True:
                    if output['values'] != (''):
                        try:
                            excel.StartpartDeletE(tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")

            #select delete
            def popup(event):
                pop_up_menu.post(event.x_root, event.y_root)
            P_list.bind('<Button-3>', popup)   #Button 3 คือ click ขวา
                    
            def getdata(tsid):
                wb = load_workbook(filename=file_path)
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=16, values_only=True):
                    if tsid ==  int(row[0]):  #TSID
                        listRow = list(row)
                        displaypopup(listRow)

            def Select_mc(event=None):
                select = P_list.selection()
                output = P_list.item(select)
                if select !=():
                    tsid = output['values'][0]
                    getdata(tsid)
                    
            def displaypopup(listRow):
                tsid = listRow[0]

                print(listRow)

                #Gui for detail
                withdraw_part = Toplevel()
                withdraw_part.title('Withdraw')
                withdraw_part.geometry('900x600-50+30')

                main_frame = LabelFrame(withdraw_part, text=f'TSID : {tsid}')
                main_frame.pack()

                L = Label(main_frame, text='Part name :', font=FONT2)
                L.grid(row=0, column=0, padx=10, pady=5)
                v_part_name = StringVar()
                v_part_name.set(listRow[1])
                E = ttk.Entry(main_frame, textvariable=v_part_name, font=FONT2, state='readonly')
                E.grid(row=0, column=1, padx=10, pady=5)

                L = Label(main_frame, text='Part no :', font=FONT2)
                L.grid(row=1, column=0, padx=10, pady=5)
                v_part_no = StringVar()
                v_part_no.set(listRow[2])
                E = ttk.Entry(main_frame, textvariable=v_part_no, font=FONT2, state='readonly')
                E.grid(row=1, column=1, padx=10, pady=5)

                L = Label(main_frame, text='Used for :', font=FONT2)
                L.grid(row=2, column=0, padx=10, pady=5)
                v_use_for = StringVar()
                v_use_for.set(listRow[3])
                E = ttk.Entry(main_frame, textvariable=v_use_for, font=FONT2, state='readonly')
                E.grid(row=2, column=1, padx=10, pady=5)

                L = Label(main_frame, text='Stock (pcs) :', font=FONT2)
                L.grid(row=3, column=0, padx=10, pady=5)
                v_stock = StringVar()
                v_stock.set(listRow[9])
                E = ttk.Entry(main_frame, textvariable=v_stock, font=FONT2, state='readonly')
                E.grid(row=3, column=1, padx=10, pady=5)

                L = Label(main_frame, text='Minimum stock  :', font=FONT2)
                L.grid(row=4, column=0, padx=10, pady=5)
                v_minimum_stock = StringVar()
                v_minimum_stock.set(listRow[10])
                E = ttk.Entry(main_frame, textvariable=v_minimum_stock, font=FONT2, state='readonly')
                E.grid(row=4, column=1, padx=10, pady=5)

                L = Label(main_frame, text='Lead time ARO :', font=FONT2)
                L.grid(row=5, column=0, padx=10, pady=5)
                v_leadtime = StringVar()
                v_leadtime.set(listRow[11])
                E = ttk.Entry(main_frame, textvariable=v_leadtime, font=FONT2, state='readonly')
                E.grid(row=5, column=1, padx=10, pady=5)

                L = Label(main_frame, text='Note :', font=FONT2)
                L.grid(row=6, column=0, padx=10, pady=5)
                v_note = StringVar()
                v_note.set(listRow[14])
                E7 = ttk.Entry(main_frame, textvariable=v_note, font=FONT2, state='readonly')
                E7.grid(row=6, column=1, padx=10, pady=5)

                L = Label(main_frame, text='Date update :', font=FONT2)
                L.grid(row=7, column=0, padx=10, pady=5)
                v_update_date = StringVar()
                v_update_date.set(listRow[15])
                E = ttk.Entry(main_frame, textvariable=v_update_date, font=FONT2, state='readonly')
                E.grid(row=7, column=1, padx=10, pady=5)

                L = Label(main_frame, text='/Day :', font=FONT2)
                L.grid(row=0, column=2)
                v_day = StringVar()
                v_day.set(listRow[4])
                E = ttk.Entry(main_frame, textvariable=v_day, font=FONT2, state='readonly')
                E.grid(row=0, column=3, padx=10, pady=5)


                L = Label(main_frame, text='/Month :', font=FONT2)
                L.grid(row=1, column=2)
                v_month = StringVar()
                v_month.set(listRow[5])
                E = ttk.Entry(main_frame, textvariable=v_month, font=FONT2, state='readonly')
                E.grid(row=1, column=3, padx=10, pady=5)



                L = Label(main_frame, text='/Quarter :', font=FONT2)
                L.grid(row=2, column=2)
                v_quarter = StringVar()
                v_quarter.set(listRow[6])
                E = ttk.Entry(main_frame, textvariable=v_quarter, font=FONT2, state='readonly')
                E.grid(row=2, column=3, padx=10, pady=5)



                L = Label(main_frame, text='/6month :', font=FONT2)
                L.grid(row=3, column=2)
                v_6month = StringVar()
                v_6month.set(listRow[7])
                E = ttk.Entry(main_frame, textvariable=v_6month, font=FONT2, state='readonly')
                E.grid(row=3, column=3, padx=10, pady=5)



                L = Label(main_frame, text='/Year :', font=FONT2)
                L.grid(row=4, column=2)
                v_year = StringVar()
                v_year.set(listRow[8])
                E = ttk.Entry(main_frame, textvariable=v_year, font=FONT2, state='readonly')
                E.grid(row=4, column=3, padx=10, pady=5)



                L = Label(main_frame, text='Supplier :', font=FONT2)
                L.grid(row=5, column=2)
                v_sub = StringVar()
                v_sub.set(listRow[12])
                E = ttk.Entry(main_frame, textvariable=v_sub, font=FONT2, state='readonly')
                E.grid(row=5, column=3, padx=10, pady=5)

                L = Label(main_frame, text='Contact phone :', font=FONT2)
                L.grid(row=6, column=2)
                v_sub_con = StringVar()
                v_sub_con.set(listRow[13])
                E = ttk.Entry(main_frame, textvariable=v_sub_con, font=FONT2, state='readonly')
                E.grid(row=6, column=3, padx=10, pady=5)

                ##--photo frame
                photo_f = LabelFrame(main_frame, text='Photo', font=FONT3, width=500, height=450)
                photo_f.grid(row=7, column=2)

                #แสดงรูปภาพ
                try:
                        photo_part = Label(photo_f)
                        photo_part.pack()
                        image = PIL.Image.open(f"{self.sparephotopath}\\{tsid}.png")
                        image = image.resize((500, 405))
                        photo = ImageTk.PhotoImage(image)
                        photo_part.config(image=photo)
                        photo_part.image = photo
                except Exception as e:
                        L = Label(photo_f, text='ไม่มีรูปภาพสำหรับ TSID นี้')
                        L.pack()
                        ()
                        
                    ##--withdraw frame
                frame_withdraw = LabelFrame(withdraw_part, text='With draw detail', font=FONT3, width=800, height=100)
                frame_withdraw.pack(padx=5, pady=5)

                ##--function withdraw
                def part_with(event=None):
                        from excel_function_spare import SpareparT
                        excel = SpareparT()
                        qty_part = v_qty_withdraw.get()
                        req_by = v_withdraw_by.get()
                        if qty_part and req_by !=(''):
                            excel.StartpartWithdraw(tsid, qty_part,req_by)
                            display_results(results)
                            on_key_release(event)
                            exit_withdraw_part()

                        else:
                            messagebox.showinfo('Spare','โปรดกรอกข้อมูลให้ครบ')
                    
                    #สั่งปิด withdraw part top level
                def exit_withdraw_part():
                        withdraw_part.destroy()
                        withdraw_part.update()

                    ##--detail
                L = Label(frame_withdraw, text='จำนวนที่ต้องการ :', font=FONT2)
                L.place(x=30, y=10)
                v_qty_withdraw = StringVar()
                E = ttk.Entry(frame_withdraw, textvariable=v_qty_withdraw, font=FONT2)
                E.place(x=160, y=10, height=30, width=100)

                L = Label(frame_withdraw, text='By :', font=FONT2)
                L.place(x=300, y=10)
                v_withdraw_by = StringVar()
                E = ttk.Combobox(frame_withdraw, textvariable=v_withdraw_by, font=FONT2, values=techname, state='readonly')
                E.place(x=350, y=10, height=30)

                ##--withdraw button
                part_withdraw = ttk.Button(frame_withdraw, text='Request', command=part_with)
                part_withdraw.place(x=700, y=20)

                withdraw_part.mainloop()

            P_list.bind('<Double-1>', Select_mc)

            pop_up_menu = Menu(self, tearoff=0)
            pop_up_menu.add_command(label='delete', command=delete_right_click)

        def clear_results():
            for iter in P_list.get_children():
                P_list.delete(iter)

        # Define Excel data for searching
        file_path = self.sparepath
        sheet_name = self.sparesheet

        #search entry
        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        enter = ttk.Entry(self, font=FONT2)
        enter.place(x=120, y=50, height=30, width=150)
        enter.bind('<KeyRelease>', on_key_release)

        #create list part
        header = ["TSID","Part name", "Part no" , 'Used for',"Stock (pcs)","Note"]
        headerw = [50,50,50,50,20,50]
        P_list = ttk.Treeview(self, columns=header, show='headings')
        P_list.place(x=20, y=100, width=1200, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',14))
        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")
        #--header and width
        for h,w in zip(header,headerw):
            P_list.heading(h,text=h)
            P_list.column(h,width=w,anchor='center')
        
        #scroll bar vertical
        P_list_scroll = Scrollbar(P_list)
        P_list_scroll.pack(side=RIGHT,fill=Y)
        P_list_scroll.config(command=P_list.yview)


# from tkinter import *
# gui = Tk()
# gui.title('Bord')
# gui.geometry('1500x1500')
# a = PartViews(gui)
# a.pack()
# gui.mainloop()

