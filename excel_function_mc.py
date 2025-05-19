import os
import pathlib
import openpyxl as xl
from tkinter import messagebox, Toplevel, Label
import threading
from PIL import Image
from datetime import datetime
import configparser

#font
FONT = ('Angsana New',18)

class MachineManagE:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.machinepath = self.config['DATABASE']['dbmachinepath'] 
        self.machinesheet = self.config['DATABASE']['machinesheet']

        self.machinephoto = self.config['DATABASE']['dbphotomachinepath']

        from send_mail import SendMail
        self.songmail = SendMail()


    ##--progress
    def StartmCReG(self,tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,desc1,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photo_path_check):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลงทะเบียนข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = MachineManagE.RegmC(self,tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,desc1,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photo_path_check))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Machine Register', 'ลงทะเบียนข้อมูลสำเร็จ')
        
    #insert function
    def RegmC(self,tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,service,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photo_path_check):
        
        #load worksheet
        route = pathlib.Path(self.machinepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.machinesheet]
        self.sheet = self.wb[self.machinesheet]
        rows = self.sheet.max_row
        date = datetime.now().strftime("%d/%b/%Y")

        try:
            if photo_path_check !=(''):
                image = Image.open(photo_path_check)
                target_path = self.machinephoto
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

            self.sheet.cell(row=rows+1,column=1).value=int(tsid)
            self.sheet.cell(row=rows+1,column=2,value=mc)
            self.sheet.cell(row=rows+1,column=3,value=mceds)
            self.sheet.cell(row=rows+1,column=4,value=brand)
            self.sheet.cell(row=rows+1,column=5,value=serial)
            self.sheet.cell(row=rows+1,column=6,value=cano)
            self.sheet.cell(row=rows+1,column=7,value=ora)
            self.sheet.cell(row=rows+1,column=8,value=wi)
            self.sheet.cell(row=rows+1,column=9,value=form)

            #อ้างอิงจาก pms
            if m ==1 :
                self.sheet.cell(row=rows+1,column=10,value='Y')
            if m3 == 1:
                self.sheet.cell(row=rows+1,column=11,value='Y')
            if m6 == 1:
                self.sheet.cell(row=rows+1,column=12,value='Y')
            if year == 1:
                self.sheet.cell(row=rows+1,column=13,value='Y')

            #แ้างอิงจาก wi dcc
            if mdcc ==1 :
                self.sheet.cell(row=rows+1,column=14,value='Y')
            if m3dcc == 1:
                self.sheet.cell(row=rows+1,column=15,value='Y')
            if m6dcc == 1:
                self.sheet.cell(row=rows+1,column=16,value='Y')
            if yeardcc == 1:
                self.sheet.cell(row=rows+1,column=17,value='Y')

            self.sheet.cell(row=rows+1,column=18).value = bld
            self.sheet.cell(row=rows+1,column=19).value = service
            self.sheet.cell(row=rows+1,column=20).value = comment
            self.sheet.cell(row=rows+1,column=21).value = date
            
            self.wb.save(self.machinepath)

            #Send mail notifucation
            self.songmail.Machinesongmail(mc, mceds, brand, serial, cano, ora, wi, form, bld, service, [{'M':m}, {'3M':m3}, {'6M':m6}, {'Year':year}], [{'M':mdcc}, {'3M':m3dcc}, {'6M':m6dcc}, {'Year':yeardcc}], comment, photo_path_check)
           
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    ##--progress
    def StartmCUpdatE(self,tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,service,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photoupdate):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'

        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังอัปเดตข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = MachineManagE.update_mc(self,tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,service,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photoupdate))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Machine Management', 'อัปเดตข้อมูลสำเร็จ')                        

    def update_mc(self,tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,service,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photoupdate):
        
        #load worksheet
        route = pathlib.Path(self.machinepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.machinesheet]
        self.sheet = self.wb[self.machinesheet]
        rows = self.sheet.max_row
        try:
            if photoupdate !=(''):
                image = Image.open(photoupdate)
                target_path = self.machinephoto
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)
            for i in range(2,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    self.sheet.cell(row=i,column=2,value=mc)
                    self.sheet.cell(row=i,column=3,value=mceds)
                    self.sheet.cell(row=i,column=4,value=brand)
                    self.sheet.cell(row=i,column=5,value=serial)
                    self.sheet.cell(row=i,column=6,value=cano)
                    self.sheet.cell(row=i,column=7,value=ora)
                    self.sheet.cell(row=i,column=8,value=wi)
                    self.sheet.cell(row=i,column=9,value=form)

                    #clear periodic
                    self.sheet.cell(row=i,column=10,value='')
                    self.sheet.cell(row=i,column=11,value='')
                    self.sheet.cell(row=i,column=12,value='')
                    self.sheet.cell(row=i,column=13,value='')
                    self.sheet.cell(row=i,column=14,value='')
                    self.sheet.cell(row=i,column=15,value='')
                    self.sheet.cell(row=i,column=16,value='')
                    self.sheet.cell(row=i,column=17,value='')

                    #อ้างอิงจาก pms
                    if m ==1 :
                        self.sheet.cell(row=i,column=10,value='Y')
                    if m3 == 1:
                        self.sheet.cell(row=i,column=11,value='Y')
                    if m6 == 1:
                        self.sheet.cell(row=i,column=12,value='Y')
                    if year == 1:
                        self.sheet.cell(row=i,column=13,value='Y')

                    #อ้างอิงจาก wi dcc
                    if mdcc ==1 :
                        self.sheet.cell(row=i,column=14,value='Y')
                    if m3dcc == 1:
                        self.sheet.cell(row=i,column=15,value='Y')
                    if m6dcc == 1:
                        self.sheet.cell(row=i,column=16,value='Y')
                    if yeardcc == 1:
                        self.sheet.cell(row=i,column=17,value='Y')

                    self.sheet.cell(row=i,column=18).value = bld
                    self.sheet.cell(row=i,column=19).value = service
                    self.sheet.cell(row=i,column=20).value = comment
                    self.wb.save(self.machinepath)

                    #notification update
                    self.songmail.Machinesongmail(mc, mceds, brand, serial, cano, ora, wi, form, bld, service, '-', '-', 'Update', '')
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    ##--progress
    def StartmCDeletE(self,tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลบข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = MachineManagE.delete_mc(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Machine Management', 'ลบข้อมูลสำเร็จ')                        

    def delete_mc(self,tsid):
        #load worksheet
        route = pathlib.Path(self.machinepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.machinesheet]
        self.sheet = self.wb[self.machinesheet]
        rows = self.sheet.max_row
        try:
            try:
                os.remove(f"{self.machinephoto}\\{tsid}.png")
            except:
                ()
            for i in range(2,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    mc = self.sheet.cell(row=i,column=2).value
                    mceds = self.sheet.cell(row=i,column=3).value
                    brand = self.sheet.cell(row=i,column=4).value
                    serial = self.sheet.cell(row=i,column=5).value
                    cano = self.sheet.cell(row=i,column=6).value
                    ora = self.sheet.cell(row=i,column=7).value
                    wi = self.sheet.cell(row=i,column=8).value
                    form = self.sheet.cell(row=i,column=9).value
                    bld = self.sheet.cell(row=i,column=18).value
                    service = self.sheet.cell(row=i,column=19).value

                    self.sheet.delete_rows(i,1)
                    self.wb.save(self.machinepath)

                    #delete notify
                    self.songmail.Machinesongmail(mc, mceds, brand, serial, cano, ora, wi, form, bld, service, '-', '-', 'Delete', '')

        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")


# A = MachineManagE()
# A.StartmCReG('tsid','mc','mceds','brand','serial','cano','ora','wi','form','bld','desc1','m','m3','m6','year','mdcc','m3dcc','m6dcc','yeardcc','comment',' ')

