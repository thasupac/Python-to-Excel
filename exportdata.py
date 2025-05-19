from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk, messagebox
import configparser
from calendar_widget import MyDatePicker
from datetime import datetime, timedelta
import threading
import os

#font size
FONT1 = ('Angsana New', 25, 'bold')
FONT2 = ('Angsana New', 18, )
FONT3 = ('Angsana New', 16, )
FONT4 = ('Angsana New', 12, )

class ExportDataStencil(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=500, height=500)

        #config
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        #Access data
        #path
        self.dBtoolPath = self.config['DATABASE']['dBtoolpath']
        self.stencilSheet = self.config['DATABASE']['stencilSheet']
       
        #template
        self.stencil_tem = self.config['TEMPLATE']['stencil']
        
        #exportpath
        self.stencil_export_path = self.config['EXPORTPATH']['stencil']

        #exportdata pname
        self.stencil_name = self.config['EXPORTNAME']['stencil']
        
        def frompopup(event):
            MyDatePicker(widget=Efrom)

        def topopup(event):
            MyDatePicker(widget=Eto)

        def exportdata():
            query = v_query.get()
            fromdate = v_dateForm.get()
            todate = v_dataTo.get()
            stencilNum = v_stencilNum.get()
            bld_select = v_bld.get()
            if fromdate and todate and query:
                ExportDataStencil.startexport(self, query, stencilNum, bld_select, fromdate, todate)
            else:
                messagebox.showinfo('Export data', 'โปรดระบุข้อมูลให้ครบ')
        #mainframe
        MF = LabelFrame(self)
        MF.pack(padx=10, pady=10)

        #query frame
        QF = LabelFrame(MF, text='Query by', font=FONT3)
        QF.pack(padx=10, pady=10)

        #query
        def blank(event):
            query = v_query.get()
            if query == 'Stencil number':
                Eblank.config(state='normal')
            else:
                Eblank.config(state='readonly')
            
            if query == 'BLD':
                bld_select.config(state='readonly')
            else:
                bld_select.config(state='disabled')

        L = Label(QF, text='Query By :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_query = StringVar()
        Es = ttk.Combobox(QF, textvariable=v_query, font=FONT2, width=17, values=['Stencil number','BLD','Date','All'], state='readonly')
        Es.grid(row=0, column=1, padx=5, pady=5)
        Es.bind('<<ComboboxSelected>>', blank)

        #stencil num
        L = Label(QF, text='Stencil no :', font=FONT2)
        L.grid(row=1, column=0, padx=5, pady=5)
        v_stencilNum = StringVar()
        Eblank = ttk.Entry(QF, textvariable=v_stencilNum, font=FONT2)
        Eblank.grid(row=1, column=1, padx=5, pady=5)

        #bld
        L = Label(QF, text='Building :', font=FONT2)
        L.grid(row=2, column=0, padx=5, pady=5)
        v_bld = StringVar()
        bld_select = ttk.Combobox(QF, textvariable=v_bld, font=FONT2, values=['BLD4#2','BLD5#10/1','BLD5#10/2','BLD6#15/1','BLD6#15/2'],state='readonly', width=18, )
        bld_select.grid(row=2, column=1, padx=5, pady=5)

        #date frame
        DF = LabelFrame(MF, text='Date', font=FONT3)
        DF.pack(padx=10, pady=10)

        #from
        L = Label(DF, text='From :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_dateForm = StringVar()
        Efrom = ttk.Entry(DF, textvariable=v_dateForm, font=FONT2)
        Efrom.grid(row=0, column=1, padx=10, pady=5)
        Efrom.bind('<Double-1>', frompopup)

        #to
        L = Label(DF, text='To :', font=FONT2)
        L.grid(row=1, column=0, padx=10, pady=5)
        v_dataTo = StringVar()
        Eto = ttk.Entry(DF, textvariable=v_dataTo, font=FONT2)
        Eto.grid(row=1, column=1, padx=10, pady=5)
        Eto.bind('<Double-1>', topopup)

        #button
        B = ttk.Button(self, text='Export', command=exportdata)
        B.pack(ipadx=5, ipady=5, pady=5)
    
    def startexport(self, query, stencilNum, bld_select, startdate, stopdate):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังส่งออกข้อมูล...', font=FONT4)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExportDataStencil.Exportstencil(self, query, stencilNum, bld_select, startdate, stopdate))
        thread.start()
        window.destroy()
        window.update()

    def Exportstencil(self, query, stencilNum, bld_select, startdate, stopdate):
        self.dBtool = load_workbook(filename=self.dBtoolPath)
        self.stencil = self.dBtool[self.stencilSheet]

        # Step 1: Define the start and stop dates
        start_date = datetime.strptime(startdate, "%d/%b/%Y")
        stop_date = datetime.strptime(stopdate, "%d/%b/%Y")

        # Step 2: Generate the date range
        current_date = start_date
        date_range = []

        while current_date <= stop_date:
            date_range.append(current_date.strftime("%d/%b/%Y"))  # Add the formatted date to the list
            current_date += timedelta(days=1)  # Move to the next day

        if query == 'Date':
            stencildata = []
            for date in date_range:
                for row in self.stencil.iter_rows(min_row=4, max_row=self.stencil.max_row, min_col=0, max_col=27, values_only=True):
                    if row[5] == date:
                        stencildata.append(row)

            #load template
            stenciltemplate = load_workbook(filename=self.stencil_tem )
            stencilexport = stenciltemplate.active
            start_row = 6
            for d in stencildata:
                stencilexport.cell(row=start_row,column=1).value=d[0]
                stencilexport.cell(row=start_row,column=2).value=d[1]
                stencilexport.cell(row=start_row,column=3).value=d[2]
                stencilexport.cell(row=start_row,column=4).value=d[3]
                stencilexport.cell(row=start_row,column=5).value=d[4]
                stencilexport.cell(row=start_row,column=6).value=d[5]
                stencilexport.cell(row=start_row,column=7).value=d[6]
                stencilexport.cell(row=start_row,column=8).value=d[7]
                stencilexport.cell(row=start_row,column=9).value=d[8]
                stencilexport.cell(row=start_row,column=10).value=d[9]
                stencilexport.cell(row=start_row,column=11).value=d[10]
                stencilexport.cell(row=start_row,column=12).value=d[11]
                stencilexport.cell(row=start_row,column=13).value=d[12]
                stencilexport.cell(row=start_row,column=14).value=d[13]
                stencilexport.cell(row=start_row,column=15).value=d[14]
                stencilexport.cell(row=start_row,column=16).value=d[15]
                stencilexport.cell(row=start_row,column=17).value=d[16]
                stencilexport.cell(row=start_row,column=18).value=d[17]
                stencilexport.cell(row=start_row,column=19).value=d[18]
                stencilexport.cell(row=start_row,column=20).value=d[19]
                stencilexport.cell(row=start_row,column=21).value=d[20]
                stencilexport.cell(row=start_row,column=22).value=d[21]
                stencilexport.cell(row=start_row,column=23).value=d[22]
                stencilexport.cell(row=start_row,column=24).value=d[23]
                stencilexport.cell(row=start_row,column=25).value=d[24]
                stencilexport.cell(row=start_row,column=26).value=d[25]
                stencilexport.cell(row=start_row,column=27).value=d[26]
                
                start_row+=1
            stenciltemplate.save(f'{self.stencil_export_path}/{self.stencil_name }.xlsx')

            #excel popup open
            excelopen = '{}/{}.xlsx'.format(self.stencil_export_path,self.stencil_name)
            os.system('start "excel" "{}"'.format(excelopen))
           
        elif query == 'Stencil number':
            stencildataNum = []
            for date in date_range:
                for row in self.stencil.iter_rows(min_row=4, max_row=self.stencil.max_row, min_col=0, max_col=27, values_only=True):
                    if date == row[5] and stencilNum == row[1]:
                        stencildataNum.append(row)
            
            #load template
            stenciltemplate = load_workbook(filename=self.stencil_tem )
            stencilexport = stenciltemplate.active
            start_row = 6
            for d in stencildataNum:
                stencilexport.cell(row=start_row,column=1).value=d[0]
                stencilexport.cell(row=start_row,column=2).value=d[1]
                stencilexport.cell(row=start_row,column=3).value=d[2]
                stencilexport.cell(row=start_row,column=4).value=d[3]
                stencilexport.cell(row=start_row,column=5).value=d[4]
                stencilexport.cell(row=start_row,column=6).value=d[5]
                stencilexport.cell(row=start_row,column=7).value=d[6]
                stencilexport.cell(row=start_row,column=8).value=d[7]
                stencilexport.cell(row=start_row,column=9).value=d[8]
                stencilexport.cell(row=start_row,column=10).value=d[9]
                stencilexport.cell(row=start_row,column=11).value=d[10]
                stencilexport.cell(row=start_row,column=12).value=d[11]
                stencilexport.cell(row=start_row,column=13).value=d[12]
                stencilexport.cell(row=start_row,column=14).value=d[13]
                stencilexport.cell(row=start_row,column=15).value=d[14]
                stencilexport.cell(row=start_row,column=16).value=d[15]
                stencilexport.cell(row=start_row,column=17).value=d[16]
                stencilexport.cell(row=start_row,column=18).value=d[17]
                stencilexport.cell(row=start_row,column=19).value=d[18]
                stencilexport.cell(row=start_row,column=20).value=d[19]
                stencilexport.cell(row=start_row,column=21).value=d[20]
                stencilexport.cell(row=start_row,column=22).value=d[21]
                stencilexport.cell(row=start_row,column=23).value=d[22]
                stencilexport.cell(row=start_row,column=24).value=d[23]
                stencilexport.cell(row=start_row,column=25).value=d[24]
                stencilexport.cell(row=start_row,column=26).value=d[25]
                stencilexport.cell(row=start_row,column=27).value=d[26]
                
                start_row+=1
            stenciltemplate.save(f'{self.stencil_export_path}/{self.stencil_name }.xlsx')

            #excel popup open
            excelopen = '{}/{}.xlsx'.format(self.stencil_export_path,self.stencil_name)
            os.system('start "excel" "{}"'.format(excelopen))

        #BLD
        elif query == 'BLD':
            stencilBLD = []
            for date in date_range:
                for row in self.stencil.iter_rows(min_row=4, max_row=self.stencil.max_row, min_col=0, max_col=27, values_only=True):
                    if date == row[5] and bld_select == row[2]:
                        stencilBLD.append(row)

            #load template
            stenciltemplate = load_workbook(filename=self.stencil_tem )
            stencilexport = stenciltemplate.active
            start_row = 6
            for d in stencilBLD:
                stencilexport.cell(row=start_row,column=1).value=d[0]
                stencilexport.cell(row=start_row,column=2).value=d[1]
                stencilexport.cell(row=start_row,column=3).value=d[2]
                stencilexport.cell(row=start_row,column=4).value=d[3]
                stencilexport.cell(row=start_row,column=5).value=d[4]
                stencilexport.cell(row=start_row,column=6).value=d[5]
                stencilexport.cell(row=start_row,column=7).value=d[6]
                stencilexport.cell(row=start_row,column=8).value=d[7]
                stencilexport.cell(row=start_row,column=9).value=d[8]
                stencilexport.cell(row=start_row,column=10).value=d[9]
                stencilexport.cell(row=start_row,column=11).value=d[10]
                stencilexport.cell(row=start_row,column=12).value=d[11]
                stencilexport.cell(row=start_row,column=13).value=d[12]
                stencilexport.cell(row=start_row,column=14).value=d[13]
                stencilexport.cell(row=start_row,column=15).value=d[14]
                stencilexport.cell(row=start_row,column=16).value=d[15]
                stencilexport.cell(row=start_row,column=17).value=d[16]
                stencilexport.cell(row=start_row,column=18).value=d[17]
                stencilexport.cell(row=start_row,column=19).value=d[18]
                stencilexport.cell(row=start_row,column=20).value=d[19]
                stencilexport.cell(row=start_row,column=21).value=d[20]
                stencilexport.cell(row=start_row,column=22).value=d[21]
                stencilexport.cell(row=start_row,column=23).value=d[22]
                stencilexport.cell(row=start_row,column=24).value=d[23]
                stencilexport.cell(row=start_row,column=25).value=d[24]
                stencilexport.cell(row=start_row,column=26).value=d[25]
                stencilexport.cell(row=start_row,column=27).value=d[26]
                
                start_row+=1
            stenciltemplate.save(f'{self.stencil_export_path}/{self.stencil_name }.xlsx')

            #excel popup open
            excelopen = '{}/{}.xlsx'.format(self.stencil_export_path,self.stencil_name)
            os.system('start "excel" "{}"'.format(excelopen))

        elif query == 'All':
            stencildataAll = []
            for date in date_range:
            
                for row in self.stencil.iter_rows(min_row=4, max_row=self.stencil.max_row, min_col=0, max_col=27, values_only=True):
                    if row[5] == date:
                        stencildataAll.append(row)
            
            #load template
            stenciltemplate = load_workbook(filename=self.stencil_tem )
            stencilexport = stenciltemplate.active
            start_row = 6
            for d in stencildataAll:
                stencilexport.cell(row=start_row,column=1).value=d[0]
                stencilexport.cell(row=start_row,column=2).value=d[1]
                stencilexport.cell(row=start_row,column=3).value=d[2]
                stencilexport.cell(row=start_row,column=4).value=d[3]
                stencilexport.cell(row=start_row,column=5).value=d[4]
                stencilexport.cell(row=start_row,column=6).value=d[5]
                stencilexport.cell(row=start_row,column=7).value=d[6]
                stencilexport.cell(row=start_row,column=8).value=d[7]
                stencilexport.cell(row=start_row,column=9).value=d[8]
                stencilexport.cell(row=start_row,column=10).value=d[9]
                stencilexport.cell(row=start_row,column=11).value=d[10]
                stencilexport.cell(row=start_row,column=12).value=d[11]
                stencilexport.cell(row=start_row,column=13).value=d[12]
                stencilexport.cell(row=start_row,column=14).value=d[13]
                stencilexport.cell(row=start_row,column=15).value=d[14]
                stencilexport.cell(row=start_row,column=16).value=d[15]
                stencilexport.cell(row=start_row,column=17).value=d[16]
                stencilexport.cell(row=start_row,column=18).value=d[17]
                stencilexport.cell(row=start_row,column=19).value=d[18]
                stencilexport.cell(row=start_row,column=20).value=d[19]
                stencilexport.cell(row=start_row,column=21).value=d[20]
                stencilexport.cell(row=start_row,column=22).value=d[21]
                stencilexport.cell(row=start_row,column=23).value=d[22]
                stencilexport.cell(row=start_row,column=24).value=d[23]
                stencilexport.cell(row=start_row,column=25).value=d[24]
                stencilexport.cell(row=start_row,column=26).value=d[25]
                stencilexport.cell(row=start_row,column=27).value=d[26]
                
                start_row+=1
            stenciltemplate.save(f'{self.stencil_export_path}/{self.stencil_name }.xlsx')

            #excel popup open
            excelopen = '{}/{}.xlsx'.format(self.stencil_export_path,self.stencil_name)
            os.system('start "excel" "{}"'.format(excelopen))

class ExportDatadowntime(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=500, height=500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        #path
        self.downtimepath = self.config['DATABASE']['dbdowntimepath']
        self.downtimesheet = self.config['DATABASE']['downtimesheet']
       
        #template
        self.downtime_tem = self.config['TEMPLATE']['downtime']
        
        #exportpath
        self.downtime_export_path = self.config['EXPORTPATH']['downtime']

        #exportdata pname
        self.downtime_name = self.config['EXPORTNAME']['downtime']

        def frompopup(event):
            MyDatePicker(widget=Efrom)

        def topopup(event):
            MyDatePicker(widget=Eto)

        def exportdata():
            query = v_query.get()
            fromdate = v_dateForm.get()
            todate = v_dataTo.get()
            bld_select = v_bld.get()
            if fromdate and todate and query:
                ExportDatadowntime.startexportdowntime(self, query, bld_select, fromdate, todate)
            else:
                messagebox.showinfo('Export data', 'โปรดระบุข้อมูลให้ครบ')

        #query
        def blank(event):
            query = v_query.get()
            if query == 'All':
                bld_select.config(state='disabled')
            else:
                bld_select.config(state='normal')
        #mainframe
        MF = LabelFrame(self)
        MF.pack(padx=10, pady=10)

        #query frame
        QF = LabelFrame(MF, text='Query by', font=FONT3)
        QF.pack(padx=10, pady=10)
        L = Label(QF, text='Query By :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_query = StringVar()
        Es = ttk.Combobox(QF, textvariable=v_query, font=FONT2, width=17, values=['BLD','All'], state='readonly')
        Es.grid(row=0, column=1, padx=5, pady=5)
        Es.bind('<<ComboboxSelected>>', blank)

        #bld
        L = Label(QF, text='Building :', font=FONT2)
        L.grid(row=1, column=0, padx=5, pady=5)
        v_bld = StringVar()
        bld_select = ttk.Combobox(QF, textvariable=v_bld, font=FONT2, values=['BLD4#2','BLD5#10/1','BLD5#10/2','BLD6#15/1','BLD6#15/2'],state='readonly', width=18, )
        bld_select.grid(row=1, column=1, padx=5, pady=5)

        #date frame
        DF = LabelFrame(MF, text='Date', font=FONT3)
        DF.pack(padx=10, pady=10)

        #from
        L = Label(DF, text='From :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_dateForm = StringVar()
        Efrom = ttk.Entry(DF, textvariable=v_dateForm, font=FONT2)
        Efrom.grid(row=0, column=1, padx=10, pady=5)
        Efrom.bind('<Double-1>', frompopup)

        #to
        L = Label(DF, text='To :', font=FONT2)
        L.grid(row=1, column=0, padx=10, pady=5)
        v_dataTo = StringVar()
        Eto = ttk.Entry(DF, textvariable=v_dataTo, font=FONT2)
        Eto.grid(row=1, column=1, padx=10, pady=5)
        Eto.bind('<Double-1>', topopup)

        #button
        B = ttk.Button(self, text='Export', command=exportdata)
        B.pack(ipadx=5, ipady=5, pady=5)

    def startexportdowntime(self, query, bld_select, fromdate, todate):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังส่งออกข้อมูล...', font=FONT4)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExportDatadowntime.Exportdowntime(self, query, bld_select, fromdate, todate))
        thread.start()
        window.destroy()
        window.update()
    
    def Exportdowntime(self, query, bld_select, fromdate, todate):
        self.dbdowntime = load_workbook(filename=self.downtimepath)
        self.downtime = self.dbdowntime[self.downtimesheet]

        # Step 1: Define the start and stop dates
        start_date = datetime.strptime(fromdate, "%d/%b/%Y")
        stop_date = datetime.strptime(todate, "%d/%b/%Y")

        # Step 2: Generate the date range
        current_date = start_date
        date_range = []

        while current_date <= stop_date:
            date_range.append(current_date.strftime("%d/%b/%Y"))  # Add the formatted date to the list
            current_date += timedelta(days=1)  # Move to the next day

        if query == 'BLD':
            downtimedata = []
            for date in date_range:
                for row in self.downtime.iter_rows(min_row=2, max_row=self.downtime.max_row, min_col=0, max_col=12, values_only=True):
                    if row[1].split(':')[-1].strip() == date and row[10].strip() == bld_select:
                        downtimedata.append(row)

        elif query == 'All':
            downtimedata = []
            for date in date_range:
                for row in self.downtime.iter_rows(min_row=3, max_row=self.downtime.max_row, min_col=0, max_col=12, values_only=True):
                    if row[1].split(':')[-1].strip() == date:
                        downtimedata.append(row)

            # load template
            downtimetemplate = load_workbook(filename=self.downtime_tem )
            downtimeexport = downtimetemplate.active
            start_row = 6
            for d in downtimedata:
                try:
                    downtimeexport.cell(row=start_row,column=1).value=d[1].split(':')[-1]
                    downtimeexport.cell(row=start_row,column=2).value=(d[1].split()[0])
                except Exception as e:
                    e
                downtimeexport.cell(row=start_row,column=3).value=d[2]  #machine
                downtimeexport.cell(row=start_row,column=4).value=d[3]  #inform
                downtimeexport.cell(row=start_row,column=5).value=d[4]  #problem
                downtimeexport.cell(row=start_row,column=6).value=d[5]  #solution
                try:
                    downtimeexport.cell(row=start_row,column=7).value=d[6].split(':')[-1]  #start action
                    downtimeexport.cell(row=start_row,column=8).value=(d[6].split()[0])  #start action
                    downtimeexport.cell(row=start_row,column=9).value=d[7].split(':')[-1]   #finishedaction
                    downtimeexport.cell(row=start_row,column=10).value=(d[7].split()[0])   #finishedaction
                    downtimeexport.cell(row=start_row,column=11).value=d[8] #actionby
                except Exception as e:
                    e

                downtimeexport.cell(row=start_row,column=12).value=d[9] #note
                
                start_row+=1
            downtimetemplate.save(f'{self.downtime_export_path}/{self.downtime_name }.xlsx')

            #excel popup open
            excelopen = '{}/{}.xlsx'.format(self.downtime_export_path,self.downtime_name)
            os.system('start "excel" "{}"'.format(excelopen))

class ExportDatatooling(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=500, height=500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        #path
        self.toolingpath = self.config['DATABASE']['dbtoolpath']
        self.toolingsheet = self.config['DATABASE']['regtoolsheet']
       
        #template
        self.tool_tem = self.config['TEMPLATE']['tool']
        
        #exportpath
        self.tool_export_path = self.config['EXPORTPATH']['tool']

        #exportdata pname
        self.tool_name = self.config['EXPORTNAME']['tool']

        def frompopup(event):
            MyDatePicker(widget=Efrom)

        def topopup(event):
            MyDatePicker(widget=Eto)

        def exportdata():
            query = v_query.get()
            fromdate = v_dateForm.get()
            todate = v_dataTo.get()
            bld_select = v_bld.get()
            if query and fromdate and todate:
                ExportDatatooling.startexporttoolingdata(self, query, bld_select, fromdate, todate)
            else:
                messagebox.showinfo('Export data', 'โปรดระบุข้อมูลให้ครบ')

        #query
        def blank(event):
            query = v_query.get()
            if query == 'All':
                bld_select.config(state='disabled')
                
            else:
                bld_select.config(state='normal')
        #mainframe
        MF = LabelFrame(self)
        MF.pack(padx=10, pady=10)

        #query frame
        QF = LabelFrame(MF, text='Query by', font=FONT3)
        QF.pack(padx=10, pady=10)
        L = Label(QF, text='Query By :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_query = StringVar()
        Es = ttk.Combobox(QF, textvariable=v_query, font=FONT2, width=17, values=['BLD','All'], state='readonly')
        Es.grid(row=0, column=1, padx=5, pady=5)
        Es.bind('<<ComboboxSelected>>', blank)

        #bld
        L = Label(QF, text='Building :', font=FONT2)
        L.grid(row=1, column=0, padx=5, pady=5)
        v_bld = StringVar()
        bld_select = ttk.Combobox(QF, textvariable=v_bld, font=FONT2, values=['BLD4#2','BLD5#10/1','BLD5#10/2','BLD6#15/1','BLD6#15/2'],state='readonly', width=18, )
        bld_select.grid(row=1, column=1, padx=5, pady=5)

        #date frame
        DF = LabelFrame(MF, text='Date', font=FONT3)
        DF.pack(padx=10, pady=10)

        #from
        L = Label(DF, text='From :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_dateForm = StringVar()
        Efrom = ttk.Entry(DF, textvariable=v_dateForm, font=FONT2)
        Efrom.grid(row=0, column=1, padx=10, pady=5)
        Efrom.bind('<Double-1>', frompopup)

        #to
        L = Label(DF, text='To :', font=FONT2)
        L.grid(row=1, column=0, padx=10, pady=5)
        v_dataTo = StringVar()
        Eto = ttk.Entry(DF, textvariable=v_dataTo, font=FONT2)
        Eto.grid(row=1, column=1, padx=10, pady=5)
        Eto.bind('<Double-1>', topopup)

        #button
        B = ttk.Button(self, text='Export', command=exportdata)
        B.pack(ipadx=5, ipady=5, pady=5)
    
    def startexporttoolingdata(self, query, bld_select, fromdate, todate):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังส่งออกข้อมูล...', font=FONT4)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExportDatatooling.Exporttooling(self, query, bld_select, fromdate, todate))
        thread.start()
        window.destroy()
        window.update()

    def Exporttooling(self, query, bld_select, fromdate, todate):
        self.dbtool = load_workbook(filename=self.toolingpath)
        self.dbtool_sheet = self.dbtool[self.toolingsheet]

        # Step 1: Define the start and stop dates
        start_date = datetime.strptime(fromdate, "%d/%b/%Y")
        stop_date = datetime.strptime(todate, "%d/%b/%Y")

        # Step 2: Generate the date range
        current_date = start_date
        date_range = []

        while current_date <= stop_date:
            date_range.append(current_date.strftime("%d/%b/%Y"))  # Add the formatted date to the list
            current_date += timedelta(days=1)  # Move to the next day
        
        if query == 'BLD':
            downtimedata = []
            for date in date_range:
                for row in self.dbtool_sheet.iter_rows(min_row=2, max_row=self.dbtool_sheet.max_row, min_col=0, max_col=12, values_only=True):
                    if row[7].lower() == bld_select.lower() and row[9] == date:
                        downtimedata.append(row)
        
        elif query == 'All':
            downtimedata = []
            for date in date_range:
                for row in self.dbtool_sheet.iter_rows(min_row=2, max_row=self.dbtool_sheet.max_row, min_col=0, max_col=12, values_only=True):
                    if row[9] == date:
                        downtimedata.append(row)

            # load template
            toolingtemplate = load_workbook(filename=self.tool_tem )
            toolingexport = toolingtemplate.active
            start_row = 2
            for d in downtimedata:
                toolingexport.cell(row=start_row,column=1).value=d[0]
                toolingexport.cell(row=start_row,column=2).value=d[1]
                toolingexport.cell(row=start_row,column=3).value=d[2]
                toolingexport.cell(row=start_row,column=4).value=d[3]
                toolingexport.cell(row=start_row,column=5).value=d[4]
                toolingexport.cell(row=start_row,column=6).value=d[5]
                toolingexport.cell(row=start_row,column=7).value=d[6]
                toolingexport.cell(row=start_row,column=8).value=d[7]
                toolingexport.cell(row=start_row,column=9).value=d[8]
                toolingexport.cell(row=start_row,column=10).value=d[9]
                toolingexport.cell(row=start_row,column=11).value=d[10]
                toolingexport.cell(row=start_row,column=12).value=d[11]
                
                start_row+=1
            toolingtemplate.save(f'{self.tool_export_path}/{self.tool_name }.xlsx')

            #excel popup open
            excelopen = '{}/{}.xlsx'.format(self.tool_export_path,self.tool_name)
            os.system('start "excel" "{}"'.format(excelopen))

class ExportDatasparepart(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=500, height=500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        #path
        self.spare_path = self.config['DATABASE']['dbsparepath']
        self.spare_sheet = self.config['DATABASE']['sparepartsheet']
       
        #template
        self.spare_tem = self.config['TEMPLATE']['spare']
        
        #exportpath
        self.spare_ex_path = self.config['EXPORTPATH']['spare']

        #exportdata pname
        self.spare_name = self.config['EXPORTNAME']['spare']

        def frompopup(event):
            MyDatePicker(widget=Efrom)

        def topopup(event):
            MyDatePicker(widget=Eto)

        #query
        def blank(event):
            query = v_query.get()
            if query == 'All':
                bld_select.config(state='disabled')
                
            else:
                bld_select.config(state='disabled')

        def exportdata():
            query = v_query.get()
            fromdate = v_dateForm.get()
            todate = v_dataTo.get()
            bld_select = v_bld.get()
            if query and fromdate and todate:
                ExportDatasparepart.startexportspare(self, query, bld_select, fromdate, todate)
            else:
                messagebox.showinfo('Export data', 'โปรดระบุข้อมูลให้ครบ')

        #mainframe
        MF = LabelFrame(self)
        MF.pack(padx=10, pady=10)

        #query frame
        QF = LabelFrame(MF, text='Query by', font=FONT3)
        QF.pack(padx=10, pady=10)
        L = Label(QF, text='Query By :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_query = StringVar()
        Es = ttk.Combobox(QF, textvariable=v_query, font=FONT2, width=17, values=['All'], state='readonly')
        Es.grid(row=0, column=1, padx=5, pady=5)
        Es.bind('<<ComboboxSelected>>', blank)

        #bld
        L = Label(QF, text='Building :', font=FONT2)
        L.grid(row=1, column=0, padx=5, pady=5)
        v_bld = StringVar()
        bld_select = ttk.Combobox(QF, textvariable=v_bld, font=FONT2, values=['BLD4#2','BLD5#10/1','BLD5#10/2','BLD6#15/1','BLD6#15/2'],state='readonly', width=18, )
        bld_select.grid(row=1, column=1, padx=5, pady=5)

        #date frame
        DF = LabelFrame(MF, text='Date', font=FONT3)
        DF.pack(padx=10, pady=10)

        #from
        L = Label(DF, text='From :', font=FONT2)
        L.grid(row=0, column=0, padx=10, pady=5)
        v_dateForm = StringVar()
        Efrom = ttk.Entry(DF, textvariable=v_dateForm, font=FONT2)
        Efrom.grid(row=0, column=1, padx=10, pady=5)
        Efrom.bind('<Double-1>', frompopup)

        #to
        L = Label(DF, text='To :', font=FONT2)
        L.grid(row=1, column=0, padx=10, pady=5)
        v_dataTo = StringVar()
        Eto = ttk.Entry(DF, textvariable=v_dataTo, font=FONT2)
        Eto.grid(row=1, column=1, padx=10, pady=5)
        Eto.bind('<Double-1>', topopup)

        #button
        B = ttk.Button(self, text='Export', command=exportdata)
        B.pack(ipadx=5, ipady=5, pady=5)

    def startexportspare(self, query, bld_select, fromdate, todate):
            def center_windows(w,h):
                ws = window.winfo_screenwidth() #screen width
                hs = window.winfo_screenheight() #screen height
                x = (ws/2) - (w/2)
                y = (hs/2) - (h/2)
                return f'{w}x{h}+{x:.0f}+{y:.0f}'
            
            window = Toplevel()
            win_size = center_windows(200,100)
            window.geometry(win_size)
            window.title('Progress functions')
            label = Label(window, text = 'กำลังส่งออกข้อมูล...', font=FONT4)
            label.pack()
            window.update()
            thread = threading.Thread(target = ExportDatasparepart.exportspare(self, query, bld_select, fromdate, todate))
            thread.start()
            window.destroy()
            window.update()

    def exportspare(self, query, bld_select, fromdate, todate):
        self.spare = load_workbook(filename=self.spare_path)
        self.spare_sheet = self.spare[self.spare_sheet]

        # Step 1: Define the start and stop dates
        start_date = datetime.strptime(fromdate, "%d/%b/%Y")
        stop_date = datetime.strptime(todate, "%d/%b/%Y")

        # Step 2: Generate the date range
        current_date = start_date
        date_range = []

        while current_date <= stop_date:
            date_range.append(current_date.strftime("%d/%b/%Y"))  # Add the formatted date to the list
            current_date += timedelta(days=1)  # Move to the next day
            
        if query == 'All':
            downtimedata = []
            for date in date_range:
                for row in self.spare_sheet.iter_rows(min_row=2, max_row=self.spare_sheet.max_row, min_col=0, max_col=22, values_only=True):
                    if row[15] == date:
                        downtimedata.append(row)
        
            # load template
            sparetemplate = load_workbook(filename=self.spare_tem )
            spareexport = sparetemplate.active
            start_row = 4
            for d in downtimedata:
                spareexport.cell(row=start_row,column=1).value=start_row-3
                spareexport.cell(row=start_row,column=2).value=d[1]
                spareexport.cell(row=start_row,column=3).value=d[2]
                spareexport.cell(row=start_row,column=4).value=d[3]
                spareexport.cell(row=start_row,column=5).value=d[4]
                spareexport.cell(row=start_row,column=6).value=d[5]
                spareexport.cell(row=start_row,column=7).value=d[6]
                spareexport.cell(row=start_row,column=8).value=d[7]
                spareexport.cell(row=start_row,column=9).value=d[8]
                spareexport.cell(row=start_row,column=10).value=d[9]
                spareexport.cell(row=start_row,column=11).value=d[10]
                spareexport.cell(row=start_row,column=12).value=str(d[11])
                spareexport.cell(row=start_row,column=13).value=d[12]
                spareexport.cell(row=start_row,column=14).value=d[13]
                spareexport.cell(row=start_row,column=15).value=d[14]
                    
                start_row+=1
            sparetemplate.save(f'{self.spare_ex_path}/{self.spare_name }.xlsx')

            #excel popup open
            excelopen = '{}/{}.xlsx'.format(self.spare_ex_path,self.spare_name)
            os.system('start "excel" "{}"'.format(excelopen))

        
# root = Tk()
# root.title('Export data')
# # root.geometry('1500x1500')
# # app = ExportDataStencil(root)
# app = ExportDatadowntime(root)
# app.pack()
# root.mainloop()


