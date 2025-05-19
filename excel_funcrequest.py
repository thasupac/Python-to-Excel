from openpyxl import load_workbook
import configparser
from tkinter import messagebox, Toplevel, Label
from datetime import datetime
import threading

FONT = ('Angsana New',18)

class RequestSpare:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.requestpath = self.config['DATABASE']['dbrequestpath']
        self.requestsheet = self.config['DATABASE']['requestlogsheet']

    ##--progress
    def StartRequest(self,tsid, serialnum, partname, machine, line, quantity, unit, requestby, comment):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = RequestSpare.Requestlog(self,tsid, serialnum, partname, machine, line, quantity, unit, requestby, comment))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Spare part', 'ลงทะเบียนข้อมูลสำเร็จ')

    def Requestlog(self,tsid, serialnum, partname, machine, line, quantity, unit, requestby, comment):

        self.excelfile = load_workbook(filename=self.requestpath)
        self.excelfile.active = self.excelfile[self.requestsheet]
        self.sheet = self.excelfile[self.requestsheet]
        rows = self.sheet.max_row
        date = datetime.now().strftime("%d/%b/%Y")
        try:
            self.sheet.cell(row=rows+1,column=1).value=int(tsid)
            self.sheet.cell(row=rows+1,column=2,value=serialnum)
            self.sheet.cell(row=rows+1,column=3,value=partname)
            self.sheet.cell(row=rows+1,column=4,value=machine)
            self.sheet.cell(row=rows+1,column=5,value=line)
            self.sheet.cell(row=rows+1,column=6,value=quantity)
            self.sheet.cell(row=rows+1,column=7,value=unit)
            self.sheet.cell(row=rows+1,column=8,value=date)
            self.sheet.cell(row=rows+1,column=9,value=requestby)
            self.sheet.cell(row=rows+1,column=10,value=comment)
            self.excelfile.save(self.requestpath)
           
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    def StartDeletelog(self, tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลบข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = RequestSpare.DeleteLog(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Spare part', 'ลบข้อมูลสำเร็จ')

    def DeleteLog(self, tsid):
        self.excelfile = load_workbook(filename=self.requestpath)
        self.excelfile.active = self.excelfile[self.requestsheet]
        self.sheet = self.excelfile[self.requestsheet]
        rows = self.sheet.max_row

        for i in range(2,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    self.sheet.delete_rows(i,1)
                    self.excelfile.save(self.requestpath)
                    
# A = RequestSpare()
# A.Requestlog(1,2,3,4,5,6,7,8,9)