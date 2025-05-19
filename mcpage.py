from tkinter import ttk,messagebox,filedialog,Frame,END,LabelFrame,Label,StringVar,Text,Toplevel,Menu,Scrollbar,Y,RIGHT
from datetime import datetime
from PIL import ImageTk
from openpyxl import load_workbook
from excel_function_mc import MachineManagE
import tkinter as tk
import subprocess
import PIL.Image
import configparser
import threading
exmc = MachineManagE()

#font size
FONT1 = ('Angsana New', 25, 'bold')
FONT2 = ('Angsana New', 18, )
FONT3 = ('Angsana New', 16, )
FONT4 = ('Angsana New', 12, )

#line value
line = ['BLD4#2', 'BLD5#10', 'BLD6#15/2']

###class 
class MCreg(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        ##--function save reg m/c
        def RegMC():
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            mc = v_mc.get()
            mceds = v_eqd.get()
            brand = v_brand.get()
            serial = v_serial.get()
            cano = v_ca.get()
            ora = v_ora.get()
            wi = v_widoc.get()
            form = v_recordForm.get()
            bld = v_line.get()
            service = EText.get('1.0', END).strip()
            m = monthly.get()
            m3 = threemonth.get()
            m6 = sixmonth.get()
            year = yearly.get()
            mdcc = monthlydcc.get()
            m3dcc = threemonthdcc.get()
            m6dcc = sixmonthdcc.get()
            yeardcc = yearlydcc.get()
            comment = ECom.get('1.0', END).strip()
            photo_path_check = v_photo_path.get()

            #check blank
            if mc !=('') and mceds !=('') and brand !=('') and serial !=('') and cano !=('') and ora !=('') and wi !=('') and form !=('') and bld !=('') :
                try:
                    exmc.StartmCReG(tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,service,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photo_path_check)
                    regmcreset()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to data: {e}")
            else:
                messagebox.showinfo('Reg M/C','โปรดกรอกข้อมูลให้ครบ')
        
        ##--function after save
        def regmcreset():
            v_mc.set('')
            v_eqd.set('')
            v_brand.set('')
            v_serial.set('')
            v_ca.set('')
            v_ora.set('')
            v_widoc.set('')
            v_recordForm.set('')
            v_line.set('')
            EText.delete('1.0', END)
            monthly.set('')
            threemonth.set('')
            sixmonth.set('')
            yearly.set('')
            monthlydcc.set('')
            threemonthdcc.set('')
            sixmonthdcc.set('')
            yearlydcc.set('')
            ECom.delete('1.0', END)
            v_photo_path.set('')
            monthly.set(0)
            threemonth.set(0)
            sixmonth.set(0)
            yearly.set(0)
            monthlydcc.set(0)
            threemonthdcc.set(0)
            sixmonthdcc.set(0)
            yearlydcc.set(0)
            photoRemove()
        
        ##insert photo
        def selectPhoto():
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            v_photo_path.set(file_path)

            ##
            try:
                image = PIL.Image.open(v_photo_path.get())
                image = image.resize((190, 140))
                photo = ImageTk.PhotoImage(image)
                photo_mc.config(image=photo)
                photo_mc.image = photo

            except Exception as e:
                messagebox.showinfo('Equipment', 'Failed to load photo {e}')
            
        def photoRemove():
            photo_mc.config(image=None)
            photo_mc.image = None
           
         #Reg machine
        MF = LabelFrame(self, width=1200, height=980)
        MF.pack(padx=0, pady=10)

        #general detail
        G = LabelFrame(MF, text='general',width=385, height=610)
        G.grid(row=0, column=0, padx=5, pady=5)
        #m/c reg label
        ##--M/C Name :
        ML = Label(G, text='Equipment name :', font=FONT2)
        ML.grid(row=0, column=0)
        v_mc = StringVar()
        ML = ttk.Entry(G, textvariable=v_mc, font=FONT2, width=25)
        ML.grid(row=0, column=1, padx=5, pady=2.5)

        L = Label(G, text='Equipment description :', font=FONT2)
        L.grid(row=1, column=0)
        v_eqd = StringVar()
        E = ttk.Entry(G, textvariable=v_eqd, font=FONT2, width=25)
        E.grid(row=1 ,column=1, pady=5, padx=5)

        L = Label(G, text='Brand/Model :', font=FONT2)
        L.grid(row=2, column=0)
        v_brand = StringVar()
        E = ttk.Entry(G, textvariable=v_brand, font=FONT2, width=25)
        E.grid(row=2, column=1, pady=2.5, padx=5)

        L = Label(G, text='Serial no :', font=FONT2)
        L.grid(row=3, column=0)
        v_serial = StringVar()
        E = ttk.Entry(G, textvariable=v_serial, font=FONT2, width=25)
        E.grid(row=3, column=1, pady=2.5, padx=5)

        L = Label(G, text='CA#no :', font=FONT2)
        L.grid(row=4, column=0)
        v_ca = StringVar()
        E = ttk.Entry(G, textvariable=v_ca, font=FONT2, width=25)
        E.grid(row=4, column=1, pady=2.5, padx=5)

        #m/c photo tap          
        MP = LabelFrame(G, text='Photo EQ', width=190, height=200, font=FONT2)          
        MP.grid(row=6, column=0, padx=10, pady=20)

        #m/c photo tap          
        SP = LabelFrame(G, text='Photo EQ', width=190, height=200, font=FONT2)          
        SP.grid(row=6, column=1, padx=10, pady=20)

        ##--path photo
        v_photo_path = StringVar()
        Path_photo = ttk.Entry(MP, textvariable=v_photo_path, font=FONT2)
        Path_photo.place(x=5, y=25, height=25, width=175)

        #m/c button photo machine
        BMP = ttk.Button(MP, text='Select photo', command=selectPhoto)   
        BMP.place(x=50, y=70)

        #general detail
        Fabrinet = LabelFrame(MF, text='detail',width=385, height=610)
        Fabrinet.grid(row=0, column=1, padx=2.5, pady=5)

        #ora
        L = Label(Fabrinet, text='ORA# :', font=FONT2)
        L.grid(row=0, column=0)
        v_ora = StringVar()
        E = ttk.Entry(Fabrinet, textvariable=v_ora, font=FONT2, width=25)
        E.grid(row=0 ,column=1, padx=30, pady=5)

        L = Label(Fabrinet, text='WI Number :', font=FONT2)
        L.grid(row=1, column=0)
        v_widoc = StringVar()
        E = ttk.Entry(Fabrinet, textvariable=v_widoc, font=FONT2, width=25)
        E.grid(row=1, column=1, padx=30, pady=5)

        L = Label(Fabrinet, text='Form record :', font=FONT2)
        L.grid(row=2, column=0)
        v_recordForm = StringVar()
        E = ttk.Entry(Fabrinet, textvariable=v_recordForm, font=FONT2, width=25)
        E.grid(row=2, column=1, padx=30, pady=5)

        L = Label(Fabrinet, text='Line :', font=FONT2)
        L.grid(row=3, column=0)
        v_line = StringVar()
        E = ttk.Combobox(Fabrinet, textvariable=v_line, font=FONT2, values=line, state='readonly', width=23)
        E.grid(row=3, column=1, padx=30, pady=5)

        L = Label(Fabrinet, text='Services :', font=FONT2)
        L.grid(row=4, column=0)
        EText = Text(Fabrinet, font=FONT2, width=25, height=3)
        EText.grid(row=4, column=1, pady=5, padx=5)

        L = Label(Fabrinet, text='', font=FONT2)
        L.grid(row=5, column=0, pady=95)

        #preventive maintenance pms
        PM = LabelFrame(MF, text='preventive maintenance',width=385, height=610)
        PM.grid(row=0, column=2, padx=5, pady=5)

        L = Label(PM, text='PMS', font=FONT2)
        L.grid(row=0, column=0, sticky='w')

        monthly = tk.IntVar()
        c1 = tk.Checkbutton(PM, text='Monthly',variable=monthly, onvalue=1, offvalue=0, font=FONT2)
        c1.grid(row=1, column=0)

        threemonth = tk.IntVar()
        c2 = tk.Checkbutton(PM, text='3Month',variable=threemonth, onvalue=1, offvalue=0, font=FONT2)
        c2.grid(row=2, column=0)
        
        sixmonth = tk.IntVar()
        c3 = tk.Checkbutton(PM, text='6Month',variable=sixmonth, onvalue=1, offvalue=0, font=FONT2)
        c3.grid(row=3, column=0)

        yearly = tk.IntVar()
        c3 = tk.Checkbutton(PM, text='Yearly',variable=yearly, onvalue=1, offvalue=0, font=FONT2)
        c3.grid(row=4, column=0)

        #DCC
        L = Label(PM, text='DCC', font=FONT2)
        L.grid(row=0, column=1, padx=10, sticky='w')

        L = Label(PM, text='', font=FONT2)
        L.grid(row=0, column=1, padx=30)

        monthlydcc = tk.IntVar()
        c1dcc = tk.Checkbutton(PM, text='Monthly',variable=monthlydcc, onvalue=1, offvalue=0, font=FONT2)
        c1dcc.grid(row=1, column=1, padx=20)

        threemonthdcc = tk.IntVar()
        c2dcc = tk.Checkbutton(PM, text='3Month',variable=threemonthdcc, onvalue=1, offvalue=0, font=FONT2)
        c2dcc.grid(row=2, column=1)
        
        sixmonthdcc = tk.IntVar()
        c3dcc = tk.Checkbutton(PM, text='6Month',variable=sixmonthdcc, onvalue=1, offvalue=0, font=FONT2)
        c3dcc.grid(row=3, column=1)

        yearlydcc = tk.IntVar()
        c4dcc = tk.Checkbutton(PM, text='Yearly',variable=yearlydcc, onvalue=1, offvalue=0, font=FONT2)
        c4dcc.grid(row=4, column=1)

        L = Label(PM, text='Comment', font=FONT2)
        L.grid(row=5, column=0)
        ECom = Text(PM, font=FONT2, width=10, height=3)
        ECom.grid(row=5, column=1, padx=10, pady=10)

        L = Label(PM, text='', font=FONT2)
        L.grid(row=7, column=1, pady=75)

        #m/c button reg machine
        BMS = ttk.Button(G, text='Save', command=RegMC)       #button machine save
        BMS.grid(row=9, column=1, sticky='e', padx=10)
        BMC = ttk.Button(G, text='Clear', command=regmcreset)      #button machine clear
        BMC.grid(row=9, column=1, pady=20, sticky='w')

        #photo machine
        photo_mc = tk.Label(SP)
        photo_mc.pack()

    ##--progress
    def StartmCSwitchPagE(self):
        def center_windows(self,w,h):
            ws = self.winfo_screenwidth() #screen width
            hs = self.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        win_size = center_windows(self,200,100)
        window = Toplevel()
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กรุณารอสักครู่...', font=FONT1)
        label.pack()
        window.update()
        thread = threading.Thread(target = MCreg.Switch(self))
        thread.start()
        window.destroy()
        window.update()
 
    def Switch(self):
        subprocess.Popen(["python", "ASMT_Store_Part.py"])
        self.quit()

###class view new
class MCview(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.machinepath = self.config['DATABASE']['dbmachinepath']
        self.machinesheet = self.config['DATABASE']['machinesheet']
        self.machinephotopath = self.config['DATABASE']['dbphotomachinepath']

        #search_type
        search_type = Label(self, text='Search type :', font=FONT2)
        search_type.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('EQ name')
        search_type_combo = ttk.Combobox(self , textvariable=v_search_type, values=['EQ name','ORA','Line','All'], state='readonly')
        search_type_combo.place(x=120, y=10, height=30)
        
        #function search
        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()     
            
            #mc search
            if check_search_type == 'EQ name':
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=20, values_only=True):
                        found = False
                        for col_idx in columns:
                            if col_idx <= len(row):
                                cell_value = str(row[1]).strip()  #EQname
                                if cell_value.lower().startswith(search_prefix.lower()):
                                    found = True
                                    break
                        if found:
                            results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[4]).strip(), str(row[6]).strip(), str(row[17]).strip()])
                    return results
                
            #ora search
            elif check_search_type == 'Line':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=1, max_col=20, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[17]).strip() #Line
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[4]).strip(), str(row[6]).strip(), str(row[17]).strip()])
                return results
            
            #line search
            elif check_search_type == 'ORA':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=1, max_col=20, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[6]).strip()  #ORA
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[4]).strip(), str(row[6]).strip(), str(row[17]).strip()])
                return results
            
            #all search
            elif check_search_type == 'All':
                
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=20, values_only=True):
                    results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[4]).strip(), str(row[6]).strip(), str(row[17]).strip()])
                return results
                
        def on_key_release(event):
            search_prefix = EFID.get().strip().lower()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [1], search_prefix)  # Adjust columns as needed
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def display_results(results):
            for iter in M_Clist.get_children():
                M_Clist.delete(iter)

            if results:
                for row in results:
                    #row = tuple(row)
                    M_Clist['columns']=("TSID","Equipment name", "Serial" , 'ORA#',"Line")
                    M_Clist.column('TSID', anchor="center", width=5)
                    M_Clist.column('Equipment name', anchor="center", width=5)
                    M_Clist.column('Serial', anchor="center", width=5)
                    M_Clist.column('ORA#', anchor="center", width=5)  
                    M_Clist.column('Line', anchor="center", width=5) 
                    ##heading
                    M_Clist.heading('TSID', text='TSID',anchor="center")
                    M_Clist.heading('Equipment name', text='Equipment name', anchor="center")
                    M_Clist.heading('Serial', text='Serial',anchor="center")
                    M_Clist.heading('ORA#', text='ORA#', anchor="center")   
                    M_Clist.heading('Line', text='Line', anchor="center")  
                    M_Clist.insert('', 'end', values=row)

            else:
                M_Clist['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                M_Clist.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                M_Clist.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")

            #Select data to edit
            def Select_mc(event=None):
                select = M_Clist.selection()
                output = M_Clist.item(select)

                if select !=():
                    tsid = output['values'][0]
                    wb = load_workbook(filename=file_path)
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=21, values_only=True):
                        if tsid ==  int(row[0]):  #TSID
                            listRow = list(row)
                            eqn = listRow[1]
                            eqd = listRow[2]
                            brand = listRow[3]
                            serial = listRow[4]
                            caNo = listRow[5]
                            ora = listRow[6]
                            wi = listRow[7]
                            form = listRow[8]

                            #pms
                            m = listRow[9]
                            m3 = listRow[10]
                            m6 = listRow[11]
                            y = listRow[12]

                            #dcc
                            mdcc = listRow[13]
                            m3dcc = listRow[14]
                            m6dcc = listRow[15]
                            ydcc = listRow[16]
                            
                            location = listRow[17]
                            service = listRow[18]
                            comment = listRow[19]
                            datereg = listRow[20]

                            GUIDetail = Toplevel()
                            GUIDetail.title('Equipment')
                            GUIDetail.geometry('1500x1500')

                            # #Update m/c
                            def Updatemc():
                                mc = v_mc.get()
                                mceds = v_eqd.get()
                                brand = v_brand.get()
                                serial = v_serial.get()
                                cano = v_ca.get()
                                ora = v_ora.get()
                                wi = v_widoc.get()
                                form = v_recordForm.get()
                                bld = v_line.get()
                                service = EText.get('1.0', END).strip()
                                m = monthly.get()
                                m3 = threemonth.get()
                                m6 = sixmonth.get()
                                year = yearly.get()
                                mdcc = monthlydcc.get()
                                m3dcc = threemonthdcc.get()
                                m6dcc = sixmonthdcc.get()
                                yeardcc = yearlydcc.get()
                                comment = ECom.get('1.0', END).strip()
                                photoupdate = f'{v_photo_path.get()}'
                                check = messagebox.askyesno('ยืนยันการอัปเดต','คุณต้องการอัปเดตข้อมูลใช่หรือไม่?')
                                if check == True:
                                    try:
                                        exmc.StartmCUpdatE(tsid,mc,mceds,brand,serial,cano,ora,wi,form,bld,service,m,m3,m6,year,mdcc,m3dcc,m6dcc,yeardcc,comment,photoupdate)
                                        display_results(results)
                                        on_key_release(event)
                                        close_top_level()
                                    except Exception as e:
                                        messagebox.showerror("Error", f"Failed to data: {e}")

                            ##insert photo
                            def selectPhoto():
                                LP.config(text= '')
                                LP.text = None
                                file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                                v_photo_path.set(file_path)
                                try:
                                    image = PIL.Image.open(v_photo_path.get())
                                    image = image.resize((190, 140))
                                    photo = ImageTk.PhotoImage(image)
                                    photo_mc.config(image=photo)
                                    photo_mc.image = photo
                                except Exception as e:
                                    messagebox.showinfo('Machine photo',f'Failed to load photo {e}')
                            
                            def close_top_level():
                                GUIDetail.destroy()
                                GUIDetail.update()

                            #=delete abd close gui
                            def delete_mc_guiclose(event=None):
                                select = M_Clist.selection()
                                output = M_Clist.item(select)
                                check = messagebox.askyesno('Equipment','Do you want to delete equipment?')
                                if check == True:
                                    if output['values'] != (''):
                                        tsid = output['values'][0]
                                        try:
                                            exmc.StartmCDeletE(tsid)
                                            display_results(results)
                                            on_key_release(event)
                                            GUIDetail.destroy()
                                            GUIDetail.update()

                                        except Exception as e:
                                            messagebox.showerror("Error", f"Failed to data: {e}")

                            MF = LabelFrame(GUIDetail, text=f'TSID :{tsid}',width=1200, height=980, font=FONT4)
                            MF.pack(padx=0, pady=10)

                            #general detail
                            G = LabelFrame(MF, text='general',width=385, height=610)
                            G.grid(row=0, column=0, padx=5, pady=5)
                            #m/c reg label
                            ##--M/C Name :
                            ML = Label(G, text='Equipment name :', font=FONT2)
                            ML.grid(row=0, column=0)
                            v_mc = StringVar()
                            v_mc.set(eqn)
                            ML = ttk.Entry(G, textvariable=v_mc, font=FONT2, width=25)
                            ML.grid(row=0, column=1, padx=5, pady=2.5)

                            L = Label(G, text='Equipment description :', font=FONT2)
                            L.grid(row=1, column=0)
                            v_eqd = StringVar()
                            v_eqd.set(eqd)
                            E = ttk.Entry(G, textvariable=v_eqd, font=FONT2, width=25)
                            E.grid(row=1 ,column=1, pady=5, padx=5)

                            L = Label(G, text='Brand/Model :', font=FONT2)
                            L.grid(row=2, column=0)
                            v_brand = StringVar()
                            v_brand.set(brand)
                            E = ttk.Entry(G, textvariable=v_brand, font=FONT2, width=25)
                            E.grid(row=2, column=1, pady=2.5, padx=5)

                            L = Label(G, text='Serial no :', font=FONT2)
                            L.grid(row=3, column=0)
                            v_serial = StringVar()
                            v_serial.set(serial)
                            E = ttk.Entry(G, textvariable=v_serial, font=FONT2, width=25)
                            E.grid(row=3, column=1, pady=2.5, padx=5)

                            L = Label(G, text='CA#no :', font=FONT2)
                            L.grid(row=4, column=0)
                            v_ca = StringVar()
                            v_ca.set(caNo)
                            E = ttk.Entry(G, textvariable=v_ca, font=FONT2, width=25)
                            E.grid(row=4, column=1, pady=2.5, padx=5)

                            #m/c photo tap          
                            MP = LabelFrame(G, text='Photo EQ', width=190, height=200, font=FONT2)          
                            MP.grid(row=6, column=0, padx=10, pady=20)

                            #m/c photo tap          
                            SP = LabelFrame(G, text='Photo EQ', width=190, height=200, font=FONT2)          
                            SP.grid(row=6, column=1, padx=10, pady=20)

                            #label สำหรับรูปภาพ
                            photo_mc = tk.Label(SP)
                            photo_mc.pack()

                            LP = tk.Label(SP)
                            LP.pack()
                            
                            try:
                                image = PIL.Image.open(f"{self.machinephotopath}\\{tsid}.png")
                                image = image.resize((190, 140))
                                photo = ImageTk.PhotoImage(image)
                                photo_mc.config(image=photo)
                                photo_mc.image = photo
                            except Exception as e:
                                LP.config(text='ไม่มีรูปภาพของ TSID นี้')
                               
                            ##--path photo
                            v_photo_path = StringVar()
                            Path_photo = ttk.Entry(MP, textvariable=v_photo_path, font=FONT2)
                            Path_photo.place(x=5, y=25, height=25, width=175)

                            #m/c button photo machine
                            BMP = ttk.Button(MP, text='Select photo', command=selectPhoto)   
                            BMP.place(x=50, y=70)

                            #general detail
                            Fabrinet = LabelFrame(MF, text='detail',width=385, height=610)
                            Fabrinet.grid(row=0, column=1, padx=2.5, pady=5)

                            #ora
                            L = Label(Fabrinet, text='ORA# :', font=FONT2)
                            L.grid(row=0, column=0)
                            v_ora = StringVar()
                            v_ora.set(ora)
                            E = ttk.Entry(Fabrinet, textvariable=v_ora, font=FONT2, width=25)
                            E.grid(row=0 ,column=1, padx=30, pady=5)

                            L = Label(Fabrinet, text='WI Number :', font=FONT2)
                            L.grid(row=1, column=0)
                            v_widoc = StringVar()
                            v_widoc.set(wi)
                            E = ttk.Entry(Fabrinet, textvariable=v_widoc, font=FONT2, width=25)
                            E.grid(row=1, column=1, padx=30, pady=5)

                            L = Label(Fabrinet, text='Form record :', font=FONT2)
                            L.grid(row=2, column=0)
                            v_recordForm = StringVar()
                            v_recordForm.set(form)
                            E = ttk.Entry(Fabrinet, textvariable=v_recordForm, font=FONT2, width=25)
                            E.grid(row=2, column=1, padx=30, pady=5)

                            L = Label(Fabrinet, text='Line :', font=FONT2)
                            L.grid(row=3, column=0)
                            v_line = StringVar()
                            v_line.set(location)
                            E = ttk.Combobox(Fabrinet, textvariable=v_line, font=FONT2, values=line, state='readonly', width=23)
                            E.grid(row=3, column=1, padx=30, pady=5)

                            L = Label(Fabrinet, text='Services :', font=FONT2)
                            L.grid(row=4, column=0)
                            EText = Text(Fabrinet, font=FONT2, width=25, height=3)
                            EText.grid(row=4, column=1, pady=5, padx=5)
                            EText.insert(END, str(service))

                            #register date
                            L = Label(Fabrinet, text='Reg date :', font=FONT2)
                            L.grid(row=5, column=0, pady=5)
                            v_regdate = StringVar()
                            v_regdate.set(datereg)
                            E = ttk.Entry(Fabrinet, textvariable=v_regdate, font=FONT2, state='readonly')
                            E.grid(row=5, column=1, padx=20, pady=5)

                            #entry
                            L = Label(Fabrinet, text='')
                            L.grid(row=6, column=0, pady=50)

                            #preventive maintenance pms
                            PM = LabelFrame(MF, text='preventive maintenance',width=385, height=610)
                            PM.grid(row=0, column=2, padx=5, pady=5)

                            L = Label(PM, text='PMS', font=FONT2)
                            L.grid(row=0, column=0, sticky='w')

                            monthly = tk.IntVar()
                            if m == 'Y':
                                monthly.set(1)
                            c1 = tk.Checkbutton(PM, text='Monthly',variable=monthly, onvalue=1, offvalue=0, font=FONT2)
                            c1.grid(row=1, column=0)

                            threemonth = tk.IntVar()
                            if m3 == 'Y':
                                threemonth.set(1)
                            c2 = tk.Checkbutton(PM, text='3Month',variable=threemonth, onvalue=1, offvalue=0, font=FONT2)
                            c2.grid(row=2, column=0)
                            
                            sixmonth = tk.IntVar()
                            if m6 == 'Y':
                                sixmonth.set(1)
                            c3 = tk.Checkbutton(PM, text='6Month',variable=sixmonth, onvalue=1, offvalue=0, font=FONT2)
                            c3.grid(row=3, column=0)

                            yearly = tk.IntVar()
                            if y == 'Y':
                                yearly.set(1)
                            c3 = tk.Checkbutton(PM, text='Yearly',variable=yearly, onvalue=1, offvalue=0, font=FONT2)
                            c3.grid(row=4, column=0)

                            #DCC
                            L = Label(PM, text='DCC', font=FONT2)
                            L.grid(row=0, column=1, padx=10, sticky='w')

                            L = Label(PM, text='', font=FONT2)
                            L.grid(row=0, column=1, padx=30)

                            monthlydcc = tk.IntVar()
                            if mdcc == 'Y':
                                monthlydcc.set(1)
                            c1dcc = tk.Checkbutton(PM, text='Monthly',variable=monthlydcc, onvalue=1, offvalue=0, font=FONT2)
                            c1dcc.grid(row=1, column=1, padx=20)

                            threemonthdcc = tk.IntVar()
                            if m3dcc == 'Y':
                                threemonthdcc.set(1)
                            c2dcc = tk.Checkbutton(PM, text='3Month',variable=threemonthdcc, onvalue=1, offvalue=0, font=FONT2)
                            c2dcc.grid(row=2, column=1)
                            
                            sixmonthdcc = tk.IntVar()
                            if m6dcc == 'Y':
                                sixmonthdcc.set(1)
                            c3dcc = tk.Checkbutton(PM, text='6Month',variable=sixmonthdcc, onvalue=1, offvalue=0, font=FONT2)
                            c3dcc.grid(row=3, column=1)

                            yearlydcc = tk.IntVar()
                            if ydcc == 'Y':
                                yearlydcc.set(1)
                            c4dcc = tk.Checkbutton(PM, text='Yearly',variable=yearlydcc, onvalue=1, offvalue=0, font=FONT2)
                            c4dcc.grid(row=4, column=1)

                            L = Label(PM, text='Comment', font=FONT2)
                            L.grid(row=5, column=0)
                            ECom = Text(PM, font=FONT2, width=10, height=3)
                            ECom.grid(row=5, column=1, padx=10, pady=10)
                            ECom.insert(END, str(comment))

                            L = Label(PM, text='', font=FONT2)
                            L.grid(row=7, column=1, pady=75)

                            #m/c button reg machine
                            BMS = ttk.Button(G, text='Update', command=Updatemc)       #button machine save
                            BMS.grid(row=9, column=1, sticky='e', padx=10)
                            BMC = ttk.Button(G, text='Delete', command=delete_mc_guiclose)      #button machine cleark
                            BMC.grid(row=9, column=1, pady=20, sticky='w')

                            GUIDetail.mainloop()

            M_Clist.bind('<Double-1>', Select_mc)

            #right click to delete and update
            def delete_mc(event=None):
                select = M_Clist.selection()
                output = M_Clist.item(select)
                check = messagebox.askyesno('Equipment','Do you want to delete equipment?')
                if check == True:
                    if output['values'] != (''):
                        tsid = output['values'][0]
                        try:
                            exmc.StartmCDeletE(tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")
                
            delete_mc_menu = Menu(self, tearoff=0)
            delete_mc_menu.add_command(label='delete', command=delete_mc)

            #select delete
            def popup(event):
                delete_mc_menu.post(event.x_root, event.y_root)
            M_Clist.bind('<Button-3>', popup)   #Button 3 คือ click ขวา

        def clear_results():
            for iter in M_Clist.get_children():
                M_Clist.delete(iter)

        # Define Excel data for searching
        file_path = self.machinepath
        sheet_name = self.machinesheet

        # #--label 
        # #--entry fixture ID
        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        EFID = ttk.Entry(self, font=FONT2)
        EFID.place(x=120, y=50, width=150, height=30)
        EFID.bind('<KeyRelease>', on_key_release)

        #create list machine
        header = ["TSID","Equipment name", "Serial" , 'ORA#',"Line"]
        headerw = [150,150,150,150,120]
        M_Clist = ttk.Treeview(self, columns=header, show='headings')
        M_Clist.place(x=20, y=90, width=1200, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',14))

        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")

        #--header and width
        for h,w in zip(header,headerw):
            M_Clist.heading(h,text=h)
            M_Clist.column(h,width=w,anchor='center')

        #scroll bar vertical
        M_Clist_Scroll = Scrollbar(M_Clist)
        M_Clist_Scroll.pack(side=RIGHT, fill=Y)
        M_Clist_Scroll.config(command=M_Clist.yview)
       
#เรียกมาใช้ก่อนน
# root = Tk()
# root.title('MC')
# root.geometry('1500x1500')
# app = MCview(root)
# app.pack()
# root.mainloop()


