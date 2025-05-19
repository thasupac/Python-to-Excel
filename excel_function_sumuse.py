from datetime import datetime
from openpyxl import load_workbook
from tkinter import *
import configparser

class Sumuse:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.toolpath = self.config['DATABASE']['dbtoolpath']
        self.toolsumuse = self.config['DATABASE']['summarysheet']


    def Tool(self,stencilNum,stencilNumsendcheck,type,quantityuse, status, qty_old_build):
        #loadworkbook
        self.excelfile = load_workbook(filename=self.toolpath)
        self.excelfile.active = self.excelfile[self.toolsumuse]
        self.sheet = self.excelfile[self.toolsumuse]
        rows = self.sheet.max_row
        #date
        date = datetime.now().strftime("%d/%b/%Y")
        duplicate_check = False
        rowstart = 2
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, min_col=1, max_col=10, values_only=True):
            if row[0].lower().replace(' ','').strip() == stencilNumsendcheck:

                try:
                    if self.sheet.cell(row=rowstart,column=3).value == None:
                        self.sheet.cell(row=rowstart,column=3).value = int(quantityuse)
                    else:
                        if status == 'not receive':
                            self.sheet.cell(row=rowstart,column=3).value += int(quantityuse)
                        elif status == 'receive':
                            self.sheet.cell(row=rowstart,column=3).value -= int(qty_old_build)
                            self.sheet.cell(row=rowstart,column=3).value += int(quantityuse)

                    self.sheet.cell(row=rowstart,column=4).value=date
                    self.excelfile.save(self.toolpath)
                    duplicate_check = True
                    break
                except Exception as e:
                    ()
            rowstart +=1
        while duplicate_check == False:
            self.sheet.cell(row=rows+1,column=1).value=stencilNum
            self.sheet.cell(row=rows+1,column=2).value=type
            self.sheet.cell(row=rows+1,column=3).value=quantityuse
            self.sheet.cell(row=rows+1,column=4).value=date
            self.excelfile.save(self.toolpath)
            
            break

        
