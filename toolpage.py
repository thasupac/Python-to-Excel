# from tkinter import *
from tkinter import ttk, messagebox,Frame,END,LabelFrame,Label,StringVar,Text,Toplevel,Menu,Scrollbar,RIGHT,Y,IntVar,NORMAL,filedialog, messagebox
from datetime import datetime
from PIL import Image, ImageTk
from help import Help
from openpyxl import load_workbook
import PIL.Image,threading,subprocess,configparser
import tkinter as tk

from excel_function_tool import ExcelB,ExcelF,ToolReg,stenCil
from excel_functionChecktooL import ToolCheck
from excel_function_sumuse import Sumuse
exbt = ExcelB()
exft = ExcelF()
exT = ToolReg()
exS = stenCil()
exCF = ToolCheck()
exSUM = Sumuse()

###---Font
FONT1 = ('Angsana New',25,'bold')
FONT2 = ('Angsana New',18)
FONT3 = ('Angsana New',12)

### technician
techname = ['Thanongsak_su','Don P','Somchai L','Adirek C','Sangworn D',
            'Pratchaya S','Supot P','Kriangsak H','Anan C',
            'Thanatorn K','Anong J','Thanongsak D','Apicha K',
            'Sompong L', 'Kritchanaphong K','Sarawut N','Narong L',
            'Surasak N','Narong K','Other']

###classical tooling
class_tool = ['Fixture','Bord profile','Stencil','PPtool','Fliptool',
              'Feeder','Vacumm base','Vernier','Vacuum cleaner','Hand Grease Pump',
              'Crimping Tool','Digital multimeter','Thermometer','Hex L-wrench','Nozzle','Ejector needle',
              'Flux plate','Flux gage','Screwdriver','Clamp meter','Squeegee','Wrench','Adapter']

#class register tool
class RegisterTool(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        #save reg
        def saveToolReg(event=None):
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            model = v_model.get()
            modelNum = v_model_number.get()
            modelNum_for_check = v_model_number.get().lower().replace(' ','').strip()
            customer = v_customer.get()
            clsaaTypes = v_types.get()
            clsaaTypes_check = v_types.get().lower()
            qty = v_qty.get()
            unit = v_unit.get()
            line = v_line.get()
            desc = E_desc.get('1.0',END).strip()
            regBy = v_reg_by.get()
            check_photo_save = v_photo_path.get()
            if model and modelNum and customer and clsaaTypes and qty and unit and line and regBy:
                try:
                    exCF.reg_check(tsid,model,modelNum,modelNum_for_check,customer,clsaaTypes,clsaaTypes_check,qty,unit,line,regBy,desc,check_photo_save)
                    # exT.StartToolReG(tsid,model,modelNum,customer,clsaaTypes,qty,unit,line,regBy,desc,check_photo_save)
                    reset()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to data: {e}")
                
            else:
                messagebox.showinfo('Tooling register', 'โปรดกรอกข้อมูลให้ครบถ้วน')

        def reset():
            v_model.set('')
            v_model_number.set('')
            v_customer.set('')
            v_types.set('')
            v_qty.set('')
            v_unit.set('')
            v_line.set('')
            E_desc.delete('1.0',END)
            v_reg_by.set('')
            v_photo_path.set('')
            photo_mc.config(image=None)
            photo_mc.image = None

        #select photo
        def selectPhoto():
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            v_photo_path.set(file_path)
            try:
                image = PIL.Image.open(v_photo_path.get())
                image = image.resize((190,190))
                photo = ImageTk.PhotoImage(image)
                photo_mc.config(image=photo)
                photo_mc.image = photo
            except Exception as e:
                messagebox.showerror('Register photo', f'Failed to load photo {e}')

        #mainframe
        F = LabelFrame(self, text='Reg tool', font=FONT3, width=800, height=500)
        F.pack()

        #maindetail
        MD = LabelFrame(F, width=450, height=500)
        MD.grid(row=0, column=0, padx=10, pady=10)

        # #model
        L = Label(MD, text='Model :', font=FONT2)
        L.grid(row=0, column=0, padx=30, pady=10)
        v_model = StringVar()
        E = ttk.Combobox(MD, textvariable=v_model, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Ciena','Other']), state='readonly')
        E.grid(row=0, column=1, padx=30, pady=10)

        # #model number
        L = Label(MD, text='Model number :', font=FONT2)
        L.grid(row=1, column=0, padx=30, pady=10)
        v_model_number = StringVar()
        E = ttk.Entry(MD, textvariable=v_model_number, font=FONT2, width=22)
        E.grid(row=1, column=1, padx=30, pady=10)

        # #customer
        L = Label(MD, text='Customer :', font=FONT2)
        L.grid(row=2, column=0, padx=30, pady=10)
        v_customer = StringVar()
        E = ttk.Combobox(MD, textvariable=v_customer, font=FONT2, values=(['Acacia','Nokia','Cisco','Lumentum','Ciena','Other']), state='readonly')
        E.grid(row=2, column=1, padx=30, pady=10)

        # #types
        L = Label(MD, text='Types :', font=FONT2)
        L.grid(row=3, column=0, padx=30, pady=10)
        v_types = StringVar()
        E = ttk.Combobox(MD, textvariable=v_types, font=FONT2, values=class_tool, state='readonly')
        E.grid(row=3, column=1, padx=30, pady=10)

        # #q'ty
        L = Label(MD, text="Q'ty (จำนวน Tooling) :", font=FONT2)
        L.grid(row=4, column=0, padx=30, pady=10)
        v_qty = StringVar()
        E = ttk.Entry(MD, textvariable=v_qty, font=FONT2)
        E.grid(row=4, column=1, padx=30, pady=10)

        # #unit
        L = Label(MD, text="Unit :", font=FONT2)
        L.grid(row=5, column=0, padx=30, pady=10)
        v_unit = StringVar()
        E = ttk.Combobox(MD, textvariable=v_unit, font=FONT2, values=(['ชิ้น','คู่',
                                                                         'ชุด']), state='readonly')
        E.grid(row=5, column=1, padx=30, pady=10)

        # #line
        L = Label(MD, text='Line :', font=FONT2)
        L.grid(row=6, column=0, padx=30, pady=10)
        v_line = StringVar()
        E = ttk.Combobox(MD, textvariable=v_line, font=FONT2, values=(['Bld4#2','Bld5#10','Bld6#15']), state='readonly')
        E.grid(row=6, column=1, padx=30, pady=10)

        #other frame
        OF = LabelFrame(F, width=450, height=500)
        OF.grid(row=0, column=1, padx=10, pady=10)

        # #descriptions
        L = Label(OF, text="Descriptions :", font=FONT2)
        L.grid(row=0, column=0, padx=30, pady=10)
        E_desc = Text(OF, font=FONT2, width=20, height=3)
        E_desc.grid(row=0, column=1, padx=30, pady=10)

        # #reg by
        L = Label(OF, text='Reg by :', font=FONT2)
        L.grid(row=1, column=0, padx=30, pady=10)
        v_reg_by = StringVar()
        E = ttk.Combobox(OF, textvariable=v_reg_by, font=FONT2, values=techname, state='readonly')
        E.grid(row=1, column=1, padx=30, pady=10)

        #photo frame
        F2 = LabelFrame(OF, text='Photo', font=FONT3, width=200, height=200)
        F2.grid(row=2, column=0, padx=10, pady=10)

        #photo frame show
        F2S = LabelFrame(OF, text='Photo', font=FONT3, width=200, height=200)
        F2S.grid(row=2, column=1, padx=10, pady=10)

        # #photo path
        v_photo_path = StringVar()
        E = ttk.Entry(F2, textvariable=v_photo_path, font=FONT2)
        E.grid(row=0, column=0, padx=20, pady=20)

        #photo label
        photo_mc = tk.Label(F2S)
        photo_mc.pack(padx=10, pady=10)

        #button
        B = ttk.Button(F2, text='Select', command=selectPhoto)
        B.grid(row=1, column=0, padx=20, pady=20)

        B2 = ttk.Button(self, text='Save', command=saveToolReg)
        B2.pack(padx=20 ,pady=2.5)

        B3 = ttk.Button(self, text='Clear', command=reset)
        B3.pack(padx=20 ,pady=2.5)

    #back to home
    ##--progress
    def StartmCSwitchPagE(self):
        def center_windows(self,w,h):
            ws = self.winfo_screenwidth() #screen width
            hs = self.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        win_size = center_windows(self,200,100)
        window = Toplevel()
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กรุณารอสักครู่...', font=FONT1)
        label.pack()
        window.update()
        thread = threading.Thread(target = RegisterTool.Switch(self))
        thread.start()
        window.destroy()
        window.update()
 
    def Switch(self):
        subprocess.Popen(["python", "ASMT_Store_Part.py"])
        self.quit()
#class views tool
class ViewsTool(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.toolpath = self.config['DATABASE']['dbtoolpath']
        self.toolsheet = self.config['DATABASE']['regtoolsheet']
        self.toolphotopath = self.config['DATABASE']['dbphototoolpath']

        #search type
        search_type = Label(self, text='Search type :', font=FONT2)
        search_type.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('All')
        search_type_combo = ttk.Combobox(self, values=['Model','Model number','Customer','Types','Line','All'], state='readonly', textvariable=v_search_type)
        search_type_combo.place(x=120, y=10, height=30)

        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()

            ##Model:
            if check_search_type == 'Model':    #S/N search
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[1]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip(),str(row[10]).strip()])
                return results  
            
            #model num
            elif check_search_type == 'Model number':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[2]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip(),str(row[10]).strip()])
                return results 
            
            #customer
            elif check_search_type == 'Customer':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[3]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip(),str(row[10]).strip()])
                return results
            
            #types
            elif check_search_type == 'Types':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[4]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip(),str(row[10]).strip()])
                return results
            
            #line
            elif check_search_type == 'Line':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[7]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip(),str(row[10]).strip()])
                return results
            
            #all (11 นับแบบ 0 1 2 3)
            elif check_search_type == 'All':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip(),str(row[10]).strip()])
                return results


        def on_key_release(event):
            search_prefix =  EFID.get().strip()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [1], search_prefix)  # Adjust columns as needed
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def display_results(results):
            for iter in toolList.get_children():
                toolList.delete(iter)
            if results:
                for row in results:
                    #row = tuple(row)
                    toolList['columns']=("TSID","Model", "Model num","Customer" , "Types","Q'ty","Unit", "Line","Register by","Reg date", "Descriptions")
                    toolList.column('TSID', anchor="center", width=2)
                    toolList.column('Model', anchor="center", width=2)
                    toolList.column('Model num', anchor="center", width=150) 
                    toolList.column('Customer', anchor="center", width=2) 
                    toolList.column('Types', anchor="center", width=2) 
                    toolList.column("Q'ty", anchor="center", width=2) 
                    toolList.column('Unit', anchor="center", width=2) 
                    toolList.column('Line', anchor="center", width=2) 
                    toolList.column('Register by', anchor="center", width=2) 
                    toolList.column('Reg date', anchor="center", width=2) 
                    toolList.column('Descriptions', anchor="center", width=2) 

                    ##heading
                    toolList.heading('TSID', text='TSID',anchor="center")
                    toolList.heading('Model', text='Model', anchor="center")
                    toolList.heading('Model num', text='Model num',anchor="center")
                    toolList.heading('Customer', text='Customer', anchor="center")   
                    toolList.heading('Types', text='Types', anchor="center")  
                    toolList.heading("Q'ty", text="Q'ty", anchor="center")  
                    toolList.heading('Unit', text='Unit', anchor="center")  
                    toolList.heading('Line', text='Line', anchor="center")  
                    toolList.heading('Register by', text='Register by', anchor="center")  
                    toolList.heading('Reg date', text='Reg date', anchor="center") 
                    toolList.heading('Descriptions', text='Descriptions', anchor="center")  
                    
                    toolList.insert('', 'end', values=row)

            else:
                toolList['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                toolList.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                toolList.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")

            def delete_right_click(event=None):
                select = toolList.selection()
                output = toolList.item(select)
                check = messagebox.askyesno('ยืนยันการลบ','คุณต้องการลบข้อมูล?...')
                if check == True:
                    if output['values'] != (''):
                        tsid = output['values'][0]

                        try:
                            exT.StartToolDeletE(tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")

            def popup(event):
                pop_up_menu.post(event.x_root, event.y_root)
            toolList.bind('<Button-3>', popup)   #Button 3 คือ click ขวา

            def Select_edit(event=None):
                select = toolList.selection()
                output = toolList.item(select)
                if select !=():
                    tsid = output['values'][0]
                    model = output['values'][1]
                    modelNum = output['values'][2]
                    customer = output['values'][3]
                    types = output['values'][4]
                    qty = output['values'][5]
                    unit = output['values'][6]
                    line = output['values'][7]
                    reg_by = output['values'][8]
                    reg_date = output['values'][9]
                    desc = output['values'][10]

                    toolViews = Toplevel()
                    toolViews.geometry('800x500-50+50')
                    toolViews.title('Tooling management')

                    #function update
                    def upDatetool():
                        U_model = v_model.get()
                        U_modelNum = v_modelNum.get()
                        U_customer = v_customer.get()
                        U_types = v_types.get()
                        U_qty = v_qty.get()
                        U_unit = v_unit.get()
                        U_line = v_line.get()
                        U_desc = v_desc.get() 
                        file_path_add = v_photo_path.get()
                        check = messagebox.askyesno('ยืนยันการอัปเดต','คุณต้องการอัปเดตข้อมูลใช่หรือไม่?')
                        
                        if check == True:
                            exT.StartToolUpdatE(tsid,U_model,U_modelNum,U_customer,U_types,U_qty,U_unit,U_line,U_desc,file_path_add)
                            display_results(results)
                            on_key_release(event)
                            closegui()

                    def closegui():
                        toolViews.destroy()
                        toolViews.update()
                    
                    #main frame
                    MF = LabelFrame(toolViews)
                    MF.pack(padx=10, pady=5)

                    #general frame
                    GD = LabelFrame(MF)
                    GD.grid(row=0, column=0, padx=10, pady=10)

                    #tsid
                    L = Label(GD, text=f'TSID : {tsid}', font=FONT3)
                    L.grid(row=0, column=0)

                    #model
                    L = Label(GD, text='Model :', font=FONT2)
                    L.grid(row=1, column=0, padx=20, pady=5, sticky='w')
                    v_model = StringVar()
                    v_model.set(model)
                    E = ttk.Entry(GD, textvariable=v_model, font=FONT2)
                    E.grid(row=1, column=1, padx=20, pady=5)

                    #model num
                    L = Label(GD, text='Model number :', font=FONT2)
                    L.grid(row=2, column=0, padx=20, pady=5, sticky='w')
                    v_modelNum = StringVar()
                    v_modelNum.set(modelNum)
                    E = ttk.Entry(GD, textvariable=v_modelNum, font=FONT2)
                    E.grid(row=2, column=1, padx=20, pady=5)

                    #customer
                    L = Label(GD, text='Customer :', font=FONT2)
                    L.grid(row=3, column=0, padx=20, pady=5, sticky='w')
                    v_customer = StringVar()
                    v_customer.set(customer)
                    E = ttk.Entry(GD, textvariable=v_customer, font=FONT2)
                    E.grid(row=3, column=1, padx=20, pady=5)

                    #types
                    L = Label(GD, text='Types :', font=FONT2)
                    L.grid(row=4, column=0, padx=20, pady=5, sticky='w')
                    v_types = StringVar()
                    v_types.set(types)
                    E = ttk.Entry(GD, textvariable=v_types, font=FONT2)
                    E.grid(row=4, column=1, padx=20, pady=5)

                    #qty
                    L = Label(GD, text="Q'ty :", font=FONT2)
                    L.grid(row=5, column=0, padx=20, pady=5)
                    v_qty = StringVar()
                    v_qty.set(qty)
                    E = ttk.Entry(GD, textvariable=v_qty, font=FONT2)
                    E.grid(row=5, column=1, padx=20, pady=5)

                    #unit
                    L = Label(GD, text='Unit :', font=FONT2)
                    L.grid(row=6, column=0, padx=20, pady=5)
                    v_unit = StringVar()
                    v_unit.set(unit)
                    E = ttk.Entry(GD, textvariable=v_unit, font=FONT2)
                    E.grid(row=6, column=1, padx=20, pady=5)

                    #line
                    L = Label(GD, text='Line :', font=FONT2)
                    L.grid(row=7, column=0, padx=20, pady=5)
                    v_line = StringVar()
                    v_line.set(line)
                    E = ttk.Entry(GD, textvariable=v_line, font=FONT2)
                    E.grid(row=7, column=1, padx=20, pady=5)

                    #regis by
                    L = Label(GD, text='Reg by:', font=FONT2)
                    L.grid(row=8, column=0, padx=20, pady=5)
                    v_regBy = StringVar()
                    v_regBy.set(reg_by)
                    E = ttk.Entry(GD, textvariable=v_regBy, font=FONT2, state='readonly')
                    E.grid(row=8, column=1, padx=20, pady=5)

                    #reg date
                    L = Label(GD, text='Reg date :', font=FONT2)
                    L.grid(row=9, column=0, padx=20, pady=10)
                    v_regDate = StringVar()
                    v_regDate.set(reg_date)
                    E = ttk.Entry(GD, textvariable=v_regDate, font=FONT2, state='readonly')
                    E.grid(row=9, column=1, padx=20, pady=5)

                    #ldescriptions
                    L = Label(GD, text='Descriptions :', font=FONT2)
                    L.grid(row=10, column=0, padx=20, pady=5)
                    v_desc = StringVar()
                    v_desc.set(desc)
                    E = ttk.Entry(GD, textvariable=v_desc, font=FONT2)
                    E.grid(row=10, column=1, padx=20, pady=5)

                    #frame photo
                    F = LabelFrame(MF, text='Photo', font=FONT3, width=350, height=350)
                    F.grid(row=0, column=1, padx=10, pady=10)

                    #buttom select photo
                    def add_photo():
                        file_path_add = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                        v_photo_path.set(file_path_add)
                        
                        try:
                            photo_mc = tk.Label(F)
                            photo_mc.pack()
                            image = PIL.Image.open(file_path_add)
                            image = image.resize((350, 350))
                            photo = ImageTk.PhotoImage(image)
                            photo_mc.config(image=photo)
                            photo_mc.image = photo
                        except Exception as e:
                            messagebox.showerror('Register photo', f'Failed to load photo {e}')

                    b_select = ttk.Button(MF, text='Add photo', command=add_photo)
                    b_select.grid(row=0, column=2, padx=10)

                    #photo path
                    v_photo_path = StringVar()
                    L_path = ttk.Entry(MF, textvariable=v_photo_path)
                    L_path.grid(row=0, column=3)

                    #try to previews
                    try:
                        photo_mc = tk.Label(F)
                        photo_mc.pack()
                        image = PIL.Image.open(f"{self.toolphotopath}\\{tsid}.png")
                        image = image.resize((350, 350))
                        photo = ImageTk.PhotoImage(image)
                        photo_mc.config(image=photo)
                        photo_mc.image = photo

                    except Exception as e:
                        L = Label(F, text='ไม่มีรูปภาพสำหรับ TSID นี้')
                        L.pack()
                        ()

                    #button update
                    B = ttk.Button(toolViews, text='Update', command=upDatetool)
                    B.pack(pady=2.5)

                    #button delete
                    B = ttk.Button(toolViews, text='Delete',  command=lambda:[delete_right_click(),closegui()])
                    B.pack(pady=2.5)
                    toolViews.mainloop()

            toolList.bind('<Double-1>', Select_edit)

            pop_up_menu = Menu(self, tearoff=0)
            pop_up_menu.add_command(label='delete', command=delete_right_click) 

        def clear_results():
            for iter in toolList.get_children():
                toolList.delete(iter)

        #defind path
        file_path = self.toolpath
        sheet_name = self.toolsheet

        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        EFID = ttk.Entry(self, font=FONT2)
        EFID.place(x=120, y=50, width=145, height=30)
        EFID.bind('<KeyRelease>',on_key_release)

         #New solution
        header = ["TSID","Model", "Model num","Customer" , "Types","Q'ty","Unit", "Line","Register by","Reg date", "Descriptions"]
        headerw = [10,10,100,10,10,10,10,10,10,10,10]

        # Create result display area
        toolList = ttk.Treeview(self, columns=header, show='headings')
        toolList.place(x=5, y=90, width=1250, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',14))

        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")

        #--header and width
        for h,w in zip(header,headerw):
            toolList.heading(h,text=h)
            toolList.column(h,width=w,anchor='center')

        #scroll bar vertical
        result_text_scroll = Scrollbar(toolList)
        result_text_scroll.pack(side=RIGHT, fill=Y)
        result_text_scroll.config(command=toolList.yview)

##class fixture withdraw
class fixWithdraw(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        ##--save function
        def save(event=None):
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            fixid_for_check = v_fixid.get().lower().replace(' ','').strip()  ##ทำให้เป็นตัวเล็กและตัดช่องว่างออก
            fixid = v_fixid.get()
            line = v_line.get()
            side = v_side.get()
            customer = v_customer.get()
            qty = v_qty.get()
            by = v_by.get()
            comment = Edesc.get('1.0', END)
            photofix = v_fixpathPhoto.get()

            # Generate Data to excel
            if fixid !=('') and line !=('') and side !=('') and customer !=('') and qty > (0) and by != (''):
                exCF.fixTure(tsid,fixid,side,line,customer,qty,by,fixid_for_check,comment,photofix)
                fixwireset()
                
            else:
                messagebox.showinfo('Fixture manage','โปรดกรอกข้อมูลให้ครบ') 
        def fixwireset():
            # reset
            v_fixid.set('')
            v_side.set('')
            v_line.set('')
            v_customer.set('')
            v_by.set('')
            v_qty.set('')
            Edesc.delete('1.0', END)
            v_fixpathPhoto.set('')
            photofix.config(image=None)
            photofix.image = None

        ##insert photo
        def fixselectphoto():
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            v_fixpathPhoto.set(file_path)
            ##
            try:
                image = PIL.Image.open(v_fixpathPhoto.get())
                image = image.resize((190, 140))
                photo = ImageTk.PhotoImage(image)
                photofix.config(image=photo)
                photofix.image = photo

            except Exception as e:
                messagebox.showinfo('Equipment', 'Failed to load photo {e}')
        
        GD = LabelFrame(self)
        GD.pack(padx=10, pady=10)

        #main frame
        MF = LabelFrame(GD, text='general detail', width=700, height=500)
        MF.grid(row=0, column=0, padx=10, pady=10)

        ##Fixture ID :
        L = Label(MF, text='Fixture ID :', font=FONT2)
        L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
        v_fixid = StringVar()
        E = ttk.Entry(MF, textvariable=v_fixid, font=FONT2)
        E.grid(row=0, column=1, padx=30, pady=10, sticky='w')
        E.bind('<Return>', lambda x:E.focus())

        ##--Side :
        L = Label(MF, text='Side :', font=FONT2)
        L.grid(row=1, column=0, padx=50, pady=10, sticky='e')
        v_side = StringVar()
        E = ttk.Combobox(MF, textvariable=v_side, state="readonly", values=("Top Side", "Bottom Side"), font=FONT2, width=18)
        E.grid(row=1, column=1, padx=30, pady=10, sticky='w')
        E.bind('<Return>', lambda x:E.focus())

        ##--Fixture customer :
        L = Label(MF, text='Fixture customer :', font=FONT2)
        L.grid(row=2, column=0, padx=50, pady=10, sticky='e')
        v_customer = StringVar()
        E = ttk.Combobox(MF, textvariable=v_customer, state='readonly',  values=(
                                                                            'Acacia',
                                                                            'Ciena',
                                                                            'Cisco',
                                                                            'Other'), font=FONT2, width=18)
        E.grid(row=2, column=1, padx=30, pady=10, sticky='w')
        E.bind('<Return>', lambda x:E.focus())

        ##--Line :
        L = Label(MF, text='Line :', font=FONT2)
        L.grid(row=3, column=0, padx=50, pady=10, sticky='e')
        v_line = StringVar()
        E = ttk.Combobox(MF, textvariable=v_line, state='readonly', values=(
                                                                        'BLD6_15/2_Line_1',
                                                                        'BLD6_15/2_Line_2',
                                                                        'BLD5_10_Line_1',
                                                                        'BLD5_10_Line_2',
                                                                        'BLD4'), font=FONT2, width=18)
        E.grid(row=3, column=1, padx=30, pady=10, sticky='w')
        E.bind('<Return>', lambda x:E.focus())

         ##--Quantity withdraw :
        L = Label(MF, text='Quantity withdraw :', font=FONT2)
        L.grid(row=4, column=0, padx=50, pady=10, sticky='e')
        v_qty = IntVar()
        E = ttk.Entry(MF, textvariable=v_qty, font=FONT2)
        E.grid(row=4, column=1, padx=30, pady=10, sticky='w')

        ##--By :
        L = Label(MF, text='By :', font=FONT2)
        L.grid(row=5, column=0, padx=50, pady=10, sticky='e')
        v_by = StringVar()
        E = ttk.Combobox(MF, textvariable=v_by, state='readonly', values=techname, font=FONT2, width=18)
        E.grid(row=5, column=1, padx=30, pady=10, sticky='w')
        E.bind('<Return>', lambda x:E.focus())

        #other frame
        OD = LabelFrame(GD, text='other details', width=700, height=500)
        OD.grid(row=0, column=1, padx=10, pady=10)

        #comment
        L = Label(OD, text='Descriptions :', font=FONT2)
        L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
        Edesc = Text(OD, width=20, height=5)
        Edesc.grid(row=0, column=1, padx=20, pady=10, sticky='e')

        #select photo
        SP = LabelFrame(OD, width=200, height=150)
        SP.grid(row=1, column=0, padx=10, pady=10)

        #pathphoto
        v_fixpathPhoto = StringVar()
        E = ttk.Entry(SP, textvariable=v_fixpathPhoto, font=FONT2)
        E.pack(padx=20, pady=20)

        #select photo button
        B = ttk.Button(SP, text='Select', command=fixselectphoto)
        B.pack(padx=10, pady=20)

        #show photo
        PS = LabelFrame(OD, width=200, height=150)
        PS.grid(row=1, column=1, padx=10, pady=10)

        #entry
        L = Label(OD, text='')
        L.grid(row=2, column=0, pady=30)

        # #save button
        B = ttk.Button(OD, text='Save', command=save)
        B.grid(row=2, column=1, sticky='w')

        # #clear 
        B2 = ttk.Button(OD, text='Clear', command=fixwireset)
        B2.grid(row=2, column=1, sticky='e', padx=10)

        #label photo
        photofix = Label(PS)
        photofix.pack()

##class fixture return
class fixPageReturn(Frame):
    def __init__(self,GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.fixturepath = self.config['DATABASE']['dbtoolpath']
        self.fixturesheet = self.config['DATABASE']['palletsheet']
        self.fixturephotopath = self.config['DATABASE']['dbphototoolpath']

        #define path
        file_path = self.fixturepath
        sheet_name = self.fixturesheet

        #search type
        search_type = Label(self, text='Search type :', font=FONT2)
        search_type.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('Fixture ID')
        search_type_combo = ttk.Combobox(self, values=['Fixture ID','Date','Customer','All'], state='readonly', textvariable=v_search_type)
        search_type_combo.place(x=120, y=10, height=30)  

        ##--def search dynamic
        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()

            #fixture id
            if check_search_type == 'Fixture ID':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=1, max_col=19, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[1]).strip()    #วิ่งตรวจ column ที่ 2
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), 
                                        str(row[3]).strip(), str(row[4]).strip(), str(row[5]).strip(), 
                                        str(row[6]).strip(), str(row[7]).strip(), str(row[18]).strip()])
                return results
            
            #date
            elif check_search_type == 'Date':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=1, max_col=19, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[7]).strip()    #วิ่งตรวจ column ที่ 7
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), 
                                        str(row[3]).strip(), str(row[4]).strip(), str(row[5]).strip(), 
                                        str(row[6]).strip(), str(row[7]).strip(), str(row[18]).strip()])
                return results
            
            #customer
            elif check_search_type == 'Customer':
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=1, max_col=19, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[4]).strip()    #วิ่งตรวจ column ที่ 4
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), 
                                        str(row[3]).strip(), str(row[4]).strip(), str(row[5]).strip(), 
                                        str(row[6]).strip(), str(row[7]).strip(), str(row[18]).strip()])
                return results
            
            #all 
            elif check_search_type == 'All':
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=19, values_only=True):
                    results.append([str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), 
                                    str(row[3]).strip(), str(row[4]).strip(), str(row[5]).strip(), 
                                    str(row[6]).strip(), str(row[7]).strip(), str(row[18]).strip()])
                return results
        
        def on_key_release(event):
            search_prefix = EFID.get().strip()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [1], search_prefix)  # 2 = fixtureID
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")
        
        def display_results(results):
            for iter in result_text.get_children():
                result_text.delete(iter)
        
            if results:
                for row in results: 
                    result_text['columns']=("tsid","Fixture ID", "Side","Line" , "Fixture customer" ,"Q'ty","By", "วันที่เอาออกมาใช้","สถานะ")
                    result_text.column("tsid", anchor="center", width=5)
                    result_text.column("Fixture ID", anchor="center", width=5)
                    result_text.column("Side", anchor="center", width=5)
                    result_text.column("Line", anchor="center", width=5)  
                    result_text.column("Fixture customer", anchor="center", width=5) 
                    result_text.column("Q'ty", anchor="center", width=5) 
                    result_text.column("By",anchor="center", width=5)
                    result_text.column("สถานะ", anchor="center", width=5)
                    result_text.column("วันที่เอาออกมาใช้", anchor="center", width=5)

                    #heading
                    result_text.heading("tsid", text="tsid", anchor="center")
                    result_text.heading("Fixture ID", text="Fixture ID", anchor="center")
                    result_text.heading("Side", text="Side",anchor="center")
                    result_text.heading("Line", text="Line", anchor="center")   
                    result_text.heading("Fixture customer", text="Fixture customer", anchor="center")  
                    result_text.heading("Q'ty",text="Q'ty", anchor="center") 
                    result_text.heading( "By", text="By",  anchor="center") 
                    result_text.heading("วันที่เอาออกมาใช้", text="วันที่เอาออกมาใช้", anchor="center") 
                    result_text.heading("สถานะ", text="สถานะ", anchor="center")
                    
                    result_text.insert('', 'end', values=row)

            else:
                result_text['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                result_text.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                result_text.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")

            def Select_edit(event=None):
                select = result_text.selection()
                output = result_text.item(select)

                if select != ():
                    #คำสั่ง selection เป็นข้อมูลแบบ dict {'text':data, 'image':data, 'values':data}
                    tsid = output['values'][0]  #values = key
                    wb = load_workbook(filename=file_path)
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=20, values_only=True):
                        if tsid ==  int(row[0]):
                            listRow = list(row)
                            f_id = listRow[1]
                            f_side = listRow[2]
                            f_line = listRow[3]
                            f_cus = listRow[4]
                            f_qty = listRow[5]
                            f_by = listRow[6]
                            f_date = listRow[7]
                            con1 = listRow[8]
                            con2 = listRow[9]
                            con3 = listRow[10]
                            con4 = listRow[11]
                            con5 = listRow[12]
                            con6 = listRow[13]
                            f_qtyre = listRow[14]
                            f_store = listRow[15]
                            f_datea = listRow[16]
                            f_note = listRow[17]
                            f_status = listRow[18]

                            #GUIpopup tool select
                            toolselect = Toplevel()
                            toolselect.title('Fixture detail')
                            toolselect.geometry('1200x550+50+50')

                            # #-save
                            def SaveReturn():
                                c1 = v_c1a.get()
                                c2 = v_c2a.get()
                                c3 = v_c3a.get()
                                c4 = v_c4a.get()
                                c5 = v_c5a.get()
                                c6 = v_c6a.get()
                                qty = v_qtyre.get()
                                receive = v_receive.get()
                                desc = EDesc.get('1.0', END)
                                check_path_photo = v_pathPhoto.get()

                                if c1 !=('') and c2 !=('') and c3 !=('') and c4 !=('') and c5 != ('') and c6 != ('') and qty !=('') and receive !=(''):

                                    try:
                                        exft.StartToolReturnwFixture(tsid,c1,c2,c3,c4,c5,c6,qty,receive,desc,check_path_photo)
                                        Reset()
                                        
                                    except Exception as e:
                                        messagebox.showerror("Error", f"Failed to data: {e}")

                                else:
                                    messagebox.showinfo('Fixture detail','โปรดกรอกข้อมูลให้ครบ')
                            
                            #close
                            def Close():
                                toolselect.destroy()
                                toolselect.update()
                                display_results(results)
                                on_key_release(event)

                            #-reset
                            def Reset():
                                v_c1a.set('')
                                v_c2a.set('')
                                v_c3a.set('')
                                v_c4a.set('')
                                v_c5a.set('')
                                v_c6a.set('')
                                v_qtyre.set('')
                                v_receive.set('')
                                EDesc.delete('1.0', END)
                                v_pathPhoto.set('')
                                Close()

                            MF = LabelFrame(toolselect)
                            MF.pack(padx=10, pady=10)

                            #eneral detail
                            GD = LabelFrame(MF, text=f'general detail TSID:{tsid}', font=FONT3, width=350, height=500)
                            GD.grid(row=0, column=0, padx=10, pady=10)

                            # ##fixture id
                            L = Label(GD, text='Fixture ID :', font=FONT2)
                            L.grid(row=0, column=0, padx=20, pady=10, sticky='e')
                            v_fid = StringVar()
                            v_fid.set(f_id)
                            E1 = ttk.Entry(GD, textvariable=v_fid, font=FONT2, state='readonly')
                            E1.grid(row=0, column=1, padx=10, pady=10, sticky='w')

                            # ##side
                            L = Label(GD, text='Side :', font=FONT2)
                            L.grid(row=1, column=0, padx=20, pady=10, sticky='e')
                            v_side = StringVar()
                            v_side.set(f_side)
                            E2 = ttk.Entry(GD, textvariable=v_side, font=FONT2, state='readonly')
                            E2.grid(row=1, column=1, padx=10, pady=10, sticky='w')

                             # #line
                            L = Label(GD, text='Line :', font=FONT2)
                            L.grid(row=2, column=0, padx=20 ,pady=10, sticky='e')
                            v_line = StringVar()
                            v_line.set(f_line)
                            E3 = ttk.Entry(GD, textvariable=v_line, font=FONT2, state='readonly')
                            E3.grid(row=2, column=1, padx=10, pady=10 ,sticky='w')

                            # #customer
                            L = Label(GD, text='Fixture customer :', font=FONT2)
                            L.grid(row=3, column=0, padx=20 ,pady=10, sticky='e')
                            v_cus = StringVar()
                            v_cus.set(f_cus)
                            E4 = ttk.Entry(GD, textvariable=v_cus, font=FONT2, state='readonly')
                            E4.grid(row=3, column=1, padx=10, pady=10, sticky='w')

                            # #QTY
                            L = Label(GD, text='Qty :', font=FONT2)
                            L.grid(row=4, column=0, padx=20 ,pady=10, sticky='e')
                            v_qty = StringVar()
                            v_qty.set(f_qty)
                            E5 = ttk.Entry(GD, textvariable=v_qty, font=FONT2, state='readonly')
                            E5.grid(row=4, column=1, padx=10, pady=10, sticky='w')

                            # #by 
                            L = Label(GD, text='Withdraw by :', font=FONT2)
                            L.grid(row=5, column=0, padx=20, pady=10, sticky='e')
                            v_by = StringVar()
                            v_by.set(f_by)
                            E6 = ttk.Entry(GD, textvariable=v_by, font=FONT2, state='readonly')
                            E6.grid(row=5, column=1, padx=10 ,pady=10, sticky='w')

                            # #date
                            L = Label(GD, text='Date of use :', font=FONT2)
                            L.grid(row=6, column=0, padx=20, pady=10, sticky='e')
                            v_date = StringVar()
                            v_date.set(f_date)
                            E7 = ttk.Entry(GD, textvariable=v_date, font=FONT2, state='readonly')
                            E7.grid(row=6, column=1, padx=10, pady=10, sticky='w')

                            #status
                            L = Label(GD, text='Status :', font=FONT2)
                            L.grid(row=7, column=0, padx=20, pady=10, sticky='e')
                            v_status = StringVar()
                            if f_status == 'not receive':
                                v_status.set(f_status)
                                E = ttk.Entry(GD, textvariable=v_status, font=FONT2, foreground='red', state='readonly')
                                E.grid(row=7, column=1, padx=10, pady=10, sticky='w')
                            else:
                                v_status.set(f_status)
                                E = ttk.Entry(GD, textvariable=v_status, font=FONT2, foreground='green', state='readonly')
                                E.grid(row=7, column=1, padx=10, pady=10, sticky='w')

                            #conditions
                            BC = LabelFrame(MF, text='conditions', font=FONT3, width=350, height=500)
                            BC.grid(row=0, column=1, padx=10, pady=10)

                            # #-c1
                            L = Label(BC, text='Clean/Check body of Fixture/Cover :',font=FONT2)
                            L.grid(row=0, column=0 ,padx=10, pady=10, sticky='e')
                            v_c1a = StringVar()
                            E = ttk.Combobox(BC, textvariable=v_c1a, state='readonly', values=['OK', 'NG'])
                            E.grid(row=0, column=1, padx=10, pady=10, sticky='e')
                            if con1 != None:
                                v_c1a.set(con1)

                            # #-c2
                            L = Label(BC, text='Retighten all screws :',font=FONT2)
                            L.grid(row=1, column=0, padx=10, pady=10, sticky='e')
                            v_c2a = StringVar()
                            E = ttk.Combobox(BC, textvariable=v_c2a, state='readonly', values=['OK', 'NG'])
                            E.grid(row=1, column=1, padx=10, pady=10, sticky='e')
                            if con2 != None:
                                v_c2a.set(con2)

                            # #-c3
                            L = Label(BC, text='Check all pins of Fixture :',font=FONT2)
                            L.grid(row=2, column=0, padx=10, pady=10, sticky='e')
                            v_c3a = StringVar()
                            E = ttk.Combobox(BC, textvariable=v_c3a, state='readonly', values=['OK', 'NG'])
                            E.grid(row=2, column=1, padx=10, pady=10, sticky='e')
                            if con3 != None:
                                v_c3a.set(con3)
                            
                            # #-c4
                            L = Label(BC, text='Check all Magnet of Fixture :',font=FONT2)
                            L.grid(row=3, column=0, padx=10, pady=10, sticky='e')
                            v_c4a = StringVar()
                            E = ttk.Combobox(BC, textvariable=v_c4a, state='readonly', values=['OK', 'NG'])
                            E.grid(row=3, column=1, padx=10, pady=10, sticky='e')
                            if con4 != None:
                                v_c4a.set(con4)
                            
                            # #-c5
                            L = Label(BC, text='Check hinge of Fixture :',font=FONT2)
                            L.grid(row=4, column=0, padx=10, pady=10, sticky='e')
                            v_c5a = StringVar()
                            E = ttk.Combobox(BC, textvariable=v_c5a, state='readonly', values=['OK', 'NG'])
                            E.grid(row=4, column=1, padx=10, pady=10, sticky='e')
                            if con5 != None:
                                v_c5a.set(con5)
                            
                            # #-c6
                            L = Label(BC, text='Check locking system of Fixture :',font=FONT2)
                            L.grid(row=5, column=0, padx=10, pady=10, sticky='e')
                            v_c6a = StringVar()
                            E = ttk.Combobox(BC, textvariable=v_c6a, state='readonly', values=['OK', 'NG'])
                            E.grid(row=5, column=1, padx=10, pady=10, sticky='e')
                            if con6 != None:
                                v_c6a.set(con6)

                            #descriptions
                            L = Label(BC, text='Descriptions :', font=FONT2)
                            L.grid(row=6, column=0, padx=10, pady=10, sticky='e')
                            EDesc = Text(BC, width=20, height=3)
                            EDesc.grid(row=6, column=1, padx=10, pady=10, sticky='e')
                            if f_note != None:
                                EDesc.insert(END, str(f_note))

                            #emtry
                            L = Label(BC, text='')
                            L.grid(row=7, column=0, pady=15)

                            #other detail
                            AC = LabelFrame(MF, text='other detail', font=FONT3, width=270, height=500)
                            AC.grid(row=0, column=2, padx=10, pady=10)

                            #qtyre
                            L = Label(AC, text="Q'ty", font=FONT2)
                            L.grid(row=0, column=0, padx=10, pady=1)
                            v_qtyre = StringVar()
                            if f_qtyre != None:
                                v_qtyre.set(f_qtyre)
                            E = ttk.Entry(AC, textvariable=v_qtyre, font=FONT2)
                            E.grid(row=0, column=1, padx=10, pady=10, sticky='w')

                            #receive
                            L = Label(AC, text='Receive by :', font=FONT2)
                            L.grid(row=1, column=0, padx=10, pady=10)
                            v_receive = StringVar()
                            if f_store != None:
                                v_receive.set(f_store)
                            E = ttk.Combobox(AC, textvariable=v_receive, font=FONT2, values=techname, state='readonly', width=18)
                            E.grid(row=1, column=1, padx=10, pady=10, sticky='w')

                            #date receive
                            if f_datea !=None:
                                L = Label(AC, text='Data receive :', font=FONT2)
                                L.grid(row=2, column=0, padx=10, pady=10)
                                v_datere = StringVar()
                                v_datere.set(f_datea)
                                E = ttk.Entry(AC, textvariable=v_datere, font=FONT2, state='readonly')
                                E.grid(row=2, column=1, padx=10, pady=10, sticky='w')
                            
                            # #-insert photo
                            L = LabelFrame(AC, text='Insert photo', width=200, height=150)
                            L.grid(row=3, column=0, padx=10, pady=10)
                            
                            #ปุ่มกดบันทึกการคืน
                            B = ttk.Button(AC, text='Save', command=SaveReturn)
                            B.grid(row=3, column=1, padx=10, pady=20)

                            #show photo
                            L2 = LabelFrame(AC, width=195, height=150)
                            L2.grid(row=4, column=0, padx=10, pady=10)

                            #photo label
                            photoFixture = tk.Label(L2)
                            photoFixture.pack(padx=10, pady=10)

                            v_pathPhoto = StringVar()
                            E = ttk.Entry(L, textvariable=v_pathPhoto, font=FONT2)
                            E.grid(row=0, column=0, padx=10, pady=10, sticky='w')

                            #select photo
                            def selectPhoto():
                                file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                                v_pathPhoto.set(file_path)
                                try:
                                    image = Image.open(v_pathPhoto.get())
                                    image = image.resize((180,170))
                                    photo = ImageTk.PhotoImage(image)
                                    photoFixture.config(image=photo)
                                    photoFixture.image = photo
                                except:
                                    ()
                            try:
                                image = Image.open(f"{self.fixturephotopath}\\{tsid}.png")
                                image = image.resize((180,170))
                                photo = ImageTk.PhotoImage(image)
                                photoFixture.config(image=photo)
                                photoFixture.image = photo
                            except Exception as e:
                                ()

                            #ปุ่มเลือกรูปภาพ
                            B = ttk.Button(L, text='select file', command=selectPhoto)
                            B.grid(row=1, column=0, pady=25)

                            toolselect.mainloop()

            result_text.bind('<Double-1>', Select_edit)

            #select delete
            def popup(event):
                delete_menu.post(event.x_root, event.y_root)
            result_text.bind('<Button-3>', popup)   #Button 3 คือ click ขวา

            #function from right click
            def fixturE_delete(event=None):
                select = result_text.selection()
                output = result_text.item(select)
                tsid = output['values'][0]
                check = messagebox.askyesno('ยืนยันการลบ','คุณต้องการลบข้อมูลใช่หรือไม่?')
                if check == True:
                    if output['values'] != (''):
                        try:
                            exft.StartToolDeleteFixture(tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")
            
            #delete menu
            delete_menu = Menu(self, tearoff=0)
            delete_menu.add_command(label='delete', command=fixturE_delete)

        def clear_results():
            for iter in result_text.get_children():
                result_text.delete(iter)
        
        #--label 
        #--entry fixture ID
        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        EFID = ttk.Entry(self, font=FONT2)
        EFID.place(x=120, y=50, width=150, height=30)
        EFID.bind('<KeyRelease>',on_key_release)

         #New solution
        header = ["tsid","Fixture ID", "Side","Line" , "Fixture customer" ,"Q'ty","By" "วันที่เอาออกมาใช้","สถานะ"]
        headerw = [20,150,10,50,20,5,8,8,10]

        # Create result display area
        result_text = ttk.Treeview(self, columns=header, show='headings')
        result_text.place(x=20, y=90, width=1200, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',14))

        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")

        #--header and width
        for h,w in zip(header,headerw):
            result_text.heading(h,text=h)
            result_text.column(h,width=w,anchor='center')

        #scroll bar vertical
        result_text_scroll = Scrollbar(result_text)
        result_text_scroll.pack(side=RIGHT, fill=Y)
        result_text_scroll.config(command=result_text.yview)

##class bord withdraw
class bordwith(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        #save functions
        def save_bord(event=None):
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            bordNumcheck = v_bordnum.get().lower().replace(' ','').strip()  ##ทำให้เป็นตัวเล็กและตัดช่องว่างออก
            model = v_profile_model.get()
            num_model = v_bordnum.get()
            side = v_side_bord.get()
            line = v_line_bord.get()
            with_by = v_withdraw_bord.get()
            physical = v_physical.get()
            fixture_status = v_fixture_bord.get()
            signal_status = v_signal.get()
            desc = descriptions.get('1.0',END).strip()
            photo = v_photoPathbord.get()
            if model and num_model and side and with_by and physical and fixture_status and signal_status and line!= (''):
                try:
                    # ff =
                    exCF.bordProfile(tsid,model,num_model,line,side,with_by,physical,fixture_status,signal_status,desc,bordNumcheck,photo)
                    reset()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to data: {e}")
            else:
                messagebox.showinfo('Bord profile','โปรดกรอกข้อมูลให้ครบ')

        def reset():
            v_profile_model.set('')
            v_bordnum.set('')
            v_side_bord.set('')
            v_line_bord.set('')
            v_withdraw_bord.set('')
            v_physical.set('')
            v_fixture_bord.set('')
            v_signal.set('')
            descriptions.delete('1.0',END)
            v_photoPathbord.set('')
            photoBord.config(image=None)
            photoBord.image = None

        #select photo
        def selectPhoto():
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            v_photoPathbord.set(file_path)
            ##
            try:
                image = Image.open(v_photoPathbord.get())
                image = image.resize((190, 140))
                photo = ImageTk.PhotoImage(image)
                photoBord.config(image=photo)
                photoBord.image = photo

            except Exception as e:
                messagebox.showinfo('Equipment', 'Failed to load photo {e}')


        #main frame
        MF = LabelFrame(self, width=700, height=600)
        MF.pack()

        #general frame
        GD = LabelFrame(MF, text='general detail', width=400, height=600)
        GD.grid(row=0, column=0, padx=10, pady=10)

        #--Model :
        L = Label(GD, text='Model :', font=FONT2)
        L.grid(row=0, column=0, padx=50, pady=10, sticky='w')
        v_profile_model = StringVar()
        E = ttk.Combobox(GD, textvariable=v_profile_model, state='readonly',values=['Acacia',
                                                                   'Acacia-CoW',
                                                                    'KLA',
                                                                    'Molex',
                                                                    'Cisco',
                                                                    'Nokia'], font=FONT2, width=18)
        E.grid(row=0, column=1, padx=30, pady=10, sticky='e')

        ##--Model number :
        L = Label(GD, text='Model number :', font=FONT2)
        L.grid(row=1, column=0, padx=50 ,pady=10, sticky='w')
        v_bordnum = StringVar()
        E = ttk.Entry(GD, textvariable=v_bordnum, font=FONT2)
        E.grid(row=1, column=1, padx=30, pady=10, sticky='e')

        ##--Side :
        L = Label(GD, text='Side :', font=FONT2)
        L.grid(row=2, column=0, padx=50, pady=10, sticky='w')
        v_side_bord = StringVar()
        E = ttk.Combobox(GD, textvariable=v_side_bord, values=['Top side','Bottom side'], state='readonly', font=FONT2, width=18)
        E.grid(row=2, column=1, padx=30, pady=10, sticky='e')

        ##--line
        L = Label(GD, text='Line :', font=FONT2)
        L.grid(row=3, column=0, padx=50, pady=10, sticky='w')
        v_line_bord = StringVar()
        E = ttk.Combobox(GD, textvariable=v_line_bord, values=['BLD4','BLD5-Cell1',
                                                                       'BLD5-Cell2','BLD6-Cell1',
                                                                       'BLD6-Cell2'], state='readonly', font=FONT2, width=18)
        E.grid(row=3, column=1, padx=30, pady=10, sticky='e')

        ##--By :
        L = Label(GD, text='By :', font=FONT2)
        L.grid(row=4, column=0, padx=50 ,pady=10, sticky='w')
        v_withdraw_bord = StringVar()
        E = ttk.Combobox(GD, textvariable=v_withdraw_bord, values=techname, state='readonly', font=FONT2, width=18)
        E.grid(row=4, column=1, padx=30, pady=10, sticky='e')

        #other detail
        OD = LabelFrame(MF, text='other detail', width=400, height=600)
        OD.grid(row=0, column=1, padx=10, pady=10)

        ##--ลักษณะทางกายภาพ :
        L = Label(OD, text='ลักษณะทางกายภาพ :', font=FONT2)
        L.grid(row=0, column=0, padx=50, pady=10, sticky='w')
        v_physical = StringVar()
        E = ttk.Combobox(OD, textvariable=v_physical, values=['Good','Not good','Hold'], state='readonly', font=FONT2, width=18)
        E.grid(row=0, column=1, padx=30, pady=10, sticky='e')

        ##--Fixture :
        L = Label(OD, text='Fixture :', font=FONT2)
        L.grid(row=1, column=0, padx=50, pady=10, sticky='w')
        v_fixture_bord = StringVar()
        E = ttk.Combobox(OD, textvariable=v_fixture_bord, values=['Good','Not good','Hold'], state='readonly', font=FONT2, width=18)
        E.grid(row=1, column=1, padx=30, pady=10, sticky='e')

        ##--Signal :
        L = Label(OD, text='Signal :', font=FONT2)
        L.grid(row=2, column=0, padx=50, pady=10, sticky='w')
        v_signal = StringVar()
        E = ttk.Combobox(OD, textvariable=v_signal, values=['Good','Not good','Hold'], state='readonly', font=FONT2, width=18)
        E.grid(row=2, column=1, padx=30, pady=10, sticky='e')

        #deas
        L = Label(OD, text='Descriptions :', font=FONT2)
        L.grid(row=3, column=0, padx=50, pady=10, sticky='w')
        descriptions = Text(OD, width=18, height=3)
        descriptions.grid(row=3, column=1, padx=30, pady=10, sticky='e')

        #selectphoto
        SP = LabelFrame(OD, width=200, height=150)
        SP.grid(row=4, column=0, padx=10, pady=10)

        #selectphoto
        PS = LabelFrame(OD, width=250, height=200)
        PS.grid(row=4, column=1, padx=10, pady=10)

        #label photo
        photoBord = Label(PS)
        photoBord.pack(padx=10, pady=10)

        #photopath
        v_photoPathbord = StringVar()
        E = ttk.Entry(SP, textvariable=v_photoPathbord)
        E.pack(padx=30, pady=10)

        #selectphoto
        B = ttk.Button(SP, text='Select', command=selectPhoto)
        B.pack(pady=10)

        #Button bord
        B = ttk.Button(self, text='Save', command=save_bord)
        B.pack(padx=10, pady=2.5)

        #Button bord
        B = ttk.Button(self, text='Clear', command=reset)
        B.pack(padx=10, pady=2.5)

##class bord return
class bordReturn(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        #define path
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.bordprofilepath = self.config['DATABASE']['dbtoolpath']
        self.bordprofilesheet = self.config['DATABASE']['bordprofilesheet']
        self.bordprofilephotopath = self.config['DATABASE']['dbphototoolpath']
        self.bordprofilesummary = self.config['DATABASE']['summarysheet']

        # Define Excel data for searching
        file_path = self.bordprofilepath
        sheet_name = self.bordprofilesheet

        #search_type
        search_type = Label(self, text='Search type :', font=FONT2)
        search_type.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('Model number')
        search_type_combo = ttk.Combobox(self, values=['Model','Model number','Line','Date','All'], state='readonly', textvariable=v_search_type)
        search_type_combo.place(x=120, y=10, height=30) 

        #function search
        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()    

            #mc search
            if check_search_type == 'Model':
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=19, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[1]).strip()  #colume 2 model  นับแบบ 0 1 2 3
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), 
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[8]).strip(),
                                        str(row[9]).strip(),str(row[15]).strip(),str(row[16]).strip()])
                return results
            
            elif check_search_type == 'Model number':
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=19, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[2]).strip()  #colume 3 model number  นับแบบ 0 1 2 3
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), 
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[8]).strip(),
                                        str(row[9]).strip(),str(row[15]).strip(),str(row[16]).strip()])
                return results
            
            elif check_search_type == 'Line':
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=19, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[3]).strip()  #colume 3 model number  นับแบบ 0 1 2 3
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), 
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[8]).strip(),
                                        str(row[9]).strip(),str(row[15]).strip(),str(row[16]).strip()])
                return results
            
            elif check_search_type == 'Date':
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=19, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[8]).strip()  #colume 3 model number  นับแบบ 0 1 2 3
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), 
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[8]).strip(),
                                        str(row[9]).strip(),str(row[15]).strip(),str(row[16]).strip()])
                return results
            
            elif check_search_type == 'All':
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=19, values_only=True):
                    results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), 
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[8]).strip(),
                                        str(row[9]).strip(),str(row[15]).strip(),str(row[16]).strip()])
                
                return results

        def on_key_release(event):
            search_prefix = EFID.get().strip()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [1], search_prefix)  # [1] เริ่มต้นไว้ที่ 1
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def display_results(results):
            for iter in bord_results.get_children():
                bord_results.delete(iter)

            if results:
                for row in results:
                    #row = tuple(row)
                    bord_results['columns']=('TSID','Model', 'Model number' ,'Line','Side','Date', 'Withdraw by', 'Deacription','สถานะ')
                    bord_results.column('TSID', anchor="center", width=5)
                    bord_results.column('Model', anchor="center", width=5)
                    bord_results.column('Model number', anchor="center", width=5)
                    bord_results.column('Line', anchor="center", width=5)  
                    bord_results.column('Side', anchor="center", width=5) 
                    bord_results.column('Date', anchor="center", width=5) 
                    bord_results.column('Withdraw by', anchor="center", width=5) 
                    bord_results.column('Deacription', anchor="center", width=5) 
                    bord_results.column('สถานะ', anchor="center", width=5) 

                    ##heading
                    bord_results.heading('TSID', text='TSID',anchor="center")
                    bord_results.heading('Model', text='Model', anchor="center")
                    bord_results.heading('Model number', text='Model number',anchor="center")
                    bord_results.heading('Line', text='Line', anchor="center")   
                    bord_results.heading('Side', text='Side', anchor="center") 
                    bord_results.heading('Date', text='Date', anchor="center") 
                    bord_results.heading('Withdraw by', text='Withdraw by', anchor="center") 
                    bord_results.heading('Deacription', text='Deacription', anchor="center") 
                    bord_results.heading('สถานะ', text='สถานะ', anchor="center")  
                    bord_results.insert('', 'end', values=row)

                    #select delete
                    def popup(event):
                        pop_up_menu.post(event.x_root, event.y_root)
                    bord_results.bind('<Button-3>', popup)   #Button 3 คือ click ข

            else:
                bord_results['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                bord_results.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                bord_results.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")

            def Select_edit(event=None):
                select = bord_results.selection()
                output = bord_results.item(select)
                if select != (''):
                    tsid = output['values'][0]  #values = key
                    wb = load_workbook(filename=file_path)
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=20, values_only=True):
                        if tsid ==  int(row[0]):
                            listRow = list(row)
                            b_model = listRow[1]
                            b_num = listRow[2]
                            b_line = listRow[3]
                            b_side = listRow[4]
                            b_c1b = listRow[5]
                            b_c2b = listRow[6]
                            b_c3b = listRow[7]
                            b_date = listRow[8]
                            b_by = listRow[9]
                            b_c1a = listRow[10]
                            b_c2a = listRow[11]
                            b_c3a = listRow[12]
                            b_datea = listRow[13]
                            b_reby = listRow[14]
                            b_desc = listRow[15]
                            b_status = listRow[16]

                            #GUIpopup tool select
                            toolselect = Toplevel()
                            toolselect.title('Bord profile detail')
                            toolselect.geometry('1300x600+0+50')

                            #save return
                            def rebordSave():
                                c1a = v_c1a.get()
                                c2a = v_c2a.get()
                                c3a = v_c3a.get()
                                reby = v_receive.get()
                                desc = EDes.get('1.0', END)  #กรณีมีการเปลี่ยน desc
                                photosave = v_pathPhotobord.get()

                                if c1a != ('') and c2a != ('') and c3a != ('') and reby != (''):
                                    try:
                                        exbt.StartToolDReturnBorD(tsid,c1a,c2a,c3a,reby,desc,photosave)
                                        exSUM.Tool(b_num, 'Bordprofile',1)
                                        reset()
                                    except Exception as e:
                                        messagebox.showerror("Error", f"Failed to data: {e}")
                                else:
                                    messagebox.showinfo('Bord return', 'โปรดกรอกข้อมูลให้ครบ')
                            
                            def reset():
                                toolselect.destroy()
                                toolselect.update()
                                display_results(results)
                                on_key_release(event)

                            def selectPhoto():
                                file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                                v_pathPhotobord.set(file_path)
                                ##
                                try:
                                    image = Image.open(v_pathPhotobord.get())
                                    image = image.resize((190, 140))
                                    photo = ImageTk.PhotoImage(image)
                                    photoBord.config(image=photo)
                                    photoBord.image = photo

                                except Exception as e:
                                    ()

                            #main frame
                            MF = LabelFrame(toolselect)
                            MF.pack(padx=10, pady=10)

                            ##detail
                            GD = LabelFrame(MF, text=f'TSID : {tsid}', width=400, height=500, font=FONT3)
                            GD.grid(row=0, column=0, padx=10, pady=10)

                            ##model
                            L = Label(GD, text='Model :', font=FONT2)
                            L.grid(row=0, column=0, padx=40, pady=10, sticky='e')
                            v_model = StringVar()
                            v_model.set(b_model)
                            E = ttk.Entry(GD, textvariable=v_model, font=FONT2, state='readonly')
                            E.grid(row=0 ,column=1, padx=20, pady=10, sticky='w')

                            ##model num
                            L = Label(GD, text='Model num :', font=FONT2)
                            L.grid(row=1, column=0, padx=40, pady=10, sticky='e')
                            v_modelnum = StringVar()
                            v_modelnum.set(b_num)
                            E = ttk.Entry(GD, textvariable=v_modelnum, font=FONT2, state='readonly')
                            E.grid(row=1, column=1, padx=20, pady=10, sticky='w')

                            ##line
                            L = Label(GD, text='Line :', font=FONT2)
                            L.grid(row=2, column=0, padx=40, pady=10, sticky='e')
                            v_line = StringVar()
                            v_line.set(b_line)
                            E = ttk.Entry(GD, textvariable=v_line, font=FONT2, state='readonly')
                            E.grid(row=2, column=1, padx=20, pady=10, sticky='w')

                            ##side
                            L = Label(GD, text='Side :', font=FONT2)
                            L.grid(row=3, column=0, padx=40, pady=10, sticky='e')
                            v_side = StringVar()
                            v_side.set(b_side)
                            E = ttk.Entry(GD, textvariable=v_side, font=FONT2, state='readonly')
                            E.grid(row=3, column=1, padx=20, pady=10, sticky='w')

                            ##witggraw by
                            L = Label(GD, text='Withdraw by :', font=FONT2)
                            L.grid(row=4, column=0, padx=40, pady=10, sticky='e')
                            v_withdraw = StringVar()
                            v_withdraw.set(b_by)
                            E = ttk.Entry(GD, textvariable=v_withdraw, font=FONT2, state='readonly')
                            E.grid(row=4, column=1, padx=20, pady=10, sticky='w')

                            ##date
                            L = Label(GD, text='Date withdraw :', font=FONT2)
                            L.grid(row=5, column=0, padx=40, pady=10, sticky='e')
                            v_date = StringVar()
                            v_date.set(b_date)
                            E = ttk.Entry(GD, textvariable=v_date, font=FONT2, state='readonly')
                            E.grid(row=5, column=1, padx=20, pady=10, sticky='w')

                            #สรุปการใช้
                            sumBord_path = self.bordprofilepath
                            sumBord_sheet = self.bordprofilesummary  #spare part (reg)
                            #loadworkbook
                            excelfile = load_workbook(filename=sumBord_path)
                            excelfile.active = excelfile[sumBord_sheet]
                            sheet = excelfile[sumBord_sheet]
                            currentUse = 0

                            ##currentUse
                            L = Label(GD, text='Current use :', font=FONT2)
                            L.grid(row=6, column=0, padx=40, pady=10, sticky='e')
                            v_currentuse = IntVar()
                            v_currentuse.set(currentUse)
                            E = ttk.Entry(GD, textvariable=v_currentuse, font=FONT2, state='readonly')
                            E.grid(row=6, column=1)
                           
                            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=10, values_only=True):
                                if row[0].lower() == b_num.lower():
                                    currentUse = row[2]
                                    v_currentuse = IntVar()
                                    v_currentuse.set(currentUse)
                                    E = ttk.Entry(GD, textvariable=v_currentuse, font=FONT2, state='readonly')
                                    E.grid(row=6, column=1, padx=10, pady=10)
                                    break

                            ##status
                            L = Label(GD, text='Status :', font=FONT2)
                            L.grid(row=7, column=0, padx=40, pady=10, sticky='e')
                            v_status = StringVar()
                            v_status.set(b_status)
                            if b_status == 'not receive':
                                E = ttk.Entry(GD, textvariable=v_status, font=FONT2, foreground='red', state='readonly')
                                E.grid(row=7, column=1, padx=20, pady=10, sticky='w')
                            else:
                                E = ttk.Entry(GD, textvariable=v_status, font=FONT2, foreground='green', state='readonly')
                                E.grid(row=7, column=1, padx=20, pady=10, sticky='w')

                            #condition before
                            CB = LabelFrame(MF, text='Conditionns before', width=400, height=500, font=FONT3)
                            CB.grid(row=0, column=1)

                            #con1
                            L = Label(CB, text='ลักษณะทางกายภาพ :', font=FONT2)
                            L.grid(row=0, column=0, padx=20, pady=10, sticky='e')
                            v_c1b = StringVar()
                            v_c1b.set(b_c1b)
                            E = ttk.Combobox(CB, textvariable=v_c1b, font=FONT2, state='readonly')
                            E.grid(row=0, column=1, padx=10, pady=10, sticky='w')

                            #con2
                            L = Label(CB, text='Fixture :')
                            L.grid(row=1, column=0, padx=20, pady=10, sticky='e')
                            v_c2b = StringVar()
                            v_c2b.set(b_c2b)
                            E = ttk.Combobox(CB, textvariable=v_c2b, font=FONT2, state='readonly')
                            E.grid(row=1, column=1, padx=10, pady=10, sticky='w')

                            #con3
                            L = Label(CB, text='Signal :', font=FONT2)
                            L.grid(row=2, column=0, padx=20, pady=10, sticky='e')
                            v_c3b = StringVar()
                            v_c3b.set(b_c3b)
                            E = ttk.Combobox(CB, textvariable=v_c3b, font=FONT2, state='readonly')
                            E.grid(row=2, column=1, padx=10, pady=10 ,sticky='w')

                            #photo select
                            SP = LabelFrame(CB, width=180, height=180)
                            SP.grid(row=3, column=0, padx=10, pady=10)

                            #photo sshow
                            PS = LabelFrame(CB, width=180, height=180)
                            PS.grid(row=3, column=1, padx=10, pady=10)

                            #desc
                            ##descriptions
                            L = Label(CB, text='Descriptions :', font=FONT2)
                            L.grid(row=4, column=0, padx=40, pady=10, sticky='e')
                            EDes = Text(CB, font=FONT2, width=20, height=3)
                            EDes.grid(row=4, column=1, padx=20, pady=10, sticky='w')
                            EDes.config(state=NORMAL)
                            if b_desc != None:
                                EDes.insert(END, str(b_desc))

                             #emtry
                            L = Label(CB, text='', font=FONT2)
                            L.grid(row=5, column=0, pady=15)

                            ##return detail after
                            RD = LabelFrame(MF, text='Conditions after', width=400, height=500, font=FONT3)
                            RD.grid(row=0, column=2, padx=10, pady=10)

                            #con1
                            L = Label(RD, text='ลักษณะทางกายภาพ :', font=FONT2)
                            L.grid(row=0, column=0, padx=20, pady=10, sticky='e')
                            v_c1a = StringVar()
                            if b_c1a != None:
                                v_c1a.set(b_c1a)
                            E = ttk.Combobox(RD, textvariable=v_c1a, font=FONT2, values=['OK','NG'], state='readonly')
                            E.grid(row=0, column=1, padx=10, pady=10, sticky='w')

                            #con2
                            L = Label(RD, text='Fixture :')
                            L.grid(row=1, column=0, padx=20, pady=10, sticky='e')
                            v_c2a = StringVar()
                            if b_c2a != None:
                                v_c2a.set(b_c2a)
                            E = ttk.Combobox(RD, textvariable=v_c2a, font=FONT2, values=['OK','NG'], state='readonly')
                            E.grid(row=1, column=1, padx=10, pady=10, sticky='w')

                            #con3
                            L = Label(RD, text='Signal :', font=FONT2)
                            L.grid(row=2, column=0, padx=20, pady=10, sticky='e')
                            v_c3a = StringVar()
                            if b_c3a != None:
                                v_c3a.set(b_c3a)
                            E = ttk.Combobox(RD, textvariable=v_c3a, font=FONT2, values=['OK','NG'], state='readonly')
                            E.grid(row=2, column=1, padx=10, pady=10 ,sticky='w')

                            #receive by
                            L = Label(RD, text='Receive by :', font=FONT2)
                            L.grid(row=3, column=0, padx=20, pady=10, sticky='e')
                            v_receive = StringVar()
                            if b_reby != None:
                                v_receive.set(b_reby)
                            E = ttk.Combobox(RD, textvariable=v_receive, values=techname, font=FONT2, state='readonly')
                            E.grid(row=3, column=1, padx=10, pady=10, sticky='w')

                            #date of receive
                            if b_datea != None:
                                L = Label(RD, text='Date receive :', font=FONT2)
                                L.grid(row=4, column=0, padx=20, pady=10, sticky='e')
                                v_data_receive = StringVar()
                                v_data_receive.set(b_datea)
                                E = ttk.Entry(RD, textvariable=v_data_receive, font=FONT2, state='readonly')
                                E.grid(row=4, column=1, padx=10, pady=10, sticky='w')

                            #photo select
                            SP = LabelFrame(RD, width=180, height=180)
                            SP.grid(row=5, column=0, padx=10, pady=10)

                            #photopath
                            v_pathPhotobord = StringVar()
                            E = ttk.Entry(SP, textvariable=v_pathPhotobord, font=FONT2)
                            E.place(x=5, y=30, height=30)

                            #button
                            B = ttk.Button(SP, text='Select photo', command=selectPhoto)
                            B.place(x=50, y=75)

                            #emtry
                            L = Label(RD, text='', font=FONT2)
                            L.grid(row=6, column=0, pady=10)

                            #photo sshow
                            PS = LabelFrame(RD, width=180, height=180)
                            PS.grid(row=5, column=1, padx=10, pady=10)

                            #label photo
                            photoBord = Label(PS)
                            photoBord.pack(padx=10, pady=10)

                            try:
                                image = Image.open(f"{self.bordprofilephotopath}\\{tsid}.png")
                                image = image.resize((180,170))
                                photo = ImageTk.PhotoImage(image)
                                photoBord.config(image=photo)
                                photoBord.image = photo
                            except Exception as e:
                                ()

                            ##button
                            B = ttk.Button(RD, text='Save', command=rebordSave)
                            B.grid(row=6, column=1, padx=20, pady=20)

                            toolselect.mainloop()

            bord_results.bind('<Double-1>', Select_edit)
            
            #RIGHT CLICK DELETE
            def delete_right_click(event=None):
                select = bord_results.selection()
                output = bord_results.item(select)
                tsid = output['values'][0]
                check = messagebox.askyesno('ยืนยันการลบ','คุณต้องการลบข้อมูล?...')
                if check == True:
                    if output['values'] != (''):
                        try:
                            exbt.StartToolDeletEBorD(tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")
                                            
            pop_up_menu = Menu(self, tearoff=0)
            pop_up_menu.add_command(label='delete', command=delete_right_click)

        def clear_results():
            for iter in bord_results.get_children():
                bord_results.delete(iter)

        # #--label 
        # #--entry fixture ID
        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        EFID = ttk.Entry(self, font=FONT2)
        EFID.place(x=120, y=50, width=150, height=30)
        EFID.bind('<KeyRelease>', on_key_release)

        #create list machine
        header = ['TSID','Model', 'Model number' ,'Line','Side','Date', 'Withdraw by', 'Deacription','สถานะ']
        headerw = [150,150,150,150,120,100,100,100,100]
        bord_results = ttk.Treeview(self, columns=header, show='headings')
        bord_results.place(x=20, y=90, width=1200, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',14))

        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")

        #--header and width
        for h,w in zip(header,headerw):
            bord_results.heading(h,text=h)
            bord_results.column(h,width=w,anchor='center')

        #scroll bar vertical
        bord_results_Scroll = Scrollbar(bord_results)
        bord_results_Scroll.pack(side=RIGHT, fill=Y)
        bord_results_Scroll.config(command=bord_results.yview)

##class stencil
class stenCilwith(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        #save withdraw
        def stenCilWith():
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            stencilCheck = v_stenNum.get().lower().replace(' ','').strip()  ##ทำให้เป็นตัวเล็กและตัดช่องว่างออก
            date = datetime.now().strftime("%d/%b/%Y")
            stenNum = v_stenNum.get()
            line = v_line.get()
            slotNum = v_slotNum.get()
            by = v_By.get()
            comment = Edesc.get('1.0', END)
            photoStencilSave = v_photoStencil.get()
            if tsid and stenNum and line and slotNum and by and date:
                try:
                    exCF.stenCil(tsid,stenNum,line,slotNum,by,date,stencilCheck,photoStencilSave,comment)
                    reset()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to data: {e}")
            else:
                messagebox.showinfo('Stencil','โปรดกรอกข้อมูลให้ครบ')
        def reset():
            v_stenNum.set('')
            v_line.set('')
            v_slotNum.set('')
            v_By.set('')
            Edesc.delete('1.0', END)
            v_photoStencil.set('')
            photoStencil.config(image=None)
            photoStencil.image = None

        #select photo
        def selectPhoto():
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            v_photoStencil.set(file_path)
            ##
            if v_photoStencil.get() != (''):
                try:
                    image = Image.open(v_photoStencil.get())
                    image = image.resize((190, 140))
                    photo = ImageTk.PhotoImage(image)
                    photoStencil.config(image=photo)
                    photoStencil.image = photo

                except Exception as e:
                    messagebox.showinfo('Equipment', f'Failed to load photo {e}')

        #frame
        MF = LabelFrame(self, text='Withdraw', font=FONT3, width=700, height=500)
        MF.pack()

        #general frame
        GD = LabelFrame(MF, text='general detail', width=300, height=400)
        GD.grid(row=0, column=0, padx=10, pady=10)

        #stencil num
        L = Label(GD, text='Stencil P/N :', font=FONT2)
        L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
        v_stenNum = StringVar()
        E = ttk.Entry(GD, textvariable=v_stenNum, font=FONT2)
        E.grid(row=0, column=1, padx=30, pady=10, sticky='w')

        #line
        L = Label(GD, text='Line :', font=FONT2)
        L.grid(row=1, column=0, padx=50, pady=10, sticky='e')
        v_line = StringVar()
        E = ttk.Combobox(GD, textvariable=v_line, font=FONT2, values=(['BLD4#2','BLD5#10/1','BLD5#10/2','BLD6#15/1','BLD6#15/2']), state='readonly')
        E.grid(row=1, column=1, padx=30, pady=10, sticky='w')

        #slotNum
        L = Label(GD, text='Slot num :', font=FONT2)
        L.grid(row=2, column=0, padx=50, pady=10, sticky='e')
        v_slotNum = StringVar()
        vaLues = ['Temporary']
        for i in range(1,121): 
            vaLues.append(f"{i:03}")
        E = ttk.Combobox(GD, textvariable=v_slotNum, font=FONT2, values=vaLues, state='readonly')
        E.grid(row=2, column=1, padx=30, pady=10, sticky='w')

        #With draw by
        L = Label(GD, text='Withdraw by :', font=FONT2)
        L.grid(row=3, column=0, padx=50, pady=10, sticky='e')
        v_By = StringVar()
        E = ttk.Combobox(GD, textvariable=v_By, font=FONT2, values=techname, state='readonly')
        E.grid(row=3, column=1, padx=30, pady=10, sticky='w')

        #other detail
        OD = LabelFrame(MF, text='other detail', width=300, height=400)
        OD.grid(row=0, column=1, padx=10, pady=10)

        #descriptiond
        L = Label(OD, text='Descriptions :', font=FONT2)
        L.grid(row=0, column=0, padx=50, pady=10, sticky='e')
        Edesc = Text(OD, width=20, height=3)
        Edesc.grid(row=0, column=1, padx=30, pady=10, sticky='w')

        #photoselect
        PS = LabelFrame(OD, width=200, height=150)
        PS.grid(row=1, column=0, padx=10, pady=10)

        #photolable
        v_photoStencil = StringVar()
        E = ttk.Entry(PS, textvariable=v_photoStencil, font=FONT2)
        E.grid(row=0, column=0, padx=50, pady=10)

        #select button
        B = ttk.Button(PS, text='Select Photo', command=selectPhoto)
        B.grid(row=1, column=0, pady=20)

        #photoselect
        SP = LabelFrame(OD, width=200, height=150)
        SP.grid(row=1, column=1, padx=10, pady=10)

        #label show photo
        photoStencil = Label(SP)
        photoStencil.pack(padx=10, pady=10)

        #button
        B = ttk.Button(self, text='Save', command=stenCilWith)
        B.pack(padx=10, pady=10)

        #button
        B = ttk.Button(self, text='Clear', command=reset)
        B.pack(padx=10)

##class stencil return
class stenCilre(Frame):
    def __init__(self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        #definepath
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.stencilpath = self.config['DATABASE']['dbtoolpath']
        self.stencilsheet = self.config['DATABASE']['stencilsheet']
        self.stencilphotopath = self.config['DATABASE']['dbphototoolpath']
        self.stencilsummary = self.config['DATABASE']['summarysheet']

        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()

            ##S/N search type:
            if check_search_type == ('Stencil P/N'):    #S/N search
                for row in sheet.iter_rows(min_row=3, max_row=None, min_col=0, max_col=29, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[1]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(),
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[5]).strip(),str(row[26]).strip()])
                return results

            #line
            elif check_search_type == 'Line':
                for row in sheet.iter_rows(min_row=3, max_row=None, min_col=0, max_col=29, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[2]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(),
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[5]).strip(),str(row[26]).strip()])
                return results
            
            #slot
            elif check_search_type == 'Slot Number':
                for row in sheet.iter_rows(min_row=3, max_row=None, min_col=0, max_col=29, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[3]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(),
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[5]).strip(),str(row[26]).strip()])
                return results
            
            #date
            elif check_search_type == 'Date':
                for row in sheet.iter_rows(min_row=3, max_row=None, min_col=0, max_col=29, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[5]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(),
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[5]).strip(),str(row[26]).strip()])
                return results

            #all
            elif check_search_type == 'All':
                for row in sheet.iter_rows(min_row=4, max_row=None, min_col=0, max_col=29, values_only=True):
                    results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(),
                                        str(row[3]).strip(),str(row[4]).strip(),str(row[5]).strip(),str(row[26]).strip()])
                return results

            
        def on_key_release(event):
            search_prefix = enter.get().strip()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [1], search_prefix)  # Adjust columns as needed
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def display_results(results):
            for iter in S_list.get_children():
                S_list.delete(iter)
            if results:
                for row in results:
                    #row = tuple(row)
                    S_list['columns']=("TSID","Stencil P/N", "Line" , 'Slot',"Withdraw by","Date","Status")
                    S_list.column('TSID', anchor="center", width=1)
                    S_list.column('Stencil P/N', anchor="center", width=150)
                    S_list.column('Line', anchor="center", width=1)
                    S_list.column('Slot', anchor="center", width=1)  
                    S_list.column('Withdraw by', anchor="center", width=1) 
                    S_list.column('Date', anchor="center", width=1) 
                    S_list.column('Status', anchor="center", width=1) 

                    ##heading
                    S_list.heading('TSID', text='TSID',anchor="center")
                    S_list.heading('Stencil P/N', text='Stencil P/N', anchor="center")
                    S_list.heading('Line', text='Line',anchor="center")
                    S_list.heading('Slot', text='Slot', anchor="center")   
                    S_list.heading('Withdraw by', text='Withdraw by', anchor="center")  
                    S_list.heading('Date', text='Date', anchor="center")  
                    S_list.heading('Status', text='Status', anchor="center")          
                    S_list.insert('', 'end', values=row)

            else:
                S_list['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                S_list.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                S_list.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")
            
            def Select_mc(event=None):
                select = S_list.selection()
                output = S_list.item(select)
                if select !=():
                    tsid = output['values'][0]  #values = key
                    wb = load_workbook(filename=file_path)
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=27, values_only=True):
                        if tsid ==  int(row[0]):
                            stencilNum = row[1]
                            stencilNumsendcheck = row[1].lower().replace(' ','').strip()
                            line = row[2]
                            slot = row[3]
                            withdrawby = row[4]
                            dateb = row[5]
                            x1a = row[6]
                            x2a = row[7]
                            x3a = row[8]
                            x4a = row[9]
                            x5a = row[10]
                            dirx = row[11]
                            y1a = row[12]
                            y2a = row[13]
                            y3a = row[14]
                            y4a = row[15]
                            y5a = row[16]
                            diry = row[17]
                            datea = row[18]
                            statussten = row[19]
                            qbuild = row[20]
                            receiveby = row[21]
                            criticaldent = row[22]
                            criticalstra = row[23]
                            criticalapture = row[24]
                            comment = row[25]
                            statususe = row[26]

                            #stencil detail
                            stenCilGUI = Toplevel()
                            stenCilGUI.title('Stencil manage')
                            stenCilGUI.geometry('900x600-50+30')

                            #save functions
                            def saVe_re():
                                X1 = int(x1.get())
                                X2 = int(x2.get())
                                X3 = int(x3.get())
                                X4 = int(x4.get())
                                X5 = int(x5.get())
                                Y1 = int(y1.get())
                                Y2 = int(y2.get())
                                Y3 = int(y3.get())
                                Y4 = int(y4.get())
                                Y5 = int(y5.get())

                                DenT = dent.get()
                                StratcheD = stratched.get()
                                AperturE = aperture.get()

                                QBuilD = v_buiLd.get()
                                Ste_Sta = v_stenStatus.get()
                                ReceivE = v_recEive.get()
                                DesC = vE_desC.get('1.0',END).strip()
                                date = datetime.now().strftime("%d/%b/%Y")
                                checKPhotopatH = v_phoToPath.get()

                                qty_old_build = qbuild

                                if X1 and X2 and X3 and X4 and X5 and Y1 and Y2 and Y3 and Y4 and Y5 and DenT and StratcheD and AperturE and QBuilD and Ste_Sta  and ReceivE:
                                    # try:
                                        exS.StartToolReturnstencil(tsid,X1,X2,X3,X4,X5,Y1,Y2,Y3,Y4,Y5,date,Ste_Sta,QBuilD,ReceivE,DenT,StratcheD,AperturE,DesC,checKPhotopatH)
                                        exSUM.Tool(stencilNum,stencilNumsendcheck, 'Stencil',QBuilD, statususe, qty_old_build)
                                        reset()
                                        display_results(results)
                                        on_key_release(event)
                                        close()

                                    # except Exception as e:
                                    #     messagebox.showerror("Error", f"Failed to data: {e}")
                                else:
                                    messagebox.showinfo('Stencil','โปรดกรอกข้อมูลให้ครบ')
                            def close():
                                stenCilGUI.destroy()
                                stenCilGUI.update()
                            
                            def reset():
                                x1.set('')
                                x2.set('')
                                x3.set('')
                                x4.set('')
                                x5.set('')
                                y1.set('')
                                y2.set('')
                                y3.set('')
                                y4.set('')
                                y5.set('')

                                dent.set('')
                                stratched.set('')
                                aperture.set('')

                                v_buiLd.set('')
                                v_stenStatus.set('')
                                v_recEive.set('')
                                vE_desC.delete('1.0',END)

                                v_phoToPath.set('')
                                photoStencil.config(image=None)
                                photoStencil.image = None

                            def phoToselEcT():
                                file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
                                v_phoToPath.set(file_path)

                                try:
                                    image = PIL.Image.open(v_phoToPath.get())
                                    image = image.resize((180,170))
                                    photo = ImageTk.PhotoImage(image)
                                    photoStencil.config(image=photo)
                                    photoStencil.image = photo
                                except:
                                    ()

                            #main frame
                            MF = LabelFrame(stenCilGUI, font=FONT3, width=890, height=550)
                            MF.pack(padx=10, pady=10)

                            #general detail
                            GD = LabelFrame(MF, text='general detail', width=300, height=400)
                            GD.grid(row=0, column=0, padx=10, pady=10)

                            #tsid
                            L = Label(GD, text=f'TSID : {tsid}', font=FONT3)
                            L.grid(row=0, column=0, padx=10, pady=10)

                            #stencilNum
                            L = Label(GD, text='Stencil P/N :', font=FONT2)
                            L.grid(row=1, column=0, padx=30, pady=5, sticky='e')
                            v_stenCilnum = StringVar()
                            v_stenCilnum.set(stencilNum)
                            E = ttk.Entry(GD, textvariable=v_stenCilnum, font=FONT2, state='readonly')
                            E.grid(row=1, column=1, padx=10, pady=5, sticky='w')

                            #line
                            L = Label(GD, text='Line :', font=FONT2)
                            L.grid(row=2, column=0, padx=30, pady=5, sticky='e')
                            v_liNe = StringVar()
                            v_liNe.set(line)
                            E = ttk.Entry(GD, textvariable=v_liNe, font=FONT2, state='readonly')
                            E.grid(row=2, column=1, padx=10, pady=5, sticky='w')

                            #slot
                            L = Label(GD, text='Slot :',font=FONT2)
                            L.grid(row=3, column=0, padx=30, pady=10, sticky='e')
                            v_slOt = StringVar()
                            v_slOt.set(slot)
                            E = ttk.Entry(GD, textvariable=v_slOt, font=FONT2, state='readonly')
                            E.grid(row=3, column=1, padx=10, pady=5, sticky='w')

                            #by
                            L = Label(GD, text='Withdraw By :', font=FONT2)
                            L.grid(row=4, column=0, padx=30, pady=10, sticky='e')
                            v_bY = StringVar()
                            v_bY.set(withdrawby)
                            E = ttk.Entry(GD, textvariable=v_bY, font=FONT2, state='readonly')
                            E.grid(row=4, column=1, padx=10, pady=5, sticky='w')

                            #date
                            L = Label(GD, text='Date :', font=FONT2)
                            L.grid(row=5, column=0, padx=30, pady=5, sticky='e')
                            v_daTe = StringVar()
                            v_daTe.set(dateb)
                            E = ttk.Entry(GD, textvariable=v_daTe, font=FONT2, state='readonly')
                            E.grid(row=5, column=1, padx=10, pady=5, sticky='w')

                            # #desc
                            L = Label(GD, text='Descriptions :', font=FONT2)
                            L.grid(row=6, column=0, padx=30, pady=10, sticky='e')
                            vE_desC = Text(GD, width=20, height=4)
                            vE_desC.grid(row=6, column=1, padx=10, pady=10, sticky='w')
                            vE_desC.insert('1.0', str(comment))

                            # #current qty build
                            L = Label(GD, text='Current Qty build :', font=FONT2)
                            L.grid(row=7, column=0, padx=30, pady=10, sticky='e')

                            #สรุปการใช้
                            sumFix_path = self.stencilpath
                            sumFix_sheet = self.stencilsummary  #spare part (reg)
                            #loadworkbook
                            excelfile = load_workbook(filename=sumFix_path)
                            excelfile.active = excelfile[sumFix_sheet]
                            sheet = excelfile[sumFix_sheet]
                            currentBuild = 0
                            v_currentBuild = IntVar()
                            v_currentBuild.set(currentBuild)
                            E = ttk.Entry(GD, textvariable=v_currentBuild, font=FONT2, state='readonly')
                            E.grid(row=7, column=1, padx=10, pady=10, sticky='w')

                            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=10, values_only=True):
                                if row[0].lower() == stencilNum.lower():
                                    currentBuild = row[2]
                                    v_currentBuild = IntVar()
                                    v_currentBuild.set(currentBuild)
                                    E = ttk.Entry(GD, textvariable=v_currentBuild, font=FONT2, state='readonly')
                                    E.grid(row=7, column=1, padx=10, pady=10, sticky='w')
                                    break
                                
                            # #status
                            if statususe == 'not receive':
                                L = Label(GD, text=f'Status : {statususe}', font=FONT2, bg='red')
                                L.grid(row=8, column=0, padx=30, pady=10)
                            else:
                                L = Label(GD, text=f'Status : {statususe}', font=FONT2, bg='green')
                                L.grid(row=8, column=0, padx=30, pady=10)

                            #other frame
                            OD = LabelFrame(MF, text='other detail', width=300, height=400)
                            OD.grid(row=0, column=1, padx=10, pady=10)

                            # #tension control
                            TEN = LabelFrame(OD, text='Tension', font=FONT3, width=230, height=190)
                            TEN.grid(row=0, column=0, padx=10, pady=10)

                            #spectension
                            L = Label(TEN, text='Pass : >=32 N/cm', font=FONT3)
                            L.grid(row=0, column=0)

                            L = Label(TEN, text='Fail : <32 N/cm', font=FONT3)
                            L.grid(row=0, column=1)

                            X = LabelFrame(TEN, text='X', font=FONT3, width=230, height=190)
                            X.grid(row=1, column=0, padx=10, pady=10)

                            def AleartTension(event):
                                if E1x.get() != (''):
                                    try:
                                        cx1 = int(E1x.get())
                                        if 32 < cx1 <= 50:
                                            E1x.config(bg='lightgreen')
                                        else:
                                            E1x.config(bg='lightcoral')
                                    except:
                                        ()
                                
                                if E2x.get() != (''):
                                    cx2 = int(E2x.get())
                                    if 32 < cx2 <= 50:
                                        E2x.config(bg='lightgreen')
                                    else:
                                        E2x.config(bg='lightcoral')
                                
                                if E3x.get() != (''):
                                    cx3 = int(E3x.get())
                                    if 32 < cx3 <= 50:
                                        E3x.config(bg='lightgreen')
                                    else:
                                        E3x.config(bg='lightcoral')
                                
                                if E4x.get() != (''):
                                    cx4 = int(E4x.get())
                                    if 32 < cx4 <= 50:
                                        E4x.config(bg='lightgreen')
                                    else:
                                        E4x.config(bg='lightcoral')
                                
                                if E5x.get() != (''):
                                    cx5 = int(E5x.get())
                                    if 32 < cx5 <= 50:
                                        E5x.config(bg='lightgreen')
                                    else:
                                        E5x.config(bg='lightcoral')

                                if E1y.get() != (''):
                                    cy1 = int(E1y.get())
                                    if 32 < cy1 <= 50:
                                        E1y.config(bg='lightgreen')
                                    else:
                                        E1y.config(bg='lightcoral')

                                if E2y.get() != (''):
                                    cy2 = int(E2y.get())
                                    if 32 < cy2 <= 50:
                                        E2y.config(bg='lightgreen')
                                    else:
                                        E2y.config(bg='lightcoral')
                                    
                                if E3y.get() != (''):
                                    cy3 = int(E3y.get())
                                    if 32 < cy3 <= 50:
                                        E3y.config(bg='lightgreen')
                                    else:
                                        E3y.config(bg='lightcoral')

                                if E4y.get() != (''):
                                    cy4 = int(E4y.get())
                                    if 32 < cy4 <= 50:
                                        E4y.config(bg='lightgreen')
                                    else:
                                        E4y.config(bg='lightcoral')
                                
                                if E5y.get() != (''):
                                    cy5 = int(E5y.get())
                                    if 32 < cy5 <= 50:
                                        E5y.config(bg='lightgreen')
                                    else:
                                        E5y.config(bg='lightcoral')
                                    
                            ##x
                            L = Label(X, text='X1 :', font=FONT3)
                            L.grid(row=0, column=0, padx=10, pady=5, sticky='e')
                            x1 = StringVar()
                            E1x = tk.Entry(X, textvariable=x1, font=FONT2, width=5)
                            E1x.grid(row=0, column=1, padx=10, pady=5, sticky='w')
                            E1x.bind('<KeyRelease>', AleartTension)
                            E1x.bind('<Return>', lambda x:E2x.focus())
                            if x1a != None:
                                x1.set(x1a)

                            L = Label(X, text='X2 :', font=FONT3)
                            L.grid(row=1, column=0, padx=10, pady=5, sticky='e')
                            x2 = StringVar()
                            E2x = tk.Entry(X, textvariable=x2, font=FONT2, width=5)
                            E2x.grid(row=1, column=1, padx=10, pady=5, sticky='w')
                            E2x.bind('<KeyRelease>', AleartTension)
                            E2x.bind('<Return>', lambda x:E3x.focus())
                            if x2a != None:
                                x2.set(x2a)

                            L = Label(X, text='X3 :', font=FONT3)
                            L.grid(row=2, column=0, padx=10, pady=5, sticky='e')
                            x3 = StringVar()
                            E3x = tk.Entry(X, textvariable=x3, font=FONT2, width=5)
                            E3x.grid(row=2, column=1, padx=10, pady=5, sticky='w')
                            E3x.bind('<KeyRelease>', AleartTension)
                            E3x.bind('<Return>', lambda x:E4x.focus())
                            if x3a != None:
                                x3.set(x3a)

                            L = Label(X, text='X4 :', font=FONT3)
                            L.grid(row=3, column=0, padx=10, pady=5, sticky='e')
                            x4 = StringVar()
                            E4x = tk.Entry(X, textvariable=x4, font=FONT2, width=5)
                            E4x.grid(row=3, column=1, padx=10, pady=5, sticky='w')
                            E4x.bind('<KeyRelease>', AleartTension)
                            E4x.bind('<Return>', lambda x:E5x.focus())
                            if x4a != None:
                                x4.set(x4a)

                            L = Label(X, text='X5 :', font=FONT3)
                            L.grid(row=4, column=0, padx=10, sticky='e')
                            x5 = StringVar()
                            E5x = tk.Entry(X, textvariable=x5, font=FONT2, width=5)
                            E5x.grid(row=4, column=1, padx=10, pady=5, sticky='w')
                            E5x.bind('<KeyRelease>', AleartTension)
                            E5x.bind('<Return>', lambda x:E1y.focus())
                            if x5a != None:
                                x5.set(x5a)

                            Y = LabelFrame(TEN, text='Y', font=FONT3, width=230, height=190)
                            Y.grid(row=1, column=1, padx=10, pady=10)

                            ##y
                            L = Label(Y, text='Y1 :', font=FONT3)
                            L.grid(row=0, column=0, padx=10, pady=5, sticky='e')
                            y1 = StringVar()
                            E1y = tk.Entry(Y, textvariable=y1, font=FONT2, width=5)
                            E1y.grid(row=0, column=1, padx=10, pady=5, sticky='w')
                            E1y.bind('<KeyRelease>', AleartTension)
                            E1y.bind('<Return>', lambda x:E2y.focus())
                            if y1a != None:
                                y1.set(y1a)
                            
                            L = Label(Y, text='Y2 :', font=FONT3)
                            L.grid(row=1, column=0, padx=10, pady=5, sticky='e')
                            y2 = StringVar()
                            E2y = tk.Entry(Y, textvariable=y2, font=FONT2, width=5)
                            E2y.grid(row=1, column=1, padx=10, pady=5, sticky='w')
                            E2y.bind('<KeyRelease>', AleartTension)
                            E2y.bind('<Return>', lambda x:E3y.focus())
                            if y2a != None:
                                y2.set(y2a)
                            
                            L = Label(Y, text='Y3 :', font=FONT3)
                            L.grid(row=2, column=0, padx=10, pady=5, sticky='e')
                            y3 = StringVar()
                            E3y = tk.Entry(Y, textvariable=y3, font=FONT2, width=5)
                            E3y.grid(row=2, column=1, padx=10, pady=5, sticky='w')
                            E3y.bind('<KeyRelease>', AleartTension)
                            E3y.bind('<Return>', lambda x:E4y.focus())
                            if y3a != None:
                                y3.set(y3a)
                            
                            L = Label(Y, text='Y4 :', font=FONT3)
                            L.grid(row=3, column=0, padx=10, pady=5, sticky='e')
                            y4 = StringVar()
                            E4y = tk.Entry(Y, textvariable=y4, font=FONT2, width=5)
                            E4y.grid(row=3, column=1, padx=10, pady=5, sticky='w')
                            E4y.bind('<KeyRelease>', AleartTension)
                            E4y.bind('<Return>', lambda x:E5y.focus())
                            if y4a != None:
                                y4.set(y4a)
                            
                            L = Label(Y, text='Y5 :', font=FONT3)
                            L.grid(row=4, column=0, padx=10, pady=5, sticky='e')
                            y5 = StringVar()
                            E5y = tk.Entry(Y, textvariable=y5, font=FONT2, width=5)
                            E5y.grid(row=4, column=1, padx=10, pady=5, sticky='w')
                            E5y.bind('<KeyRelease>', AleartTension)
                            if y5a != None:
                                y5.set(y5a)

                            # #criterial
                            CRI = LabelFrame(OD, text='Criteria',font=FONT3, width=230, height=190)
                            CRI.grid(row=1, column=0, padx=10, pady=10)

                            #dent
                            L = Label(CRI, text='รอยเว้า :', font=FONT2)
                            L.grid(row=0, column=0, padx=10, pady=2.5, sticky='e')
                            dent = StringVar()
                            E = ttk.Combobox(CRI, textvariable=dent, values=(['OK','NG']), state='readonly', width=10)
                            E.grid(row=0, column=1, padx=10, pady=2.5, sticky='w')
                            if criticaldent != None:
                                dent.set(criticaldent)

                            #scratch
                            L = Label(CRI, text='รอยขีดข่วน :', font=FONT2)
                            L.grid(row=1, column=0, padx=10, pady=2.5, sticky='e')
                            stratched = StringVar()
                            E = ttk.Combobox(CRI, textvariable=stratched, values=(['OK','NG']), state='readonly', width=10)
                            E.grid(row=1, column=1, padx=10, pady=2.5, sticky='w')
                            if criticalstra != None:
                                stratched.set(criticalstra)

                            #aperture
                            L = Label(CRI, text='Aperture :', font=FONT2)
                            L.grid(row=2, column=0, padx=10, pady=2.5, sticky='e')
                            aperture = StringVar()
                            E = ttk.Combobox(CRI, textvariable=aperture, values=(['OK','NG']), state='readonly', width=10)
                            E.grid(row=2, column=1, padx=10, pady=2.5, sticky='w')
                            if criticalapture != None:
                                aperture.set(criticalapture)
                            
                            #Help
                            def WindowHelp():
                                HelpGui = Toplevel()
                                HelpGui.title('Help')
                                Helppage = Help(HelpGui)
                                Helppage.pack()
                                HelpGui.mainloop()

                            BH = tk.Button(OD, text='?', width=2, height=1, command=WindowHelp)
                            BH.grid(row=2, column=0, sticky='w', pady=5, padx=5)

                            OD2 = LabelFrame(MF)
                            OD2.grid(row=0, column=2, padx=10, pady=10)

                            # #build
                            L = Label(OD2, text='Quantity Build :', font=FONT2)
                            L.grid(row=0, column=0, padx=10, pady=10, sticky='e')
                            v_buiLd = IntVar()
                            E = ttk.Entry(OD2, textvariable=v_buiLd, font=FONT2)
                            E.grid(row=0, column=1, padx=10, pady=10, sticky='w')
                            if qbuild != None:
                                v_buiLd.set(qbuild)

                            #stencil check
                            L = Label(OD2, text='Stencil status :', font=FONT2)
                            L.grid(row=1, column=0, padx=10, pady=10, sticky='e')
                            v_stenStatus = StringVar()
                            E = ttk.Combobox(OD2, textvariable=v_stenStatus, font=FONT2, values=(['OK','NG','Inform to process']), state='readonly')
                            E.grid(row=1, column=1, padx=10, pady=10, sticky='w')
                            if statussten != None:
                                v_stenStatus.set(statussten)

                            #receive by
                            L = Label(OD2, text='Receive by :', font=FONT2)
                            L.grid(row=2, column=0, padx=10, pady=10, sticky='e')
                            v_recEive = StringVar()
                            E = ttk.Combobox(OD2, textvariable=v_recEive, font=FONT2, values=techname, state='readonly')
                            E.grid(row=2, column=1, padx=10, pady=10, sticky='w')
                            if receiveby != None:
                                v_recEive.set(receiveby)

                            #date receive
                            if datea != None:
                                L = Label(OD2, text='Date receive', font=FONT2)
                                L.grid(row=3, column=0, padx=10, pady=10, sticky='e')
                                v_datea = StringVar()
                                v_datea.set(datea)
                                E = ttk.Entry(OD2, textvariable=v_datea, font=FONT2)
                                E.grid(row=3, column=1, padx=10, pady=10, sticky='w')

                            # #select photo
                            SP = LabelFrame(OD2, width=190, height=190)
                            SP.grid(row=4, column=1, padx=10, pady=10)

                            # #show photo
                            PS = LabelFrame(OD2, width=190, height=190)
                            PS.grid(row=4, column=0, padx=10, pady=10)

                            #label photo
                            photoStencil = Label(PS)
                            photoStencil.pack(padx=10, pady=10)

                            try:
                                image = PIL.Image.open(f"{self.stencilphotopath}\\{tsid}.png")
                                image = image.resize((180,170))
                                photo = ImageTk.PhotoImage(image)
                                photoStencil.config(image=photo)
                                photoStencil.image = photo
                            except Exception as e:
                                ()

                            #photoPath
                            v_phoToPath = StringVar()
                            EphoTo = ttk.Entry(SP, textvariable=v_phoToPath, font=FONT2)
                            EphoTo.pack(padx=20, pady=10)

                            #selectphoto
                            B = ttk.Button(SP, text='Photo Select', command=phoToselEcT)
                            B.pack(padx=10, pady=20)

                            #save 
                            saVe = ttk.Button(OD2, text='Save', command=saVe_re)
                            saVe.grid(row=5, column=0, padx=10, pady=10, sticky='e')

                            #clear
                            B = ttk.Button(OD2, text='Clear', command=reset)
                            B.grid(row=5, column=1, padx=10, pady=10, sticky='w')

                            #focus
                            E1x.focus()
                            
                            stenCilGUI.mainloop()

            #select delete
            def delete_right_click(event=None):
                select = S_list.selection()
                output = S_list.item(select)
                tsid = output['values'][0]
                check = messagebox.askyesno('ยืนยันการลบ','คุณต้องการลบข้อมูล?...')
                if check == True:
                    if output['values'] != (''):
                        try:
                            exS.StartToolDeletEstencil(tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")

            pop_up_menu = Menu(self, tearoff=0)
            pop_up_menu.add_command(label='delete', command=delete_right_click)
            
            def popup(event):
                pop_up_menu.post(event.x_root, event.y_root)
            S_list.bind('<Button-3>', popup)   #Button 3 คือ click ขวา

            S_list.bind('<Double-1>', Select_mc)

        def clear_results():
            for iter in S_list.get_children():
                S_list.delete(iter)

        #define path
        file_path = self.stencilpath
        sheet_name = self.stencilsheet

        #search_type
        search_type = Label(self, text='Search type :', font=FONT2)
        search_type.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('Stencil P/N')
        search_type_combo = ttk.Combobox(self, values=['Stencil P/N','Line','Slot Number','Date','All'], state='readonly', textvariable=v_search_type)
        search_type_combo.place(x=120, y=10, height=30)  

        #entry search
        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        enter = ttk.Entry(self, font=FONT2)
        enter.place(x=120, y=50, height=30, width=150)
        enter.bind('<KeyRelease>', on_key_release)

        #create list part
        header = ["TSID","Stencil P/N", "Line" , 'Slot',"Withdraw by","Date","Status"]
        headerw = [10,100,10,10,10,10,10]
        S_list = ttk.Treeview(self, columns=header, show='headings')
        S_list.place(x=5, y=100, width=1250, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',12))
        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")
        #--header and width
        for h,w in zip(header,headerw):
            S_list.heading(h,text=h)
            S_list.column(h,width=w,anchor='center')
        
        #scroll bar vertical
        S_list_scroll = Scrollbar(S_list)
        S_list_scroll.pack(side=RIGHT,fill=Y)
        S_list_scroll.config(command=S_list.yview)


# gui = Tk()
# gui.title('Bord')
# gui.geometry('500x500')
# a = stenCilre(gui)
# # a = bordReturn(gui)
# # a = bordwith(gui)
# a.pack()
# gui.mainloop()


