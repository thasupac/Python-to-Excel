import pathlib
import openpyxl as xl
from tkinter import messagebox, Toplevel, Label
from datetime import datetime
import threading
import os
import PIL.Image
import configparser

#font
FONT = ('Angsana New',18)

##--register tool
class ToolReg:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.toolpath = self.config['DATABASE']['dbtoolpath']
        self.toolregsheet = self.config['DATABASE']['regtoolsheet']
        self.toolphotopath = self.config['DATABASE']['dbphototoolpath']

        from send_mail import SendMail
        self.songmail = SendMail()

    ##--progress
    def StartToolReG(self,tsid,model,modelNum,customer,types,qty,unit,line,regby,desc,check_photo_save):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลงทะเบียนข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ToolReg.toolReg(self,tsid,model,modelNum,customer,types,qty,unit,line,regby,desc,check_photo_save))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool Register', 'ลงทะเบียนข้อมูลสำเร็จ')
        
    def toolReg(self,tsid,model,modelNum,customer,types,qty,unit,line,regby,desc,check_photo_save):
        route = pathlib.Path(self.toolpath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.toolregsheet]
        self.sheet = self.wb[self.toolregsheet]
        date = datetime.now().strftime("%d/%b/%Y")
        rows = self.sheet.max_row

        try:
            if check_photo_save !=(''):
                image = PIL.Image.open(check_photo_save)  
                target_path = self.toolphotopath
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

            self.sheet.cell(row=rows+1,column=1,value=int(tsid))
            self.sheet.cell(row=rows+1,column=2,value=model)
            self.sheet.cell(row=rows+1,column=3,value=modelNum)
            self.sheet.cell(row=rows+1,column=4,value=customer)
            self.sheet.cell(row=rows+1,column=5,value=types)
            self.sheet.cell(row=rows+1,column=6,value=int(qty))
            self.sheet.cell(row=rows+1,column=7,value=unit)
            self.sheet.cell(row=rows+1,column=8,value=line)
            self.sheet.cell(row=rows+1,column=9,value=regby)
            self.sheet.cell(row=rows+1,column=10,value=date)
            self.sheet.cell(row=rows+1,column=11,value=desc)
            self.wb.save(self.toolpath)

            #notification to outlook
            self.songmail.Toolsongmail(model, modelNum, customer, types, qty, unit, line, regby, desc, check_photo_save)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    ##--progress
    def StartToolUpdatE(self,tsid,U_model,U_modelNum,U_customer,U_types,U_qty,U_unit,U_line,U_desc,add_photo):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังอัปเดตข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ToolReg.toolUpdate(self,tsid,U_model,U_modelNum,U_customer,U_types,U_qty,U_unit,U_line,U_desc,add_photo))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool Update', 'อัปเดตข้อมูลสำเร็จ')

    def toolUpdate(self,tsid,U_model,U_modelNum,U_customer,U_types,U_qty,U_unit,U_line,U_desc,add_photo):
        route = pathlib.Path(self.toolpath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.toolregsheet]
        self.sheet = self.wb[self.toolregsheet]
        date = datetime.now().strftime("%d/%b/%Y")
        rows = self.sheet.max_row

        #update photo
        if add_photo !=(''):
                image = PIL.Image.open(add_photo)  
                target_path = self.toolphotopath
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

        try:
            for i in range(2, rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    self.sheet.cell(row=i,column=2,value=U_model)
                    self.sheet.cell(row=i,column=3,value=U_modelNum)
                    self.sheet.cell(row=i,column=4,value=U_customer)
                    self.sheet.cell(row=i,column=5,value=U_types)
                    self.sheet.cell(row=i,column=6,value=int(U_qty))
                    self.sheet.cell(row=i,column=7,value=U_unit)
                    self.sheet.cell(row=i,column=8,value=U_line)
                    self.sheet.cell(row=i,column=11,value=U_desc)
                    self.sheet.cell(row=i,column=12,value=date)
                    self.wb.save(self.toolpath)

                    #notification to outlook
                    self.songmail.Toolsongmail(U_model, U_modelNum, U_customer, U_types, U_qty, U_unit, U_line, '-', 'Update', '')

        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

     ##--progress
    def StartToolDeletE(self,tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลบข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ToolReg.toolDelete(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool Deletee', 'ลบข้อมูลสำเร็จ')
    
    def toolDelete(self,tsid):
        route = pathlib.Path(self.toolpath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.toolregsheet]
        self.sheet = self.wb[self.toolregsheet]
        rows = self.sheet.max_row
        try:
            try:
                os.remove(f"{self.toolphotopath}\\{tsid}.png")
            except:
                ()

            for i in range(2,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    modelNum = self.sheet.cell(row=i,column=3).value
                    self.sheet.delete_rows(i,1)
                    self.wb.save(self.toolpath)

                    #notification to outlook
                    self.songmail.Toolsongmail('-', modelNum, '-', '-', '-', '-', '-', '-', 'Delete', '')
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

##--bord profile
class ExcelB:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.bordprofilepath = self.config['DATABASE']['dbtoolpath']
        self.bordprofilephotopath = self.config['DATABASE']['dbphototoolpath']
        self.bordprofilesheet = self.config['DATABASE']['bordprofilesheet']
      
    #progress
    def StartToolWithdrawBorD(self,tsid,model,num_model,line,side,with_by,physical,fixture_status,signal_status,desc,photo):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูลการยืม...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExcelB.withDrawborD(self,tsid,model,num_model,line,side,with_by,physical,fixture_status,signal_status,desc,photo))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'บันทึกการยืมสำเร็จ')

    #withDrawborD
    def withDrawborD(self,tsid,model,num_model,line,side,with_by,physical,fixture_status,signal_status,desc,photo):
        route = pathlib.Path(self.bordprofilepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.bordprofilesheet]
        self.sheet = self.wb[self.bordprofilesheet]
        rows = self.sheet.max_row
        no = rows-2
        date = datetime.now().strftime("%d/%b/%Y")

        try:
            if photo !=(''):
                image = PIL.Image.open(photo)
                target_path = self.bordprofilephotopath
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

            self.sheet.cell(row=rows+1,column=1,value=int(tsid))
            self.sheet.cell(row=rows+1,column=2,value=model)
            self.sheet.cell(row=rows+1,column=3,value=num_model)
            self.sheet.cell(row=rows+1,column=4,value=line)
            self.sheet.cell(row=rows+1,column=5,value=side)
            self.sheet.cell(row=rows+1,column=6,value=physical)
            self.sheet.cell(row=rows+1,column=7,value=fixture_status)
            self.sheet.cell(row=rows+1,column=8,value=signal_status)
            self.sheet.cell(row=rows+1,column=9,value=date)
            self.sheet.cell(row=rows+1,column=10,value=with_by)
            self.sheet.cell(row=rows+1,column=16,value=desc)
            self.sheet.cell(row=rows+1,column=17,value='not receive')
            self.wb.save(self.bordprofilepath)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")
    
    #progress
    def StartToolDReturnBorD(self,tsid,c1a,c2a,c3a,reby,desc,photosave):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูลการคืน...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExcelB.returen_bord(self,tsid,c1a,c2a,c3a,reby,desc,photosave))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'บันทึกข้อมูลการคืนสำเร็จ')

    def returen_bord(self,tsid,c1a,c2a,c3a,reby,desc,photosave):
        route = pathlib.Path(self.bordprofilepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.bordprofilesheet]
        self.sheet = self.wb[self.bordprofilesheet]
        date = datetime.now().strftime("%d/%b/%Y")
        rows = self.sheet.max_row

        try:
            if photosave !=(''):
                image = PIL.Image.open(photosave)
                target_path = self.bordprofilephotopath
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

            for i in range(2, rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:  
                    self.sheet.cell(row=i,column=11).value=c1a
                    self.sheet.cell(row=i,column=12).value=c2a
                    self.sheet.cell(row=i,column=13).value=c3a
                    self.sheet.cell(row=i,column=14).value=date
                    self.sheet.cell(row=i,column=15).value=reby
                    self.sheet.cell(row=i,column=16).value=desc
                    self.sheet.cell(row=i,column=17).value='receive'
                    self.wb.save(self.bordprofilepath)
                 
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    #progress
    def StartToolDeletEBorD(self,tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลบข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExcelB.delete_bord(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'ลบข้อมูลสำเร็จ')

    def delete_bord(self,tsid):
        route = pathlib.Path(self.bordprofilepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.bordprofilesheet]
        self.sheet = self.wb[self.bordprofilesheet]
        rows = self.sheet.max_row
        try:
            try:
                os.remove(f"{self.bordprofilephotopath}\\{tsid}.png")
            except:
                ()

            for i in range(2,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    self.sheet.delete_rows(i,1)
                    self.wb.save(self.bordprofilepath)
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

##--Fixture
class ExcelF:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.fixturepath = self.config['DATABASE']['dbtoolpath']
        self.fixturephotopath = self.config['DATABASE']['dbphototoolpath']
        self.fixturesheet = self.config['DATABASE']['palletsheet']

    #progress
    def StartToolWithdrawFixture(self,tsid,f_id,f_line,f_side,f_cus,f_qty,f_by,comment,photofix):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูลการยืม...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExcelF.WithDrawFixture(self,tsid,f_id,f_line,f_side,f_cus,f_qty,f_by,comment,photofix))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'บันทึกข้อมูลการยืมสำเร็จ')

    #WithDrawFixture
    def WithDrawFixture(self,tsid,f_id,f_line,f_side,f_cus,f_qty,f_by,comment,photofix):
        route = pathlib.Path(self.fixturepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.fixturesheet]
        self.sheet = self.wb[self.fixturesheet]
        date = datetime.now().strftime("%d/%b/%Y")
        rows = self.sheet.max_row
        try:
            if photofix !=(''):
                image = PIL.Image.open(photofix)
                target_path = self.fixturephotopath
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

            self.sheet.cell(row=rows+1,column=1,value=int(tsid))
            self.sheet.cell(row=rows+1,column=2,value=f_id)
            self.sheet.cell(row=rows+1,column=3,value=f_line)
            self.sheet.cell(row=rows+1,column=4,value=f_side)
            self.sheet.cell(row=rows+1,column=5,value=f_cus)
            self.sheet.cell(row=rows+1,column=6,value=int(f_qty))
            self.sheet.cell(row=rows+1,column=7,value=f_by)
            self.sheet.cell(row=rows+1,column=8,value=date)
            self.sheet.cell(row=rows+1,column=18,value=comment)
            self.sheet.cell(row=rows+1,column=19,value='not receive')
            self.wb.save(self.fixturepath)
   
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    #progress
    def StartToolReturnwFixture(self,tsid,c1s,c2s,c3s,c4s,c5s,c6s,re_qty,re_by,re_remark,check_path_photo):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูลการคืน...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExcelF.return_fix(self,tsid,c1s,c2s,c3s,c4s,c5s,c6s,re_qty,re_by,re_remark,check_path_photo))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'บันทึกข้อมูลการคืนสำเร็จ')

    def return_fix(self,tsid,c1s,c2s,c3s,c4s,c5s,c6s,re_qty,re_by,re_remark,check_path_photo):
        route = pathlib.Path(self.fixturepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.fixturesheet]
        self.sheet = self.wb[self.fixturesheet]
        rows = self.sheet.max_row
        date = datetime.now().strftime("%d/%b/%Y")
        ##
        try:
            if check_path_photo !=(''):
                image = PIL.Image.open(check_path_photo)
                target_path = self.fixturephotopath
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

            for i in range(1,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:  
                    self.sheet.cell(row=i,column=9).value=c1s
                    self.sheet.cell(row=i,column=10).value=c2s
                    self.sheet.cell(row=i,column=11).value=c3s
                    self.sheet.cell(row=i,column=12).value=c4s
                    self.sheet.cell(row=i,column=13).value=c5s
                    self.sheet.cell(row=i,column=14).value=c6s
                    self.sheet.cell(row=i,column=15).value=int(re_qty)
                    self.sheet.cell(row=i,column=16).value=re_by
                    self.sheet.cell(row=i,column=17).value=date
                    self.sheet.cell(row=i,column=18).value=re_remark
                    self.sheet.cell(row=i,column=19).value='receive'
                    self.wb.save(self.fixturepath)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")
    #progress
    def StartToolDeleteFixture(self,tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลบข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = ExcelF.deleteToolFixture(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'ลบข้อมูลสำเร็จ')

    def deleteToolFixture(self,tsid):
        route = pathlib.Path(self.fixturepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.fixturesheet]
        self.sheet = self.wb[self.fixturesheet]
        rows = self.sheet.max_row

        try:
            try:
                os.remove(f"{self.fixturephotopath}\\{tsid}.png")
            except:
                ()
            for i in range(2,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    self.sheet.delete_rows(i,1)
                    self.wb.save(self.fixturepath)
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

#class stencil
class stenCil:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.stencilpath = self.config['DATABASE']['dbtoolpath']
        self.stencilphotopath = self.config['DATABASE']['dbphototoolpath']
        self.stencilsheet = self.config['DATABASE']['stencilsheet']

        from send_mail import SendMail
        self.songmail = SendMail()

    #progress
    def StartToolWithdrawstencil(self,tsid,stenNum,line,slotNum,by,date,photoStencilSave,comment):
        # tsid,stenNum,line,slotNum,by,date,photoStencilSave,comment
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูลการยืม...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = stenCil.Wd_stencil(self,tsid,stenNum,line,slotNum,by,date,photoStencilSave,comment))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'บันทึกข้อมูลการยืมสำเร็จ')

    def Wd_stencil(self,tsid,stenNum,line,slotNum,by,date,photoStencilSave,comment):
        route = pathlib.Path(self.stencilpath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.stencilsheet]
        self.sheet = self.wb[self.stencilsheet]
        rows = self.sheet.max_row
        
        try:
            if photoStencilSave !=(''):
                image = PIL.Image.open(photoStencilSave)  
                target_path = self.stencilphotopath
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)

            self.sheet.cell(row=rows+1,column=1,value=int(tsid))
            self.sheet.cell(row=rows+1,column=2,value=stenNum)
            self.sheet.cell(row=rows+1,column=3,value=line)
            self.sheet.cell(row=rows+1,column=4,value=slotNum)
            self.sheet.cell(row=rows+1,column=5,value=by)
            self.sheet.cell(row=rows+1,column=6,value=date)
            self.sheet.cell(row=rows+1,column=26,value=comment)
            self.sheet.cell(row=rows+1,column=27,value='not receive')
            self.wb.save(self.stencilpath)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")
    #progress
    def StartToolReturnstencil(self,tsid,x1,x2,x3,x4,x5,y1,y2,y3,y4,y5,date,statusST,qBuild,RebY,dent,stratched,aperture,DesC,checKPhotopatH):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูลการคืน...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = stenCil.Re_stenciL(self,tsid,x1,x2,x3,x4,x5,y1,y2,y3,y4,y5,date,statusST,qBuild,RebY,dent,stratched,aperture,DesC,checKPhotopatH))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'บันทึกข้อมูลการคืนสำเร็จ')
    
    def Re_stenciL(self,tsid,x1,x2,x3,x4,x5,y1,y2,y3,y4,y5,date,statusST,qBuild,RebY,dent,stratched,aperture,DesC,checKPhotopatH):
        route = pathlib.Path(self.stencilpath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.stencilsheet]
        self.sheet = self.wb[self.stencilsheet]
        rows = self.sheet.max_row

        #find max-min
        x_valuE = [x1,x2,x3,x4,x5]
        max_X = max(x_valuE)
        min_X = min(x_valuE)
        diRX = max_X-min_X

        y_valuE = [y1,y2,y3,y4,y5]
        max_Y = max(y_valuE)
        min_Y = min(y_valuE)
        diRY = max_Y-min_Y

        if checKPhotopatH !=(''):
            image = PIL.Image.open(checKPhotopatH)
            target_path = self.stencilphotopath
            photo_name = f'{tsid}.png'
            photo_path = os.path.join(target_path, photo_name)
            image.save(photo_path)

        for i in range(3, rows+1):
            if self.sheet.cell(row=i,column=1).value == tsid:
                try:  
                    self.sheet.cell(row=i,column=7).value=int(x1)
                    self.sheet.cell(row=i,column=8).value=int(x2)
                    self.sheet.cell(row=i,column=9).value=int(x3)
                    self.sheet.cell(row=i,column=10).value=int(x4)
                    self.sheet.cell(row=i,column=11).value=int(x5)
                    self.sheet.cell(row=i,column=12).value=int(diRX)
                    
                    self.sheet.cell(row=i,column=13).value=int(y1)
                    self.sheet.cell(row=i,column=14).value=int(y2)
                    self.sheet.cell(row=i,column=15).value=int(y3)
                    self.sheet.cell(row=i,column=16).value=int(y4)
                    self.sheet.cell(row=i,column=17).value=int(y5)
                    self.sheet.cell(row=i,column=18).value=int(diRY)

                    self.sheet.cell(row=i,column=19).value=date
                    self.sheet.cell(row=i,column=20).value=statusST
                    self.sheet.cell(row=i,column=21).value=int(qBuild)
                    self.sheet.cell(row=i,column=22).value=RebY
                    self.sheet.cell(row=i,column=23).value=dent
                    self.sheet.cell(row=i,column=24).value=stratched
                    self.sheet.cell(row=i,column=25).value=aperture
                    self.sheet.cell(row=i,column=26).value=DesC
                    self.sheet.cell(row=i,column=27).value='receive'
                            
                    self.wb.save(self.stencilpath)

                    #if status stencil not good send to responsible people
                    if statusST == 'NG':
                        modelnum = self.sheet.cell(row=i,column=2).value
                        self.songmail.Stencilalert(modelnum, RebY,'Status : NG')
                    if statusST == 'Inform to process':
                        modelnum = self.sheet.cell(row=i,column=2).value
                        self.songmail.Stencilalert(modelnum, RebY,'Status : Inform to process')

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to data: {e}")

    #progress
    def StartToolDeletEstencil(self,tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลบข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = stenCil.De_stenCil(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Tool manage', 'ลบข้อมูลสำเร็จ')
    
    def De_stenCil(self,tsid):
        route = pathlib.Path(self.stencilpath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.stencilsheet]
        self.sheet = self.wb[self.stencilsheet]
        rows = self.sheet.max_row

        try:
            os.remove(f"{self.stencilphotopath}\\{tsid}.png")
        except:
            ()
            
        for i in range(3,rows+1):
            if self.sheet.cell(row=i,column=1).value == tsid:
                try:
                    self.sheet.delete_rows(i,1)
                    self.wb.save(self.stencilpath)
            
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to data: {e}")
                








