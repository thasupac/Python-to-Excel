import os
import threading
from tkinter import messagebox, Toplevel, Label
from datetime import datetime
from PIL import Image
import configparser

#font
FONT = ('Angsana New',18)

class Machinedown:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.downtimepath = self.config['DATABASE']['dbdowntimepath']
        self.downtimesheet = self.config['DATABASE']['downtimesheet']
        self.downtimephoto = self.config['DATABASE']['dbdowntimephotopath']
    
    def startrecorddown(self,tsid,machine,inform,location,problems,photo):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังส่งข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = Machinedown.downtimerecord(self,tsid,machine,inform,location,problems,photo))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Machine downtime', 'ส่งข้อมูลสำเร็จ')

    def downtimerecord(self,tsid,machine,inform,location,problems,photo):
        from openpyxl import load_workbook
        self.excelfile = load_workbook(filename=self.downtimepath)
        self.excelfile.active = self.excelfile[self.downtimesheet]
        self.sheet = self.excelfile[self.downtimesheet]
        rows = self.sheet.max_row
        timestamp = datetime.now().strftime("%H:%M:%S : %d/%b/%Y")
        rows = self.sheet.max_row

        ###noti
        from send_mail import SendMail
        machinedown = SendMail()

        try:
            if photo !=(''):
                image = Image.open(photo)
                target_path = self.downtimephoto
                photo_name = f'{tsid}.png'
                photo_path = os.path.join(target_path, photo_name)
                image.save(photo_path)
                
            self.sheet.cell(row=rows+1,column=1).value=int(tsid)
            self.sheet.cell(row=rows+1,column=2).value=timestamp
            self.sheet.cell(row=rows+1,column=3).value=machine
            self.sheet.cell(row=rows+1,column=4).value=inform
            self.sheet.cell(row=rows+1,column=5).value=problems
            self.sheet.cell(row=rows+1,column=11).value=location
            self.sheet.cell(row=rows+1,column=12).value='open downtime'
            self.excelfile.save(self.downtimepath)

            ###email noti 
            machinedown.Startmachinedowntime(timestamp, machine, inform, problems, location)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    def startmaintenance(self,tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังส่งข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = Machinedown.maintenance(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Machine downtime', 'ส่งข้อมูลสำเร็จ')

    def maintenance(self,tsid):
        from openpyxl import load_workbook
        self.excelfile = load_workbook(filename=self.downtimepath)
        self.excelfile.active = self.excelfile[self.downtimesheet]
        self.sheet = self.excelfile[self.downtimesheet]
        rows = self.sheet.max_row
        timestamp = datetime.now().strftime("%H:%M:%S : %d/%b/%Y")
        rows = self.sheet.max_row
        for i in range(3, rows+1):
            if self.sheet.cell(row=i,column=1).value == tsid:
                self.sheet.cell(row=i,column=7,value=timestamp)
                self.excelfile.save(self.downtimepath)
    
    def startfinishedmaintenance(self, tsid, action_fix, action_by, notes, listRow):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังส่งข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = Machinedown.finisheddowntime(self, tsid, action_fix, action_by, notes, listRow))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Machine downtime', 'ส่งข้อมูลสำเร็จ')
    
    def finisheddowntime(self, tsid, action_fix, action_by, notes, listRow):
        from openpyxl import load_workbook
        self.excelfile = load_workbook(filename=self.downtimepath)
        self.excelfile.active = self.excelfile[self.downtimesheet]
        self.sheet = self.excelfile[self.downtimesheet]
        rows = self.sheet.max_row
        timestamp = datetime.now().strftime("%H:%M:%S : %d/%b/%Y")
        rows = self.sheet.max_row
        for i in range(3, rows+1):
            if self.sheet.cell(row=i,column=1).value == tsid:
                self.sheet.cell(row=i,column=6).value=action_fix
                self.sheet.cell(row=i,column=8).value=timestamp
                self.sheet.cell(row=i,column=9).value=action_by
                self.sheet.cell(row=i,column=10).value=notes
                self.sheet.cell(row=i,column=12).value='closed downtime'
                self.excelfile.save(self.downtimepath)

                ###noti
                from send_mail import SendMail
                machinedown = SendMail()

                ###email noti 
                machinedown.Startfinishedmachinedowntime(listRow[1], listRow[2], listRow[3], listRow[4], action_fix, action_by, notes, listRow[10])
                # listRow[1] = เวลาเกิด down
                # listRow[2] = เครื่องจักรที่พัง
                # listRow[3] = แจ้งเสียโดย
                # listRow[4] = ปัญหาที่เกิด
                # listRow[10] = เบย์ไหน
                # self, timestamp, machine, inform, problems, location

    def startdeletemachinedowntime(self, tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังส่งข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = Machinedown.deletedowntime(self, tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Machine downtime', 'ส่งข้อมูลสำเร็จ')

    def deletedowntime(self, tsid):
        from openpyxl import load_workbook
        self.excelfile = load_workbook(filename=self.downtimepath)
        self.excelfile.active = self.excelfile[self.downtimesheet]
        self.sheet = self.excelfile[self.downtimesheet]
        rows = self.sheet.max_row
        rows = self.sheet.max_row

        try:
            os.remove(f"{self.downtimephoto}\\{tsid}.png")
        except :
            ()

        for i in range(3, rows+1):
            if self.sheet.cell(row=i,column=1).value == tsid:
                self.sheet.delete_rows(i,1)
                self.excelfile.save(self.downtimepath)

