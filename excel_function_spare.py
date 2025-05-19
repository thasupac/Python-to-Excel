import os
import pathlib
import openpyxl as xl
import threading
from tkinter import messagebox, Toplevel, Label
from datetime import datetime
from PIL import Image
import configparser

#font
FONT = ('Angsana New',18)

##--Fixture
class SpareparT:
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.sparepath = self.config['DATABASE']['dbsparepath']
        self.sparesheet = self.config['DATABASE']['sparepartsheet']
        self.sparephoto = self.config['DATABASE']['dbphotosparepath']
        self.sparetrack = self.config['DATABASE']['tracksparesheet']

    ##--progress
    def StartpartReG(self,tsid,s_n,p_name,m_name,ora,qty,unit,line,des,photo_path_check):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลงทะเบียนข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = SpareparT.reg_spare(self,tsid,s_n,p_name,m_name,ora,qty,unit,line,des,photo_path_check))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Sparepart management', 'ลงทะเบียนข้อมูลสำเร็จ')                        

    #insert function
    def reg_spare(self,tsid,s_n,p_name,m_name,ora,qty,unit,line,des,photo_path_check):
        route = pathlib.Path(self.sparepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.sparesheet]
        self.sheet = self.wb[self.sparesheet]
        date = datetime.now().strftime("%d/%b/%Y")
        rows = self.sheet.max_row
        duplicate_check = False
        for i in range(2,rows+1):
            ##--ถ้า s/n ซ้ำกันให้บันทึกเฉพาะ qty และวันที่อัพเดต
            if self.sheet.cell(row=i,column=2).value == s_n:
                try:
                    self.sheet.cell(row=i,column=6).value += int(qty)
                    self.sheet.cell(row=i,column=9,value=date)
                    self.wb.save(self.sparepath)
                    duplicate_check = True
                    break
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to data: {e}")

        ##--ถ้า s/n ไม่ซ้ำกันให้ชบันทึกใหม่หมด
        while duplicate_check == False:
            try:
                try:
                    image = Image.open(photo_path_check)
                    target_path = self.sparephoto
                    photo_name = f'{tsid}.png'
                    photo_path = os.path.join(target_path, photo_name)
                    image.save(photo_path)
                except:
                    ()
                self.sheet.cell(row=rows+1,column=1,value=int(tsid))
                self.sheet.cell(row=rows+1,column=2,value=s_n)
                self.sheet.cell(row=rows+1,column=3,value=p_name)
                self.sheet.cell(row=rows+1,column=4,value=m_name)
                self.sheet.cell(row=rows+1,column=5,value=ora)
                self.sheet.cell(row=rows+1,column=6,value=int(qty))
                self.sheet.cell(row=rows+1,column=7,value=unit)
                self.sheet.cell(row=rows+1,column=8,value=line)
                self.sheet.cell(row=rows+1,column=9,value=date)
                self.sheet.cell(row=rows+1,column=10,value=des)
                self.wb.save(self.sparepath)
                break
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

    ##--progress
    def StartpartWithdraw(self,tsid,qty,req_by):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังบันทึกข้อมูลการใช้ Part...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = SpareparT.withdraw_part(self,tsid,qty,req_by))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Sparepart management', 'บันทึกข้อมูลการใช้ Part สำเร็จ')                        

    def withdraw_part(self,tsid,qty,req_by):
        route = pathlib.Path(self.sparepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.sparesheet]
        self.sheet = self.wb[self.sparesheet]
        date = datetime.now().strftime("%d/%b/%Y")
        rows = self.sheet.max_row
        ##
        try:
            for i in range(2, rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:  
                    self.sheet.cell(row=i,column=6).value-=int(qty)
                    self.sheet.cell(row=i,column=9).value=date
                    s_n = self.sheet.cell(row=i,column=2).value
                    p_name = self.sheet.cell(row=i,column=3).value
                    m_name = self.sheet.cell(row=i,column=4).value
                    ora = self.sheet.cell(row=i,column=5).value
                    unit = self.sheet.cell(row=i,column=7).value
                    self.wb.save(self.sparepath)
                    SpareparT.withdraw_log(self,tsid,s_n,p_name,m_name,ora,unit,qty,req_by)
                    break
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")
       
    def withdraw_log(self,tsid,s_n,p_name,m_name,ora,unit,qty,req_by):
        route = pathlib.Path(self.sparepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.sparetrack]
        self.sheet = self.wb[self.sparetrack]
        date = datetime.now().strftime("%d/%b/%Y")
        rows = self.sheet.max_row
        try:
            self.sheet.cell(row=rows+1,column=1,value=int(tsid))
            self.sheet.cell(row=rows+1,column=2,value=s_n)
            self.sheet.cell(row=rows+1,column=3,value=p_name)
            self.sheet.cell(row=rows+1,column=4,value=m_name)
            self.sheet.cell(row=rows+1,column=5,value=ora)
            self.sheet.cell(row=rows+1,column=6,value=int(qty))
            self.sheet.cell(row=rows+1,column=7,value=unit)
            self.sheet.cell(row=rows+1,column=8,value=req_by)
            self.sheet.cell(row=rows+1,column=9,value=date)
            self.wb.save(self.sparepath)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")

    def StartpartDeletE(self,tsid):
        def center_windows(w,h):
            ws = window.winfo_screenwidth() #screen width
            hs = window.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        
        window = Toplevel()
        win_size = center_windows(200,100)
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กำลังลบข้อมูล...', font=FONT)
        label.pack()
        window.update()
        thread = threading.Thread(target = SpareparT.delete_spare(self,tsid))
        thread.start()
        window.destroy()
        window.update()
        messagebox.showinfo('Sparepart management', 'ลบข้อมูลสำเร็จ')                        

    def delete_spare(self,tsid):
        route = pathlib.Path(self.sparepath)
        self.wb = xl.load_workbook(route)
        self.wb.active = self.wb[self.sparesheet]
        self.sheet = self.wb[self.sparesheet]
        rows = self.sheet.max_row
        try:
            try:
                os.remove(f"{self.sparephoto}\\{tsid}.png")
            except :
                ()
            for i in range(2,rows+1):
                if self.sheet.cell(row=i,column=1).value == tsid:
                    self.sheet.delete_rows(i,1)
                    self.wb.save(self.sparepath)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to data: {e}")





