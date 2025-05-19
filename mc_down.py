from tkinter import Frame, LabelFrame, Label, StringVar, ttk, Tk, Text, END, messagebox, filedialog, Toplevel, Menu, Scrollbar,RIGHT, Y
from datetime import datetime
from openpyxl import load_workbook
from PIL import ImageTk
import PIL.Image
import tkinter as tk
import configparser





#font size
#font size
FONT1 = ('Angsana New', 25, 'bold')
FONT2 = ('Angsana New', 18, )
FONT3 = ('Angsana New', 16, )
FONT4 = ('Angsana New', 12, )

techname = ['Thanongsak_su','Don P','Somchai L','Adirek C','Sangworn D',
            'Pratchaya S','Supot P','Kriangsak H','Anan C',
            'Thanatorn K','Anong J','Thanongsak D','Apicha K',
            'Sompong L', 'Kritchanaphong K','Sarawut N','Narong L',
            'Surasak N','Narong K','Other']

#machine_master
machine_master = ['Screen print','SPI','Datacon','Reflow','AOI','Underfill','Universal'
                  ,'Cleanning machine','Wafer cleaning','X-ray','Oven']

##line
line = ['BLD4#2', 'BLD5#10', 'BLD6#15/2']

##--rec part
class InformMachinedown(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        ###start down
        def startdowntime():
            from excel_function_mcdowntime import Machinedown
            exDown = Machinedown()
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            machine = v_machine_name.get()
            inform = v_inform_by.get()
            location = v_line.get()
            problems = problem.get('1.0',END).strip()
            photo = v_photo_path1.get()
            
            if machine and inform and location and problems != (''):
                exDown.startrecorddown(tsid,machine,inform,location,problems,photo)
                reset()
            else:
                messagebox.showinfo('Error', 'โปรดกรอกข้อมูลให้ครบ')
        def reset():
            v_machine_name.set('')
            v_inform_by.set('')
            v_line.set('')
            problem.delete('1.0', END)
            v_photo_path1.set('')
            photo_mc.config(image=None)
            photo_mc.image = None

        ##selectphoto
        def selectphoto():
            file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            v_photo_path1.set(file_path)
            try:
                image = PIL.Image.open(v_photo_path1.get())
                image = image.resize((190,190))
                photo = ImageTk.PhotoImage(image)
                photo_mc.config(image=photo)
                photo_mc.image = photo
            except Exception as e:
                messagebox.showerror('Register photo', f'Failed to load photo {e}')

        ###---production frame
        pro = LabelFrame(self, text='production information', font=FONT3)
        pro.grid(row=0, column=0, padx=10, pady=10)

        ###mc_name_down
        L = Label(pro, text='Machine name/No. :', font=FONT2)
        L.grid(row=0, column=0, padx=10)
        v_machine_name = StringVar()
        E = ttk.Combobox(pro, textvariable=v_machine_name, font=FONT2, values=machine_master, width=18)
        E.grid(row=0, column=1, padx=10, pady=2)

        ###production
        L = Label(pro, text='Inform by : ', font=FONT2)
        L.grid(row=1, column=0, padx=10, pady=2)
        v_inform_by = StringVar()
        E = ttk.Entry(pro, textvariable=v_inform_by, font=FONT2)
        E.grid(row=1, column=1, padx=10, pady=2)

        ###line
        L = Label(pro, text='Line :', font=FONT2)
        L.grid(row=2, column=0, padx=10, pady=2)
        v_line = StringVar()
        E = ttk.Combobox(pro, textvariable=v_line, values=line, font=FONT2, state='readonly', width=18)
        E.grid(row=2, column=1, padx=10, pady=2)

        ###problem
        L = Label(pro, text='Problems :', font=FONT2)
        L.grid(row=3, column=0, padx=10, pady=2)
        problem = Text(pro, font=FONT2, width=40, height=4)
        problem.grid(row=3, column=1, padx=10, pady=2)

        ###photo
        photoframe = LabelFrame(pro)
        photoframe.grid(row=4, column=0, padx=10, pady=10)
        
        ###path photo
        v_photo_path1 = StringVar()
        E = ttk.Entry(photoframe, textvariable=v_photo_path1, font=FONT3)
        E.pack(padx=10, pady=10)

        ###select photo
        B = ttk.Button(photoframe, text='select photo', command=selectphoto)
        B.pack(padx=10, pady=10)

        ###photo show
        photoframe2 = LabelFrame(pro, width=150, height=150)
        photoframe2.grid(row=4, column=1, padx=10, pady=10)
        photo_mc = Label(photoframe2)
        photo_mc.pack()


        ###button
        B = ttk.Button(pro, text='Start machine down', command=startdowntime)
        B.grid(row=5, column=1, ipadx=5, ipady=5, pady=10)

class ActionMachinedown(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.downtimepath = self.config['DATABASE']['dbdowntimepath']
        self.downtimesheet = self.config['DATABASE']['downtimesheet']
        self.downtimephoto = self.config['DATABASE']['dbdowntimephotopath']

        #search_type
        search_type = Label(self, text='Search type :', font=FONT2)
        search_type.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('All')
        search_type_combo = ttk.Combobox(self , textvariable=v_search_type, values=['Machine/#ORA','Line','Date','All'], state='readonly')
        search_type_combo.place(x=120, y=10, height=30)
        
        #function search
        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()     
            
            #mc search
            if check_search_type == 'Machine/#ORA':
                    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=12, values_only=True):
                        found = False
                        for col_idx in columns:
                            if col_idx <= len(row):
                                cell_value = str(row[2]).strip()  #Machine/#ORA
                                if cell_value.lower().startswith(search_prefix.lower()):
                                    found = True
                                    break
                        if found:
                            results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), str(row[4]).strip(), str(row[10]).strip(), str(row[11]).strip()])
                    return results
                
            #ora search
            elif check_search_type == 'Line':
                for row in sheet.iter_rows(min_row=3, max_row=None, min_col=1, max_col=12, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[10]).strip() #Line
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), str(row[4]).strip(), str(row[10]).strip(), str(row[11]).strip()])
                return results
            
            #line search
            elif check_search_type == 'Date':
                for row in sheet.iter_rows(min_row=3, max_row=None, min_col=1, max_col=12, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            cell_value = str(row[1]).strip().split(':')[-1]  #ORA
                            if cell_value.lower().startswith(search_prefix.lower()):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), str(row[4]).strip(), str(row[10]).strip(), str(row[11]).strip()])
                return results
            
            #all search
            elif check_search_type == 'All':
                
                for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=12, values_only=True):
                    results.append([str(row[0]).strip(),str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip(), str(row[4]).strip(), str(row[10]).strip(), str(row[11]).strip()])
                return results
                
        def on_key_release(event):
            search_prefix = EFID.get().strip().lower()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [1], search_prefix)  # Adjust columns as needed
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def display_results(results):
            for iter in downtime_list.get_children():
                downtime_list.delete(iter)

            if results:
                for row in results:
                    #row = tuple(row)
                    downtime_list['columns']=("TSID","Timestamp", "Machine name/No." , 'Inform by',"Problem",'Line','Status')
                    downtime_list.column('TSID', anchor="center", width=2)
                    downtime_list.column('Timestamp', anchor="center", width=2)
                    downtime_list.column('Machine name/No.', anchor="center", width=2)
                    downtime_list.column('Inform by', anchor="center", width=2)  
                    downtime_list.column('Problem', anchor="center", width=15) 
                    downtime_list.column('Line', anchor="center", width=2)
                    downtime_list.column('Status', anchor="center", width=2) 

                    ##heading
                    downtime_list.heading('TSID', text='TSID',anchor="center")
                    downtime_list.heading('Timestamp', text='Timestamp', anchor="center")
                    downtime_list.heading('Machine name/No.', text='Machine name/No.',anchor="center")
                    downtime_list.heading('Inform by', text='Inform by', anchor="center")   
                    downtime_list.heading('Problem', text='Problem', anchor="center")  
                    downtime_list.heading('Line', text='Line', anchor="center")
                    downtime_list.heading('Status', text='Status', anchor="center")  
                    downtime_list.insert('', 'end', values=row)

            else:
                downtime_list['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                downtime_list.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                downtime_list.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")
       
            ###get data from tsid
            ##########################################
            ##########################################
            ##########################################
            def getdata(tsid):
                wb = load_workbook(filename=file_path)
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=12, values_only=True):
                    if tsid ==  int(row[0]):  #TSID
                        listRow = list(row)
                        displaypopup(listRow)
                        
            #Select data to edit
            def Select_mc(event=None):
                select = downtime_list.selection()
                output = downtime_list.item(select)
                if select !=():
                    tsid = output['values'][0]
                    getdata(tsid)
            downtime_list.bind('<Double-1>', Select_mc)

            ### pop display
            def displaypopup(listRow):
                tsid = listRow[0]
                timestamp = listRow[1]
                mc_no = listRow[2]
                inform_by = listRow[3]
                problems = listRow[4]
                action = listRow[5]
                startfix = listRow[6]
                finishfix = listRow[7]
                fix_by = listRow[8]
                note = listRow[9]
                lines = listRow[10]
                status = listRow[11]
                 
                GUIDetail = Toplevel()
                GUIDetail.title('Downtimes')
                GUIDetail.geometry('800x600')

                def startmaintenance():
                    from excel_function_mcdowntime import Machinedown
                    exSM = Machinedown()
                    exSM.startmaintenance(tsid)
                    displayclose() 

                def finishedmaintenance():
                    from excel_function_mcdowntime import Machinedown
                    exSM = Machinedown()
                    action_fix = actions.get('1.0', END)
                    action_by = v_action_by.get()
                    notes = remark.get('1.0', END)
                    try:
                        if action_fix and action_by and notes:
                            exSM.startfinishedmaintenance(tsid, action_fix, action_by, notes, listRow)
                            displayclose() 
                        else:
                            messagebox.showinfo('Machine downtime', 'โปรดกรอกข้อมูลให้ครบ')  
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to data: {e}")

                def displayclose(event=None):
                    GUIDetail.destroy()
                    GUIDetail.update()
                    display_results(results)
                    on_key_release(event)

                ###---production frame
                pro = LabelFrame(GUIDetail, text='production information', font=FONT3)
                pro.grid(row=1, column=0, padx=10, pady=10)

                L = Label(pro, text='Start machine downtime :', font=FONT2)
                L.grid(row=0, column=0, pady=5)
                v_timestamp = StringVar()
                v_timestamp.set(timestamp)
                E = ttk.Entry(pro, textvariable=v_timestamp, font=FONT2)
                E.grid(row=0, column=1)

                L = Label(pro, text='Status downtime :', font=FONT2)
                L.grid(row=1, column=0, pady=5)
                v_status = StringVar()
                v_status.set(status)
                E = ttk.Entry(pro, textvariable=v_status, font=FONT2)
                E.grid(row=1, column=1)

                ###mc_name_down
                L = Label(pro, text='Machine name/No. :', font=FONT2)
                L.grid(row=2, column=0, padx=10)
                v_machine_name = StringVar()
                v_machine_name.set(mc_no)
                E = ttk.Entry(pro, textvariable=v_machine_name, font=FONT2)
                E.grid(row=2, column=1, padx=10, pady=2)

                ###production
                L = Label(pro, text='Inform by : ', font=FONT2)
                L.grid(row=3, column=0, padx=10, pady=2)
                v_inform_by = StringVar()
                v_inform_by.set(inform_by)
                E = ttk.Entry(pro, textvariable=v_inform_by, font=FONT2)
                E.grid(row=3, column=1, padx=10, pady=2)

                ###line
                L = Label(pro, text='Line :', font=FONT2)
                L.grid(row=4, column=0, padx=10, pady=2)
                v_line = StringVar()
                v_line.set(lines)
                E = ttk.Entry(pro, textvariable=v_line, font=FONT2)
                E.grid(row=4, column=1, padx=10, pady=2)

                ###problem
                L = Label(pro, text='Problems :', font=FONT2)
                L.grid(row=5, column=0, padx=10, pady=2)
                problem = Text(pro, font=FONT2, width=40, height=4)
                problem.grid(row=5, column=1, padx=10, pady=2)
                problem.insert('1.0', str(problems), END )

                ###photo show
                photoframe2 = LabelFrame(pro, width=150, height=150)
                photoframe2.grid(row=6, column=1, padx=10, pady=10)
                photo_mc = Label(photoframe2)
                photo_mc.pack()
                try:
                    image = PIL.Image.open(f"{self.downtimephoto}\\{tsid}.png")
                    image = image.resize((190, 140))
                    photo = ImageTk.PhotoImage(image)
                    photo_mc.config(image=photo)
                    photo_mc.image = photo

                except Exception as e:
                    photo_mc.config(text='ไม่มีรูปภาพของ TSID นี้')
                
                ##start maintenance
                Bst = ttk.Button(pro, text='Start maintenance', command=startmaintenance)
                Bst.grid(row=7, column=1, ipadx=5, ipady=5, pady=10)

                while True:
                    if str(startfix) == 'None': 
                        Bst.configure(state='normal')
                        break

                    Bst.configure(state='disable')
                    tech = LabelFrame(GUIDetail, text='technician actions', font=FONT3)
                    tech.grid(row=1, column=1, padx=10, pady=10)

                    ###action
                    L = Label(tech, text='Actions :', font=FONT2)
                    L.grid(row=0, column=0, padx=10, pady=10)
                    actions = Text(tech, width=40, height=4, font=FONT2)
                    actions.grid(row=0, column=1, padx=10, pady=10)
                    if str(action) != 'None':
                        actions.insert('1.0', str(action))

                    ###actions by
                    L = Label(tech, text='Action by :', font=FONT2)
                    L.grid(row=1, column=0, padx=10, pady=10)
                    v_action_by = StringVar()
                    if str(fix_by) != 'None':
                        v_action_by.set(fix_by)
                    E = ttk.Combobox(tech, textvariable=v_action_by, font=FONT2, values=techname)
                    E.grid(row=1, column=1, padx=10, pady=10)

                    ###remark
                    L = Label(tech, text='Remarks :', font=FONT2)
                    L.grid(row=2, column=0)
                    remark = Text(tech, width=40, height=3)
                    remark.grid(row=2, column=1, padx=10, pady=10)
                    if str(note) != 'None':
                        remark.insert('1.0', str(note))

                    ###photo
                    photoframe = LabelFrame(tech)
                    photoframe.grid(row=3, column=0, padx=10, pady=10)

                    ###path photo
                    v_photo_path2 = StringVar()
                    E = ttk.Entry(photoframe, textvariable=v_photo_path2, font=FONT3)
                    E.pack(padx=10, pady=10)

                    ###select photo
                    B = ttk.Button(photoframe, text='select photo')
                    B.pack(padx=10, pady=10)

                    ###photo show
                    photoframe2 = LabelFrame(tech, width=150, height=150)
                    photoframe2.grid(row=3, column=1, padx=10, pady=10)

                    ###start stop
                    v_start_fix = StringVar()
                    v_start_fix.set(f'Start maintenance : {startfix}')
                    E = ttk.Entry(tech, textvariable=v_start_fix, font=FONT3, width=40)
                    E.grid(row=4, column=0, pady=5, padx=5)

                    v_stopfix = StringVar()
                    v_stopfix.set(f'Finished maintenance : {finishfix}')
                    E = ttk.Entry(tech, textvariable=v_stopfix, font=FONT3, width=40)
                    E.grid(row=4, column=1, pady=5, padx=5)

                    B = ttk.Button(tech, text='Finished maintenance', command=finishedmaintenance)
                    B.grid(row=6, column=1, ipadx=5, ipady=5, pady=10)

                    break
                        
                GUIDetail.mainloop()



            ##########################################
            ##########################################
            ##########################################

            #right click to delete and update
            def delete_mc(event=None):
                select = downtime_list.selection()
                output = downtime_list.item(select)
                check = messagebox.askyesno('Equipment','Do you want to delete equipment?')
                if check == True:
                    if output['values'] != (''):
                        tsid = output['values'][0]
                        try:
                            from excel_function_mcdowntime import Machinedown
                            Machinedown.startdeletemachinedowntime(self,tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")
                
            delete_mc_menu = Menu(self, tearoff=0)
            delete_mc_menu.add_command(label='delete', command=delete_mc)

            #select delete
            def popup(event):
                delete_mc_menu.post(event.x_root, event.y_root)
            downtime_list.bind('<Button-3>', popup)   #Button 3 คือ click ขวา

        def clear_results():
            for iter in downtime_list.get_children():
                downtime_list.delete(iter)
                        

        # Define Excel data for searching
        file_path = self.downtimepath
        sheet_name = self.downtimesheet

        # #--label 
        # #--entry fixture ID
        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        EFID = ttk.Entry(self, font=FONT2)
        EFID.place(x=120, y=50, width=150, height=30)
        EFID.bind('<KeyRelease>', on_key_release)

        #create list machine
        header = ["TSID","Timestamp", "Machine name/No." , 'Inform by',"Problem",'Line','Status']
        headerw = [50,50,50,50,200,50,50]
        downtime_list = ttk.Treeview(self, columns=header, show='headings')
        downtime_list.place(x=20, y=90, width=1200, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',14))

        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")

        #--header and width
        for h,w in zip(header,headerw):
            downtime_list.heading(h,text=h)
            downtime_list.column(h,width=w,anchor='center')

        #scroll bar vertical
        downtime_list_Scroll = Scrollbar(downtime_list)
        downtime_list_Scroll.pack(side=RIGHT, fill=Y)
        downtime_list_Scroll.config(command=downtime_list.yview)

# from tkinter import *
# gui = Tk()
# gui.title('Downtime')
# gui.geometry('1500x1500')
# # a = InformMachinedown(gui)
# a = ActionMachinedown(gui)
# # a = bordReturn(gui)
# # a = bordwith(gui)
# a.pack()
# gui.mainloop()