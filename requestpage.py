from tkinter import ttk,messagebox,filedialog,END,LabelFrame,Label,StringVar,IntVar,Text,Toplevel,Frame,Menu,Scrollbar,RIGHT,Y
from PIL import ImageTk
from send_mail import SendMail
from excel_funcrequest import RequestSpare
from datetime import datetime
from openpyxl import load_workbook
import subprocess
import tkinter as tk
import PIL.Image
import threading
import configparser

Mail = SendMail()
excelRQ = RequestSpare()

#font size
FONT1 = ('Angsana New', 25, 'bold')
FONT2 = ('Angsana New', 18, )
FONT3 = ('Angsana New', 16, )
FONT4 = ('Angsana New', 12, )

#techname
techname = ['Thanongsak Su','Khunakorn R','Pichet T','Wasan R',
            'Don P','Somchai L','Adirek C','Sangworn D',
            'Pratchaya S','Supot P','Kriangsak H','Anan C',
            'Thanatorn K','Anong J','Thanongsak D','Apicha K','Sompong L']

line = ['BLD4#2', 'BLD5#10', 'BLD6#15/2']
unit = ['กรัม','กิโลกรัม','ชิ้น','คู่', 'ชุด','เมตร','ลิตร','ตารางเมตร','เซนติเมตร','นิ้ว','หลา','ปอนด์']

class Request(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        #request part
        def RequestSend():
            tsid = str(int(datetime.now().strftime('%y%m%d%H%M%S')) + 524008)
            serialnum = v_sn.get()
            partname = v_pName.get()
            machine = v_Mcname.get()
            line = v_line.get()
            unit = v_unit.get()
            quantity = v_quantity.get()
            requestby = v_reQby.get()
            comment = Edesc.get('1.0', END)
            attachfile = v_part_path.get()
            if serialnum != ('') and partname != ('') and machine != ('') and line != ('') and unit != ('') and quantity != ('') and requestby != (''):
                Mail.Sparesongmail(serialnum, partname, machine, line, quantity, unit, requestby, comment, attachfile)
                excelRQ.StartRequest(tsid, serialnum, partname, machine, line, quantity, unit, requestby, comment)
                Reset()
            else:
                messagebox.showinfo('Request spare part', 'Please fill data!')

        #reset
        def Reset():
            v_sn.set('')
            v_pName.set('')
            v_Mcname.set('')
            v_line.set('')
            v_unit.set('')
            v_quantity.set('')
            v_reQby.set('')
            Edesc.delete('1.0', END)
            v_part_path.set('')
            RemoveFile()

        ##--function selectFile
        def selectFile():
            filetypes = (('All files', '*.*'),
                         ("Excel file","*.xlsx;*.xls;*.xlsm"),
                         ("Image files", "*.jpg;*.jpeg;*.png"),
                         ('Powerpoint file','*.pptx'),
                         ('PDF file', '*.pdf'),
                         ('text files', '*.txt'))
            file_path = filedialog.askopenfilename(filetypes=filetypes)
            v_part_path.set(file_path)
            if v_part_path.get():
                if v_part_path.get().split('.')[-1] == 'jpg' and 'jpeg' and ' png':
                    try:
                        image = PIL.Image.open(v_part_path.get())
                        image = image.resize((300,240))
                        photo = ImageTk.PhotoImage(image)
                        RequestPart.config(image=photo)
                        RequestPart.image = photo
                        
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to data: {e}")

        def RemoveFile():
            RequestPart.config(image=None)
            RequestPart.image = None

        #main frame
        RQ = LabelFrame(self, width=1200, height=700)
        RQ.pack(padx=10, pady=10)

        #general detail
        GD = LabelFrame(RQ, text='general detail', width=400, height=500)
        GD.grid(row=0, column=0, padx=10, pady=10)

        #serial number
        L = Label(GD, text='Serial number :', font=FONT2)
        L.grid(row=0, column=0, padx=30, pady=10, sticky='w')
        v_sn = StringVar()
        E = ttk.Entry(GD, textvariable=v_sn, font=FONT2)
        E.grid(row=0, column=1, padx=20, pady=10)

        #partname
        L = Label(GD, text='Part name :', font=FONT2)
        L.grid(row=1, column=0, padx=30, pady=10, sticky='w')
        v_pName = StringVar()
        E = ttk.Entry(GD, textvariable=v_pName, font=FONT2)
        E.grid(row=1, column=1, padx=20, pady=10)

        #machine name
        L = Label(GD, text='Machine name :', font=FONT2)
        L.grid(row=2, column=0, padx=30, pady=10, sticky='w')
        v_Mcname = StringVar()
        E = ttk.Entry(GD, textvariable=v_Mcname, font=FONT2)
        E.grid(row=2, column=1, padx=20, pady=10)

        #line
        L = Label(GD, text='Line :', font=FONT2)
        L.grid(row=3, column=0, padx=30, pady=10, sticky='w')
        v_line = StringVar()
        E = ttk.Combobox(GD, textvariable=v_line, font=FONT2, values=line, state='readonly', width=18)
        E.grid(row=3, column=1, padx=20, pady=10)

        #QUANTITY
        L = Label(GD, text='Quantity :', font=FONT2)
        L.grid(row=4, column=0, padx=30, pady=10, sticky='w')
        v_quantity = IntVar()
        E = ttk.Entry(GD, textvariable=v_quantity, font=FONT2)
        E.grid(row=4, column=1, padx=20, pady=10)

        #UNIT
        L = Label(GD, text='Unit :', font=FONT2)
        L.grid(row=5, column=0, padx=30, pady=10, sticky='w')
        v_unit = StringVar()
        E = ttk.Combobox(GD, textvariable=v_unit, values=unit, font=FONT2, width=18, state='readonly')
        E.grid(row=5, column=1, padx=20, pady=10)

        #request by
        L = Label(GD, text='Request by :', font=FONT2)
        L.grid(row=6, column=0, padx=30, pady=10, sticky='w')
        v_reQby = StringVar()
        E = ttk.Combobox(GD, textvariable=v_reQby, font=FONT2, values=techname, width=18, state='readonly')
        E.grid(row=6, column=1, padx=20, pady=10)

        #other detail
        OD = LabelFrame(RQ, text='other detail', width=400, height=500)
        OD.grid(row=0, column=1, padx=10, pady=10)

        #deascqriptions
        L = Label(OD, text='Descriptions :', font=FONT2)
        L.grid(row=0, column=0, padx=30, pady=10)
        Edesc = Text(OD, width=20, height=5)
        Edesc.grid(row=0, column=1, padx=10, pady=10)

        #photo
        SP = LabelFrame(OD, width=100, height=100)
        SP.grid(row=1, column=0, padx=10, pady=10)

        #pathphoto
        v_part_path = StringVar()
        E = ttk.Entry(SP, textvariable=v_part_path, font=FONT2)
        E.grid(row=1, column=0, padx=30, pady=10)
        B = ttk.Button(SP, text='Select', command=selectFile)
        B.grid(row=2, column=0, padx=20, pady=10)

        #show photo
        PS = LabelFrame(OD, width=200, height=230)
        PS.grid(row=1, column=1, padx=10, pady=10)

        #labelphoto
        RequestPart = tk.Label(PS)
        RequestPart.pack(padx=20, pady=20)

        #request
        B = ttk.Button(self, text='Request', command=RequestSend)
        B.pack(padx=2.5, pady=2.5)

        #clear
        B = ttk.Button(self, text='Clear', command=Reset)
        B.pack()

        #back to home
    ##--progress
    def StartmCSwitchPagE(self):
        def center_windows(self,w,h):
            ws = self.winfo_screenwidth() #screen width
            hs = self.winfo_screenheight() #screen height
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            return f'{w}x{h}+{x:.0f}+{y:.0f}'
        win_size = center_windows(self,200,100)
        window = Toplevel()
        window.geometry(win_size)
        window.title('Progress functions')
        label = Label(window, text = 'กรุณารอสักครู่...', font=FONT1)
        label.pack()
        window.update()
        thread = threading.Thread(target = Request.Switch(self))
        thread.start()
        window.destroy()
        window.update()
 
    def Switch(self):
        subprocess.Popen(["python", "ASMT_Store_Part.py"])
        self.quit()

class Requestview(Frame):
    def __init__ (self, GUI):
        Frame.__init__(self, GUI, width=1500, height=1500)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')

        self.requestpath = self.config['DATABASE']['dbrequestpath']
        self.requestsheet = self.config['DATABASE']['requestlogsheet']
        self.requestphotopath = self.config['DATABASE']['dbphotosparepath']

        #function search
        def search_excel(file_path, sheet_name, columns, search_prefix):
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            results = []
            check_search_type = v_search_type.get()

            ##S/N search type:
            if check_search_type == ('Serial number'):
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[1]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip()])
                return results
            
            #Part name
            elif check_search_type == ('Part name'):    
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[col_idx]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip()])
                return results
            
            #Machine search
            elif check_search_type == ('Machine name'):  
                for row in sheet.iter_rows(min_row=2, max_row=None, min_col=0, max_col=11, values_only=True):
                    found = False
                    for col_idx in columns:
                        if col_idx <= len(row):
                            # if 
                            cell_value = str(row[col_idx+1]).strip()  # ตำแหน่ง column ที่จะหา
                            if cell_value.lower().startswith(search_prefix):
                                found = True
                                break
                    if found:
                        results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), str(row[3]).strip(), 
                                        str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                        str(row[8]).strip(),str(row[9]).strip()])
                return results
            
            #all
            elif check_search_type == ('All'):
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=0, max_col=11, values_only=True):
                    results.append([str(row[0]).strip(),str(row[1]).strip(),str(row[2]).strip(), str(row[3]).strip(), 
                                    str(row[4]).strip(),str(row[5]).strip(),str(row[6]).strip(),str(row[7]).strip(),
                                    str(row[8]).strip(),str(row[9]).strip()])
                return results

        def on_key_release(event):
            search_prefix = EFID.get().strip().lower()
            try:
                if search_prefix:
                    results = search_excel(file_path, sheet_name, [1], search_prefix)  # Adjust columns as needed
                    display_results(results)
                else:
                    clear_results()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to data: {e}")

        def display_results(results):
            for iter in RQlist.get_children():
                RQlist.delete(iter)
            if results:
                for row in results:
                    #row = tuple(row)
                    RQlist['columns']=("TSID","Serial number","Part name" ,'Machine name',"Line","Quantity","Unit","Date","Request by","Comment")
                    RQlist.column('TSID', anchor="center", width=5)
                    RQlist.column('Serial number', anchor="center", width=5)
                    RQlist.column('Part name', anchor="center", width=5)
                    RQlist.column('Machine name', anchor="center", width=5)  
                    RQlist.column('Line', anchor="center", width=5) 
                    RQlist.column('Quantity', anchor="center", width=5) 
                    RQlist.column('Unit', anchor="center", width=5) 
                    RQlist.column('Date', anchor="center", width=5) 
                    RQlist.column('Request by', anchor="center", width=5) 
                    RQlist.column('Comment', anchor="center", width=5) 

                    ##heading
                    RQlist.heading('TSID', text='TSID',anchor="center")
                    RQlist.heading('Serial number', text='Serial number', anchor="center")
                    RQlist.heading('Part name', text='Part name',anchor="center")
                    RQlist.heading('Machine name', text='Machine name', anchor="center")   
                    RQlist.heading('Line', text='Line', anchor="center")  
                    RQlist.heading('Quantity', text='Quantity', anchor="center")  
                    RQlist.heading('Unit', text='Unit', anchor="center")
                    RQlist.heading('Date', text='Date', anchor="center")
                    RQlist.heading('Request by', text='Request by', anchor="center")  
                    RQlist.heading('Comment', text='Comment', anchor="center")  
                    RQlist.insert('', 'end', values=row)

            else:
                RQlist['columns']=("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา")
                RQlist.column("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center", width=500)
                RQlist.heading("ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", text="ไม่พบผลลัพธ์ที่ตรงกับคำค้นหา", anchor="center")

            #select
            def Select_mc(event=None):
                select = RQlist.selection()
                output = RQlist.item(select)
                if select !=():
                    tsid = output['values'][0]
                    wb = load_workbook(filename=file_path)
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=10, values_only=True):
                        if tsid ==  int(row[0]):  #TSID
                            listRow = list(row)
                            sn = listRow[1]
                            pn = listRow[2]
                            mc = listRow[3]
                            bld = listRow[4]
                            qty = listRow[5]
                            unit = listRow[6]
                            daterq = listRow[7]
                            reqby = listRow[8]
                            comment = listRow[9]

                            #Gui for detail
                            RQP = Toplevel()
                            RQP.title('Withdraw')
                            RQP.geometry('600x600-50+30')

                            MF = LabelFrame(RQP)
                            MF.pack(padx=10, pady=10)

                            #general frame
                            GD = LabelFrame(MF, text='general detail', font=FONT4, width=500, height=500)
                            GD.grid(row=0, column=0, padx=10, pady=10)

                            #tsid
                            L = Label(GD, text=f' TSID : {tsid} ', font=FONT4)
                            L.grid(row=0, column=0)

                            #serial
                            L = Label(GD, text='Serial number :', font=FONT2)
                            L.grid(row=1, column=0, padx=20, pady=5, sticky='w')
                            v_serial = StringVar()
                            v_serial.set(sn)
                            E = ttk.Entry(GD, textvariable=v_serial, font=FONT2, state='readonly')
                            E.grid(row=1, column=1, padx=20, pady=5)

                            #part name
                            L = Label(GD, text='Part name :', font=FONT2)
                            L.grid(row=2, column=0 ,padx=20, pady=5, sticky='w')
                            v_partname = StringVar()
                            v_partname.set(pn)
                            E = ttk.Entry(GD, textvariable=v_partname, font=FONT2, state='readonly')
                            E.grid(row=2, column=1, padx=20, pady=5)

                            #machine name
                            L = Label(GD, text='Machine name :', font=FONT2)
                            L.grid(row=3, column=0, padx=20, pady=5, sticky='w')
                            v_machine = StringVar()
                            v_machine.set(mc)
                            E = ttk.Entry(GD, textvariable=v_machine, font=FONT2, state='readonly')
                            E.grid(row=3, column=1, padx=20, pady=5, sticky='w')

                            #line
                            L = Label(GD, text='Line :', font=FONT2)
                            L.grid(row=4, column=0, padx=20, pady=5, sticky='w')
                            v_line = StringVar()
                            v_line.set(bld)
                            E = ttk.Entry(GD, textvariable=v_line, font=FONT2, state='readonly')
                            E.grid(row=4, column=1, padx=20, pady=5)

                            #quantity
                            L = Label(GD, text='Quantity :', font=FONT2)
                            L.grid(row=5, column=0, padx=20, pady=5, sticky='w')
                            v_qty = StringVar()
                            v_qty.set(qty)
                            E = ttk.Entry(GD, textvariable=v_qty, font=FONT2, state='readonly')
                            E.grid(row=5, column=1, padx=20, pady=5)

                            #unit
                            L = Label(GD, text='Unit :', font=FONT2)
                            L.grid(row=6, column=0, padx=20, pady=5, sticky='w')
                            v_unit = StringVar()
                            v_unit.set(unit)
                            E = ttk.Entry(GD, textvariable=v_unit, font=FONT2, state='readonly')
                            E.grid(row=6, column=1, padx=20, pady=5)

                            #rqdate
                            L = Label(GD, text='Date request :', font=FONT2)
                            L.grid(row=7, column=0, padx=20, pady=5, sticky='w')
                            v_date = StringVar()
                            v_date.set(daterq)
                            E = ttk.Entry(GD, textvariable=v_date, font=FONT2, state='readonly')
                            E.grid(row=7, column=1, padx=20, pady=5)

                            #requester
                            L = Label(GD, text='Requester :', font=FONT2)
                            L.grid(row=8, column=0, padx=20, pady=5, sticky='w')
                            v_requester = StringVar()
                            v_requester.set(reqby)
                            E = ttk.Entry(GD, textvariable=v_requester, font=FONT2, state='readonly')
                            E.grid(row=8, column=1, padx=20, pady=5)

                            L = Label(GD, text='Descriptions :', font=FONT2)
                            L.grid(row=9, column=0, padx=20, pady=5)
                            Edesc = Text(GD, width=20, height=5)
                            Edesc.grid(row=9, column=1, padx=10, pady=5)
                            Edesc.insert('1.0', str(comment))

                            #other frame
                            OD = LabelFrame(MF, text='other detail', font=FONT4, width=500, height=500)
                            OD.grid(row=0, column=1, padx=10, pady=10)
                            PhotoShow = Label(OD)
                            PhotoShow.pack(padx=20, pady=20)
                            #แสดงรูปภาพ
                            try:
                                photo_part = Label(PhotoShow)
                                photo_part.pack()
                                image = PIL.Image.open(f"{self.requestphotopath}\\{tsid}.png")
                                image = image.resize((500, 405))
                                photo = ImageTk.PhotoImage(image)
                                photo_part.config(image=photo)
                                photo_part.image = photo
                            except Exception as e:
                                L = Label(PhotoShow, text='ไม่มีรูปภาพสำหรับ TSID นี้')
                                L.pack()
                                ()

                            RQP.mainloop()

            RQlist.bind('<Double-1>', Select_mc)

            #delete functions
            def delete_right_click(event=None):
                select = RQlist.selection()
                output = RQlist.item(select)
                check = messagebox.askyesno('ยืนยันการลบ','คุณต้องการลบข้อมูล?...')
                if check == True:
                    if output['values'] != (''):
                        tsid = output['values'][0]
                        try:
                            excelRQ.StartDeletelog(tsid)
                            display_results(results)
                            on_key_release(event)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to data: {e}")

            #right click
            #select delete
            def popup(event):
                pop_up_menu.post(event.x_root, event.y_root)
            RQlist.bind('<Button-3>', popup)   #Button 3 คือ click ขวา
            pop_up_menu = Menu(self, tearoff=0)
            pop_up_menu.add_command(label='delete', command=delete_right_click)

        def clear_results():
            for iter in RQlist.get_children():
                RQlist.delete(iter)

        #defind path
        file_path = self.requestpath
        sheet_name = self.requestsheet

        # #--entry fixture ID
        L = Label(self, text='Search type :', font=FONT2)
        L.place(x=20, y=10)
        v_search_type = StringVar()
        v_search_type.set('Serial number')
        E = ttk.Combobox(self , textvariable=v_search_type, values=['Serial number','Part name','Machine name','All'], state='readonly')
        E.place(x=120, y=10, height=30)

        #entry search
        L = Label(self, text='Search :', font=FONT2)
        L.place(x=20, y=50, height=30)
        EFID = ttk.Entry(self, font=FONT2)
        EFID.place(x=120, y=50, width=150, height=30)
        EFID.bind('<KeyRelease>', on_key_release)

        #create list machine
        header = ["TSID","Serial number", "Part name" , 'Machine name',"Line", "Quantity", "Unit", "Date", "Request by", "Comment"]
        headerw = [100,100,100,100,100,100,100,100,100,100]
        RQlist = ttk.Treeview(self, columns=header, show='headings')
        RQlist.place(x=20, y=90, width=1200, height=450)

        #style
        style = ttk.Style()
        style.configure('Treeview.Heading',font=('Angsana New',14,'bold'))
        style.configure('Treeview',rowheight=20,font=('Angsana New',14))

        #config treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", background="yellow", foreground="black")

        #--header and width
        for h,w in zip(header,headerw):
            RQlist.heading(h,text=h)
            RQlist.column(h,width=w,anchor='center')

        #scroll bar vertical
        RQlistScorll = Scrollbar(RQlist)
        RQlistScorll.pack(side=RIGHT, fill=Y)
        RQlistScorll.config(command=RQlist.yview) 

#เรียกมาใช้ก่อนน
# root = Tk()
# root.title('Request')
# root.geometry('1500x1500')
# app = Requestview(root)
# app.pack()
# root.mainloop()